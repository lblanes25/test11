"""
Risk Taxonomy Transformer
=========================
Transforms legacy 14-pillar risk taxonomy into new 6 L1 / 23 L2 taxonomy.
Handles: LLM override file, sub-risk description lookup, deterministic mapping,
1:many keyword resolution, overlay/amplifier risks, and rating decomposition.

Resolution order for multi mappings:
  Override — LLM-classified overrides from Review Queue (highest priority)
  Evidence — Keyword matching on rationale text + sub-risk descriptions
  Default  — First primary L2, flagged for review

Workflow:
  1. Run script without overrides -> produces Review Queue
  2. Batch Review Queue through LLM prompt -> produces override file
  3. Re-run script with override file -> overrides replace low-confidence mappings

SETUP:
1. Edit taxonomy_config.yaml with your mappings, keywords, and taxonomy
2. Set file paths in main() (Section 5), including SUB_RISK_PATH
3. Run: python risk_taxonomy_transformer.py
4. (Optional) Batch Review Queue through LLM, save as overrides, re-run
"""

import pandas as pd
import re
import logging
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
import yaml
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

_PROJECT_ROOT = Path(__file__).parent

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(_PROJECT_ROOT / "logs" / "transform_log.txt", mode="w"),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# =============================================================================
# SECTION 1: LOAD CONFIGURATION FROM YAML
# =============================================================================

_CONFIG_PATH = _PROJECT_ROOT / "config" / "taxonomy_config.yaml"


def _load_config(config_path: Path = _CONFIG_PATH) -> dict:
    """Load and validate taxonomy configuration from YAML."""
    logger.info(f"Loading config from {config_path}")
    with open(config_path, "r", encoding="utf-8") as f:
        cfg = yaml.safe_load(f)

    required_keys = ["risk_rating_map", "control_rating_map", "new_taxonomy",
                     "crosswalk_config", "keyword_map"]
    missing = [k for k in required_keys if k not in cfg]
    if missing:
        raise ValueError(f"taxonomy_config.yaml missing required keys: {missing}")

    # Pre-lowercase all keyword lists so inner loops don't repeat .lower()
    for l2_key in cfg["keyword_map"]:
        cfg["keyword_map"][l2_key] = [kw.lower() for kw in cfg["keyword_map"][l2_key]]

    # Pre-lowercase condition lists in crosswalk targets
    for pillar_cfg in cfg["crosswalk_config"].values():
        for target in pillar_cfg.get("targets", []):
            if "conditions" in target:
                target["conditions"] = [c.lower() for c in target["conditions"]]

    return cfg


_CFG = _load_config()

# Expose config values as module-level names for backward compatibility
HIGH_CONFIDENCE_THRESHOLD = _CFG.get("high_confidence_threshold", 3)
NA_STRINGS = tuple(_CFG.get("na_strings", ("not applicable", "n/a", "na", "")))
RATING_WORDS = "low|medium|high|critical"  # regex fragment for rationale parsing
RISK_RATING_MAP = _CFG["risk_rating_map"]
CONTROL_RATING_MAP = _CFG["control_rating_map"]
NEW_TAXONOMY = _CFG["new_taxonomy"]
CROSSWALK_CONFIG = _CFG["crosswalk_config"]
KEYWORD_MAP = _CFG["keyword_map"]

# Flatten for validation and lookup
L2_TO_L1 = {}
for _l1, _l2_list in NEW_TAXONOMY.items():
    for _l2 in _l2_list:
        L2_TO_L1[_l2] = _l1


@dataclass
class TransformContext:
    """Bundles shared lookup data passed through the transformation pipeline."""
    crosswalk: dict
    pillar_columns: dict
    sub_risk_index: dict | None = None
    overrides: dict | None = None
    findings_index: dict | None = None


# =============================================================================
# SECTION 2: INGESTION
# =============================================================================

def ingest_legacy_data(filepath: str) -> pd.DataFrame:
    """Read the legacy entity-level risk data from Excel.

    Expects a wide-format file: one row per audit entity with columns for
    each legacy pillar's rating, rationale, control assessment, and control
    rationale. Adjust column name patterns below to match your file.
    """
    logger.info(f"Reading legacy data from {filepath}")
    df = pd.read_excel(filepath)
    logger.info(f"  Loaded {len(df)} audit entities, {len(df.columns)} columns")

    # Normalize column names: strip whitespace, lowercase
    df.columns = [str(c).strip() for c in df.columns]
    return df


def ingest_crosswalk(filepath: str = None) -> dict:
    """Return the crosswalk config loaded from taxonomy_config.yaml.

    File-based crosswalk override is not yet implemented.
    Pass filepath=None to use the YAML-loaded CROSSWALK_CONFIG.
    """
    if filepath:
        raise NotImplementedError(
            f"File-based crosswalk parsing not yet implemented (got {filepath}). "
            "Use CROSSWALK_CONFIG or pass filepath=None."
        )
    return CROSSWALK_CONFIG


def ingest_sub_risks(filepath: str, entity_id_col: str, legacy_l1_col: str,
                     risk_desc_col: str, risk_id_col: str = None,
                     rating_col: str = None) -> pd.DataFrame:
    """Read sub-risk descriptions file.

    Expected columns (configure names in main()):
      - entity_id_col:  Audit Entity ID
      - risk_id_col:    Risk ID (optional, for traceability)
      - risk_desc_col:  Risk description text
      - legacy_l1_col:  Legacy L1 pillar(s), tab-separated if multiple
      - rating_col:     Inherent risk rating (optional, not used for scoring)

    Returns DataFrame with one row per sub-risk, with legacy L1s exploded
    so each row maps to a single L1.
    """
    logger.info(f"Reading sub-risk descriptions from {filepath}")
    df = pd.read_excel(filepath)
    df.columns = [c.strip() for c in df.columns]
    logger.info(f"  Loaded {len(df)} sub-risk rows")

    # Rename to standard internal names
    col_map = {entity_id_col: "entity_id", risk_desc_col: "risk_description",
               legacy_l1_col: "legacy_l1_raw"}
    if risk_id_col:
        col_map[risk_id_col] = "risk_id"
    if rating_col:
        col_map[rating_col] = "sub_risk_rating"
    df = df.rename(columns=col_map)

    # Ensure entity_id is string
    df["entity_id"] = df["entity_id"].astype(str).str.strip()

    # Explode multi-value L1 pillars so each row has one L1
    # Excel cells may use newlines, tabs, semicolons, or pipes as separators
    df["legacy_l1_list"] = df["legacy_l1_raw"].astype(str).str.split(r"\n|\t|;|\|")
    df = df.explode("legacy_l1_list")
    df["legacy_l1"] = df["legacy_l1_list"].str.strip()
    df = df[~df["legacy_l1"].isin(["", "nan"])]  # drop empty and NaN from splitting
    df = df.drop(columns=["legacy_l1_list"])

    # Clean up description text
    df["risk_description"] = df["risk_description"].astype(str).str.strip()

    logger.info(f"  After L1 explosion: {len(df)} sub-risk-to-L1 rows")
    logger.info(f"  Unique entities with sub-risks: {df['entity_id'].nunique()}")

    return df


def _build_nested_index(df: pd.DataFrame, key1_col: str, key2_col: str,
                        value_fn) -> dict:
    """Build a two-level nested index: {key1: {key2: [values]}}.

    value_fn receives each row and returns the value to append.
    """
    index = defaultdict(lambda: defaultdict(list))
    for _, row in df.iterrows():
        index[row[key1_col]][row[key2_col]].append(value_fn(row))
    # Convert back to plain dicts for consumers that check `key in index`
    return {k1: dict(v) for k1, v in index.items()}


def build_sub_risk_index(sub_risks_df: pd.DataFrame) -> dict:
    """Build lookup: {entity_id: {legacy_pillar: [list of risk descriptions]}}.

    This enables fast lookup during entity transformation.
    """
    index = _build_nested_index(
        sub_risks_df, "entity_id", "legacy_l1",
        value_fn=lambda row: (
            str(row.get("risk_id", "")),
            row["risk_description"],
        ),
    )

    # Diagnostic: show which sub-risk L1 values match crosswalk keys
    all_l1s_in_subrisk = {l1 for eid_map in index.values() for l1 in eid_map
                          if isinstance(l1, str)}
    crosswalk_keys = set(CROSSWALK_CONFIG.keys())
    matched = all_l1s_in_subrisk & crosswalk_keys
    unmatched = all_l1s_in_subrisk - crosswalk_keys
    unused = crosswalk_keys - all_l1s_in_subrisk
    logger.info(f"  Sub-risk L1 values found: {sorted(all_l1s_in_subrisk)}")
    logger.info(f"  Matched to crosswalk keys: {sorted(matched)}")
    if unmatched:
        logger.warning(f"  Sub-risk L1s NOT in crosswalk (will be ignored): {sorted(unmatched)}")
    if unused:
        logger.info(f"  Crosswalk keys with NO sub-risks: {sorted(unused)}")

    return index


def load_overrides(filepath: str) -> dict:
    """Load LLM-classified overrides from Excel/CSV.

    Expected columns:
      - entity_id:        Audit entity identifier
      - source_legacy_pillar: Legacy pillar name (must match CROSSWALK_CONFIG keys)
      - classified_l2:    The LLM-assigned L2 risk name (must match NEW_TAXONOMY)
      - llm_confidence:   Optional — High/Medium/Low from the LLM

    Returns dict: {(entity_id, legacy_pillar): {"l2": str, "confidence": str}}
    """
    logger.info(f"Loading LLM overrides from {filepath}")

    if filepath.endswith(".csv"):
        df = pd.read_csv(filepath)
    else:
        df = pd.read_excel(filepath)

    df.columns = [c.strip() for c in df.columns]
    df["entity_id"] = df["entity_id"].astype(str).str.strip()
    df["source_legacy_pillar"] = df["source_legacy_pillar"].astype(str).str.strip()
    df["classified_l2"] = df["classified_l2"].astype(str).str.strip()

    overrides = {}
    accepted_count = 0
    skipped = 0

    for _, row in df.iterrows():
        l2 = row["classified_l2"]
        if l2 not in L2_TO_L1:
            logger.warning(f"  Override skipped: '{l2}' not in taxonomy "
                        f"(entity={row['entity_id']}, pillar={row['source_legacy_pillar']})")
            skipped += 1
            continue

        key = (row["entity_id"], row["source_legacy_pillar"])
        confidence = str(row.get("llm_confidence", "high")).strip().lower()
        if confidence not in ("high", "medium", "low"):
            confidence = "high"

        overrides[key] = {"l2": l2, "confidence": confidence}
        accepted_count += 1

    logger.info(f"  Loaded {accepted_count} valid overrides, skipped {skipped} invalid")
    return overrides


def ingest_findings(filepath: str, column_name_map: dict) -> pd.DataFrame:
    """Read findings/issues data.

    Expected columns (configure names via column_name_map):
      entity_id, issue_id, l2_risk, severity, status, issue_title, remediation_date
    """
    logger.info(f"Reading findings from {filepath}")
    df = pd.read_excel(filepath)
    df.columns = [c.strip() for c in df.columns]

    rename = {}
    for internal, actual in column_name_map.items():
        if actual and actual in df.columns:
            rename[actual] = internal
    df = df.rename(columns=rename)

    df["entity_id"] = df["entity_id"].astype(str).str.strip()

    # Only include approved findings — filter out drafts/in-review
    if "Finding Approval Status" in df.columns:
        pre_filter = len(df)
        df = df[df["Finding Approval Status"].astype(str).str.strip() == "Approved"]
        logger.info(f"  Filtered to Approved findings: {len(df)} of {pre_filter}")

    # Exclude findings with blank severity — likely incomplete, shouldn't confirm applicability
    if "severity" in df.columns:
        blank_sev = df["severity"].isna() | (df["severity"].astype(str).str.strip() == "")
        if blank_sev.any():
            logger.info(f"  Excluded {blank_sev.sum()} findings with blank severity")
            df = df[~blank_sev]

    # Explode multi-value L2 risk cells (alt+enter in Excel = newline separator)
    df["l2_risk"] = df["l2_risk"].astype(str).str.split(r"\n|\r\n|\r")
    df = df.explode("l2_risk")
    df["l2_risk"] = df["l2_risk"].str.strip()
    df = df[df["l2_risk"] != ""]

    # Normalize L2 names (strip L1 prefix, resolve aliases, drop unmappable)
    df["l2_risk"] = df["l2_risk"].apply(normalize_l2_name)
    pre_norm = len(df)
    df = df[df["l2_risk"].notna()]
    dropped = pre_norm - len(df)
    if dropped > 0:
        logger.info(f"  Dropped {dropped} findings with unmappable or blank L2 risk categories")

    # Validate remaining L2 names match taxonomy
    valid = df["l2_risk"].isin(L2_TO_L1)
    invalid_l2s = df[~valid]["l2_risk"].unique()
    if len(invalid_l2s) > 0:
        logger.warning(f"  Findings L2s NOT in taxonomy (will be ignored): {list(invalid_l2s)}")
    df = df[valid]

    logger.info(f"  Loaded {len(df)} valid findings across {df['entity_id'].nunique()} entities")
    logger.info(f"  L2s covered by findings: {sorted(df['l2_risk'].unique())}")
    return df


def build_findings_index(findings_df: pd.DataFrame) -> dict:
    """Build lookup: {entity_id: {l2_risk: [list of finding dicts]}}.

    Each finding dict: {issue_id, severity, status, issue_title, remediation_date}
    """
    def _finding_from_row(row):
        return {
            "issue_id": str(row.get("issue_id", "")),
            "severity": str(row.get("severity", "")),
            "status": str(row.get("status", "")),
            "issue_title": str(row.get("issue_title", "")),
            "remediation_date": str(row.get("remediation_date", "")),
        }

    index = _build_nested_index(findings_df, "entity_id", "l2_risk",
                                value_fn=_finding_from_row)

    total_findings = sum(len(fs) for eid_map in index.values() for fs in eid_map.values())
    logger.info(f"  Findings index built: {len(index)} entities, {total_findings} total findings")
    return index


# =============================================================================
# L2 NAME NORMALIZATION — shared by findings ingestion and auxiliary risk flags
# =============================================================================

# L1 prefix pattern to strip (e.g., "Operational - Data" -> "Data")
_L1_PREFIX_PATTERN = (
    r"^(?:Operational and Compliance|Operational|Strategic|Market|Credit|"
    r"Liquidity|Reputational)\s*[-–]\s*"
)

# Known name variations -> canonical L2 name
_L2_ALIASES = {
    "earning": "Earnings",
    "earnings": "Earnings",
    "infosec": "Information and Cyber Security",
    "info security": "Information and Cyber Security",
    "information security": "Information and Cyber Security",
    "cyber security": "Information and Cyber Security",
    "cybersecurity": "Information and Cyber Security",
    "info and cyber security": "Information and Cyber Security",
    "prudential & bank admin compliance": "Prudential & bank administration compliance",
    "prudential and bank administration compliance": "Prudential & bank administration compliance",
    "prudential & bank admin": "Prudential & bank administration compliance",
    "customer / client protection": "Customer / client protection and product compliance",
    "customer/client protection and product compliance": "Customer / client protection and product compliance",
    "client protection": "Customer / client protection and product compliance",
    "fraud": "Fraud (External and Internal)",
    "external fraud": "Fraud (External and Internal)",
    "internal fraud": "Fraud (External and Internal)",
    "fraud (external & internal)": "Fraud (External and Internal)",
    "processing execution and change": "Processing, Execution and Change",
    "processing execution & change": "Processing, Execution and Change",
    "processing, execution & change": "Processing, Execution and Change",
    "fx & price": "FX and Price",
    "fx and price risk": "FX and Price",
    "interest rate risk": "Interest Rate",
    "consumer & small business": "Consumer and Small Business",
    "third-party": "Third Party",
}

# Values that are old L1 names or otherwise unmappable to a single L2
_UNMAPPABLE_L2S = {
    "nan", "Country", "Compliance", "Market", "Operational",
    "Strategic", "Credit", "Reputational", "Liquidity",
    "Fair Lending / Regulation B", "Operational - Legal",
}

# Build case-insensitive lookup (includes exact taxonomy names)
_L2_LOOKUP = {k.lower(): v for k, v in _L2_ALIASES.items()}
for _l2_name in L2_TO_L1:
    _L2_LOOKUP[_l2_name.lower()] = _l2_name

_UNMAPPABLE_LOWER = {v.lower() for v in _UNMAPPABLE_L2S}


def normalize_l2_name(raw: str) -> str | None:
    """Normalize a raw L2 risk name to the canonical taxonomy name.

    Strips L1 prefixes, resolves aliases, and returns None for unmappable values.
    """
    import re as _re
    text = str(raw).strip()
    if not text or text.lower() in ("", "nan"):
        return None

    # Strip L1 prefix
    text = _re.sub(_L1_PREFIX_PATTERN, "", text).strip()

    # Check unmappable
    if text.lower() in _UNMAPPABLE_LOWER:
        return None

    # Resolve alias or exact match
    return _L2_LOOKUP.get(text.lower())


# =============================================================================
# SECTION 3: RATING CONVERSION & RATIONALE PARSING
# =============================================================================

def _make_row(
    entity_id: str, l1: str, l2: str, *,
    likelihood=None, impact_financial=None, impact_reputational=None,
    impact_consumer_harm=None, impact_regulatory=None,
    # Three control columns intentionally accept the same value — legacy data
    # has a single control rating; downstream consumers may differentiate later.
    iag_control_effectiveness=None, aligned_assurance_rating=None,
    management_awareness_rating=None,
    source_legacy_pillar=None, source_risk_rating_raw=None,
    source_rationale="", source_control_raw=None, source_control_rationale="",
    mapping_type="", confidence="", method="",
    dims_parsed_from_rationale=False, sub_risk_evidence="", needs_review=False,
) -> dict:
    """Build a single transformed row dict with consistent keys."""
    return {
        "entity_id": entity_id,
        "new_l1": l1,
        "new_l2": l2,
        "composite_key": f"{l2} {entity_id}",
        "likelihood": likelihood,
        "impact_financial": impact_financial,
        "impact_reputational": impact_reputational,
        "impact_consumer_harm": impact_consumer_harm,
        "impact_regulatory": impact_regulatory,
        "iag_control_effectiveness": iag_control_effectiveness,
        "aligned_assurance_rating": aligned_assurance_rating,
        "management_awareness_rating": management_awareness_rating,
        "source_legacy_pillar": source_legacy_pillar,
        "source_risk_rating_raw": source_risk_rating_raw,
        "source_rationale": source_rationale,
        "source_control_raw": source_control_raw,
        "source_control_rationale": source_control_rationale,
        "mapping_type": mapping_type,
        "confidence": confidence,
        "method": method,
        "dims_parsed_from_rationale": dims_parsed_from_rationale,
        "sub_risk_evidence": sub_risk_evidence,
        "needs_review": needs_review,
    }


def convert_risk_rating(value) -> int | None:
    """Convert legacy risk rating to 1-4 numeric scale."""
    if pd.isna(value):
        return None
    return RISK_RATING_MAP.get(str(value).strip().lower())


def convert_control_rating(value) -> int | None:
    """Convert legacy control assessment to 1-4 numeric scale."""
    if pd.isna(value):
        return None
    return CONTROL_RATING_MAP.get(str(value).strip().lower())


def parse_rationale_for_dimensions(rationale: str) -> dict:
    """Extract explicit likelihood/impact mentions from rationale text.

    Handles many free-text formats:
      "likelihood is high"           "impact: medium"
      "likelihood(high)"             "impact (medium)"
      "the likelihood is medium"     "impact is high because..."
      "L: Low, I: High"             "high likelihood"
      "likelihood - low"             "likelihood = critical"
      "likelihood rating: high"      "impact rating is medium"

    Returns dict with any found dimensions; empty dict if none found.
    """
    if not rationale or pd.isna(rationale):
        return {}

    text = str(rationale).lower()
    found = {}
    for dimension in ["likelihood", "impact"]:
        # Pattern 1: "dimension <separator> rating"
        # Handles: is, :, -, =, (, and optional words like "is rated", "rating:"
        match = re.search(
            rf"{dimension}\s*(?:rating\s*)?(?:is\s*(?:rated\s*)?|:|–|-|=|\()\s*({RATING_WORDS})",
            text
        )
        if match:
            found[dimension] = RISK_RATING_MAP.get(match.group(1))
            continue

        # Pattern 2: "the dimension ... is/of rating" (words in between, up to 5)
        match = re.search(
            rf"(?:the\s+)?{dimension}\s+(?:\w+\s+){{0,5}}(?:is|of)\s+({RATING_WORDS})",
            text
        )
        if match:
            found[dimension] = RISK_RATING_MAP.get(match.group(1))
            continue

        # Pattern 3: "rating dimension" (e.g., "high likelihood")
        match = re.search(
            rf"({RATING_WORDS})\s+{dimension}",
            text
        )
        if match:
            found[dimension] = RISK_RATING_MAP.get(match.group(1))

    # Abbreviation patterns: "L: Low" / "I: High" / "L-Low, I-Medium"
    abbrev_match = re.search(
        rf"\bL\s*[:\-=]\s*({RATING_WORDS})", text
    )
    if abbrev_match and "likelihood" not in found:
        found["likelihood"] = RISK_RATING_MAP.get(abbrev_match.group(1))

    abbrev_match = re.search(
        rf"\bI\s*[:\-=]\s*({RATING_WORDS})", text
    )
    if abbrev_match and "impact" not in found:
        found["impact"] = RISK_RATING_MAP.get(abbrev_match.group(1))

    # Specific impact types: financial, reputational, regulatory, consumer
    # Map regex word -> output key (consumer matches text, but column is consumer_harm)
    impact_key_map = {
        "financial": "impact_financial",
        "reputational": "impact_reputational",
        "regulatory": "impact_regulatory",
        "consumer": "impact_consumer_harm",
    }
    for impact_type in impact_key_map:
        # "financial impact <sep> rating" or "impact <sep> financial <sep> rating"
        match = re.search(
            rf"{impact_type}\s+impact\s*(?:rating\s*)?(?:is\s*(?:rated\s*)?|:|–|-|=|\()?\s*({RATING_WORDS})",
            text
        )
        if match:
            found[impact_key_map[impact_type]] = RISK_RATING_MAP.get(match.group(1))
            continue

        # "rating financial impact"
        match = re.search(
            rf"({RATING_WORDS})\s+{impact_type}\s+impact",
            text
        )
        if match:
            found[impact_key_map[impact_type]] = RISK_RATING_MAP.get(match.group(1))
            continue

        # "impact - financial: rating" or "impact (financial): rating"
        match = re.search(
            rf"impact\s*[\-(]\s*{impact_type}\s*[):]?\s*(?:is\s*)?({RATING_WORDS})" ,
            text
        )
        if match:
            found[impact_key_map[impact_type]] = RISK_RATING_MAP.get(match.group(1))

    return found


# Methods that represent unrated/placeholder rows — overridden by rated or confirmed rows during dedup
BLANK_METHODS = ("evaluated_no_evidence", "gap_fill", "true_gap_fill", "no_evidence_all_candidates")

# =============================================================================
# SECTION 4: MAPPING ENGINE
# =============================================================================


def _resolve_multi_mapping(
    entity_id: str,
    legacy_pillar: str,
    pillar_config: dict,
    rationale: str,
    sub_risk_index: dict | None,
    overrides: dict | None,
) -> list[dict] | None:
    """Resolve a multi-target mapping by scoring evidence for each candidate L2.

    Returns list of target dicts [{l2, confidence, method, sub_risk_evidence}],
    or None if the mapping produced no targets and has no primary fallback.
    """
    targets_to_create = []

    # Prepare text sources separately so evidence can be labeled
    rationale_text = str(rationale).lower() if rationale else ""
    entity_subs = (sub_risk_index or {}).get(entity_id, {})
    sub_risk_entries = entity_subs.get(legacy_pillar, [])  # list of (risk_id, description)

    # Look up LLM override once for this entity+pillar
    override_entry = None
    if overrides and entity_id:
        override_entry = overrides.get((entity_id, legacy_pillar))

    first_primary_l2 = None
    for target in pillar_config["targets"]:
        if target["relationship"] == "primary" and not first_primary_l2:
            first_primary_l2 = target["l2"]

        # Check LLM override for this target L2
        if override_entry and override_entry["l2"] == target["l2"]:
            targets_to_create.append({
                "l2": target["l2"],
                "confidence": override_entry["confidence"],
                "method": "llm_override",
                "sub_risk_evidence": [],
            })
            continue

        # Score this L2 against rationale and sub-risk descriptions separately
        l2_name = target["l2"]
        keywords = KEYWORD_MAP.get(l2_name, [])
        conditions = target.get("conditions", [])
        all_keywords = keywords + conditions

        labeled_evidence = []
        score = 0

        # Check rationale text
        rationale_hits = [kw for kw in all_keywords if kw in rationale_text]
        if rationale_hits:
            score += len(rationale_hits)
            labeled_evidence.append(f"rationale: {', '.join(rationale_hits)}")

        # Check each sub-risk description individually
        for risk_id, desc in sub_risk_entries:
            desc = str(desc) if desc is not None else ""
            if not desc or desc == "nan":
                continue
            desc_lower = desc.lower()
            desc_hits = [kw for kw in all_keywords if kw in desc_lower]
            if desc_hits:
                score += len(desc_hits)
                truncated = desc[:80] + "..." if len(desc) > 80 else desc
                labeled_evidence.append(f"sub-risk {risk_id} [{truncated}]: {', '.join(desc_hits)}")

        relationship = target["relationship"]

        if score > 0:
            if score >= HIGH_CONFIDENCE_THRESHOLD:
                confidence = "high"
            else:
                confidence = "medium"
            method = f"evidence_match ({relationship})"
            targets_to_create.append({
                "l2": l2_name,
                "confidence": confidence,
                "method": method,
                "sub_risk_evidence": labeled_evidence[:8],
            })

    # If no L2s had evidence, populate ALL candidate L2s and flag for team review.
    # Don't pick one for them — present the data and let them decide applicability.
    if not targets_to_create:
        candidate_l2s = [t["l2"] for t in pillar_config["targets"]]
        if candidate_l2s:
            for l2_name in candidate_l2s:
                targets_to_create.append({
                    "l2": l2_name,
                    "confidence": "low",
                    "method": "no_evidence_all_candidates",
                    "sub_risk_evidence": [],
                })
            logger.info(
                f"  Entity {entity_id}: '{legacy_pillar}' -> no evidence for any L2, "
                f"populated all {len(candidate_l2s)} candidates — FLAGGED FOR REVIEW"
            )
        else:
            logger.warning(
                f"  Entity {entity_id}: '{legacy_pillar}' multi mapping "
                f"produced no targets and has no candidates"
            )
            return None

    return targets_to_create


def _deduplicate_transformed_rows(transformed: list[dict], entity_id: str) -> list[dict]:
    """Deduplicate when multiple legacy pillars map to the same new L2.

    Rules:
      - If one row has ratings and the other doesn't, keep the one WITH ratings
        but append issue info from findings rows to sub_risk_evidence
      - If both have ratings, keep the higher (more conservative) rating
      - If issue_confirmed vs evaluated_no_evidence/gap_fill, keep issue_confirmed
    """
    seen = {}
    deduped = []
    for row in transformed:
        l2 = row["new_l2"]
        if l2 not in seen:
            seen[l2] = len(deduped)
            deduped.append(row)
        else:
            existing = deduped[seen[l2]]
            existing_rating = existing.get("likelihood") or 0
            new_rating = row.get("likelihood") or 0
            existing_method = existing.get("method", "")
            new_method = row.get("method", "")

            if new_method == "issue_confirmed" and existing_method in BLANK_METHODS:
                deduped[seen[l2]] = row
            elif existing_method == "issue_confirmed" and new_method in BLANK_METHODS:
                pass
            elif existing_method == "issue_confirmed" and new_rating > 0:
                row["sub_risk_evidence"] = (
                    (row.get("sub_risk_evidence", "") + " | " + existing.get("sub_risk_evidence", "")).strip(" | ")
                )
                row["source_legacy_pillar"] = (
                    f"{row['source_legacy_pillar']} (also: Findings)"
                )
                deduped[seen[l2]] = row
            elif new_method == "issue_confirmed" and existing_rating > 0:
                existing["sub_risk_evidence"] = (
                    (existing.get("sub_risk_evidence", "") + " | " + row.get("sub_risk_evidence", "")).strip(" | ")
                )
                existing["source_legacy_pillar"] = (
                    f"{existing['source_legacy_pillar']} (also: Findings)"
                )
            elif new_rating > existing_rating:
                row["source_legacy_pillar"] = (
                    f"{row['source_legacy_pillar']} (also: {existing['source_legacy_pillar']})"
                )
                row["method"] = f"{row['method']} (dedup: kept higher)"
                deduped[seen[l2]] = row
            else:
                existing["source_legacy_pillar"] = (
                    f"{existing['source_legacy_pillar']} (also: {row['source_legacy_pillar']})"
                )
                if "dedup" not in existing_method:
                    existing["method"] = f"{existing_method} (dedup: kept higher)"

            logger.info(
                f"  Entity {entity_id}: DEDUP '{l2}' — "
                f"'{row.get('source_legacy_pillar', '')}' [{new_method}] vs "
                f"existing [{existing_method}]"
            )
    return deduped


def transform_entity(
    entity_id: str,
    entity_row: pd.Series,
    ctx: TransformContext,
) -> tuple[list[dict], list[dict]]:
    """Transform one audit entity from legacy to new taxonomy.

    Returns:
        (transformed_rows, overlay_flags)
    """
    crosswalk = ctx.crosswalk
    pillar_columns = ctx.pillar_columns
    sub_risk_index = ctx.sub_risk_index
    overrides = ctx.overrides
    findings_index = ctx.findings_index

    transformed = []
    overlays = []
    mapped_l2s = set()

    # --- Pre-check: findings-confirmed L2s ---
    # If an entity has findings tagged to a new L2, that L2 is confirmed applicable.
    # Create placeholder rows with no ratings (ratings come from legacy pillar data).
    # These will be overridden by dedup if the crosswalk also produces rated rows.
    if findings_index:
        entity_findings = findings_index.get(entity_id, {})
        for l2, findings_list in entity_findings.items():
            if l2 in L2_TO_L1:
                issue_summaries = [
                    f"{f['issue_id']}: {f['issue_title']} ({f['severity']}, {f['status']})"
                    for f in findings_list[:5]
                ]
                l1 = L2_TO_L1[l2]
                mapped_l2s.add(l2)
                transformed.append(_make_row(
                    entity_id, l1, l2,
                    source_legacy_pillar="Findings",
                    mapping_type="findings",
                    confidence="high",
                    method="issue_confirmed",
                    sub_risk_evidence="; ".join(issue_summaries),
                ))
                logger.info(f"  Entity {entity_id}: '{l2}' confirmed applicable by {len(findings_list)} finding(s)")

    for legacy_pillar, pillar_config in crosswalk.items():
        # Get legacy data for this pillar
        cols = pillar_columns.get(legacy_pillar)
        if not cols:
            logger.warning(f"  Entity {entity_id}: No columns found for '{legacy_pillar}'")
            continue

        rating_raw = entity_row.get(cols.get("rating"))
        rationale = entity_row.get(cols.get("rationale"), "")
        control_raw = entity_row.get(cols.get("control"))
        control_rationale = entity_row.get(cols.get("control_rationale"), "")

        rating_numeric = convert_risk_rating(rating_raw)
        control_numeric = convert_control_rating(control_raw)

        # Skip N/A ratings — flag all candidate L2s as not applicable
        raw_str = str(rating_raw).strip().lower() if rating_raw and not pd.isna(rating_raw) else ""
        is_na = (rating_numeric is None and raw_str in NA_STRINGS)

        if is_na and pillar_config.get("mapping_type") != "overlay":
            # Determine which L2s this pillar would have mapped to
            na_mapping_type = pillar_config.get("mapping_type", "")
            if na_mapping_type == "direct":
                na_l2s = [pillar_config["target_l2"]]
            elif na_mapping_type == "multi":
                na_l2s = [t["l2"] for t in pillar_config["targets"]]
            else:
                na_l2s = []

            for l2_name in na_l2s:
                l1 = L2_TO_L1.get(l2_name, "UNKNOWN")
                mapped_l2s.add(l2_name)
                transformed.append(_make_row(
                    entity_id, l1, l2_name,
                    source_legacy_pillar=legacy_pillar,
                    source_risk_rating_raw=rating_raw,
                    source_rationale=str(rationale) if rationale else "",
                    source_control_raw=control_raw,
                    mapping_type=pillar_config.get("mapping_type", ""),
                    confidence="high",
                    method="source_not_applicable",
                ))
            logger.info(f"  Entity {entity_id}: '{legacy_pillar}' -> N/A, flagged {len(na_l2s)} L2s as not applicable")
            continue

        # Parse rationale for explicit dimension mentions
        parsed_dims = parse_rationale_for_dimensions(str(rationale))

        # Build the 5 risk dimension values
        likelihood = parsed_dims.get("likelihood", rating_numeric)
        impact_financial = parsed_dims.get("impact_financial", rating_numeric)
        impact_reputational = parsed_dims.get("impact_reputational", rating_numeric)
        impact_consumer_harm = parsed_dims.get("impact_consumer_harm", rating_numeric)
        impact_regulatory = parsed_dims.get("impact_regulatory", rating_numeric)
        # If generic "impact" was parsed, use it as default for all impact cols
        if "impact" in parsed_dims:
            generic_impact = parsed_dims["impact"]
            impact_financial = parsed_dims.get("impact_financial", generic_impact)
            impact_reputational = parsed_dims.get("impact_reputational", generic_impact)
            impact_consumer_harm = parsed_dims.get("impact_consumer_harm", generic_impact)
            impact_regulatory = parsed_dims.get("impact_regulatory", generic_impact)

        mapping_type = pillar_config["mapping_type"]

        if mapping_type == "overlay":
            for target_l2 in pillar_config["target_l2s"]:
                overlays.append({
                    "entity_id": entity_id,
                    "target_l2": target_l2,
                    "overlay_source": legacy_pillar,
                    "overlay_rating": rating_numeric,
                    "overlay_rationale": str(rationale),
                })
            logger.info(f"  Entity {entity_id}: '{legacy_pillar}' -> overlay on {pillar_config['target_l2s']}")
            continue

        # Build list of target L2s to create rows for
        if mapping_type == "direct":
            targets_to_create = [{
                "l2": pillar_config["target_l2"],
                "confidence": "high",
                "method": "direct",
                "sub_risk_evidence": [],
            }]

        elif mapping_type == "multi":
            # If no rationale column exists for this pillar, skip keyword scoring —
            # populate all primary L2s directly with high confidence
            has_rationale = cols.get("rationale") is not None
            if not has_rationale:
                targets_to_create = [
                    {
                        "l2": t["l2"],
                        "confidence": "high",
                        "method": "direct (no rationale column)",
                        "sub_risk_evidence": [],
                    }
                    for t in pillar_config["targets"]
                    if t["relationship"] == "primary"
                ]
            else:
                targets_to_create = _resolve_multi_mapping(
                    entity_id, legacy_pillar, pillar_config, rationale,
                    sub_risk_index, overrides,
                )
            if targets_to_create is None:
                continue

            # Track candidate L2s that were evaluated but had no evidence
            matched_l2s_this_pillar = {t["l2"] for t in targets_to_create}
            for target in pillar_config["targets"]:
                candidate_l2 = target["l2"]
                if candidate_l2 not in matched_l2s_this_pillar:
                    l1 = L2_TO_L1.get(candidate_l2, "UNKNOWN")
                    mapped_l2s.add(candidate_l2)
                    transformed.append(_make_row(
                        entity_id, l1, candidate_l2,
                        source_legacy_pillar=legacy_pillar,
                        source_risk_rating_raw=rating_raw,
                        source_rationale=str(rationale) if rationale else "",
                        source_control_raw=control_raw,
                        mapping_type=mapping_type,
                        confidence="none",
                        method="evaluated_no_evidence",
                    ))

        else:
            logger.error(f"  Unknown mapping_type '{mapping_type}' for '{legacy_pillar}'")
            continue

        dims_were_parsed = bool(parsed_dims)

        for target_match in targets_to_create:
            selected_l2 = target_match["l2"]
            l1 = L2_TO_L1.get(selected_l2, "UNKNOWN")
            mapped_l2s.add(selected_l2)

            row = _make_row(
                entity_id, l1, selected_l2,
                likelihood=likelihood,
                impact_financial=impact_financial,
                impact_reputational=impact_reputational,
                impact_consumer_harm=impact_consumer_harm,
                impact_regulatory=impact_regulatory,
                iag_control_effectiveness=control_numeric,
                aligned_assurance_rating=control_numeric,
                management_awareness_rating=control_numeric,
                source_legacy_pillar=legacy_pillar,
                source_risk_rating_raw=rating_raw,
                source_rationale=str(rationale) if rationale else "",
                source_control_raw=control_raw,
                source_control_rationale=str(control_rationale) if control_rationale else "",
                mapping_type=mapping_type,
                confidence=target_match["confidence"],
                method=target_match["method"],
                dims_parsed_from_rationale=dims_were_parsed,
                sub_risk_evidence="; ".join(target_match["sub_risk_evidence"]) if target_match["sub_risk_evidence"] else "",
                needs_review=target_match["confidence"] == "low",
            )
            transformed.append(row)
            logger.info(
                f"  Entity {entity_id}: '{legacy_pillar}' -> {l1} / {selected_l2} "
                f"[{target_match['method']}, conf={target_match['confidence']}]"
            )

    transformed = _deduplicate_transformed_rows(transformed, entity_id)

    # Identify any new L2 risks with NO legacy mapping at all (true gaps)
    # With the current crosswalk this should be zero.
    for l2 in L2_TO_L1:
        if l2 not in mapped_l2s:
            l1 = L2_TO_L1[l2]
            transformed.append(_make_row(
                entity_id, l1, l2,
                mapping_type="no_legacy_source",
                confidence="none",
                method="true_gap_fill",
            ))

    return transformed, overlays


# =============================================================================
# SECTION 5: PIPELINE & EXPORT
# =============================================================================

def _log_transformation_summary(transformed_df: pd.DataFrame, overlays_df: pd.DataFrame):
    """Log aggregate statistics about the transformation results."""
    total = len(transformed_df)
    if total == 0:
        logger.info("TRANSFORMATION COMPLETE — no rows produced")
        return

    # Compute all counts from a single pass over the method column
    method_counts = transformed_df["method"].value_counts()
    def method_contains(substr):
        return transformed_df["method"].str.contains(substr, na=False)

    conf_counts = transformed_df["confidence"].value_counts()
    needs_review = transformed_df["needs_review"].sum()
    evidence_mask = method_contains("evidence_match")
    dims_parsed = transformed_df["dims_parsed_from_rationale"].sum()

    logger.info("=" * 60)
    logger.info("TRANSFORMATION COMPLETE")
    logger.info(f"  Total rows: {total}")
    logger.info(f"  High confidence: {conf_counts.get('high', 0)} ({conf_counts.get('high', 0)/total*100:.1f}%)")
    logger.info(f"  Medium confidence: {conf_counts.get('medium', 0)} ({conf_counts.get('medium', 0)/total*100:.1f}%)")
    logger.info(f"  Low confidence / needs review: {conf_counts.get('low', 0)} ({conf_counts.get('low', 0)/total*100:.1f}%)")
    logger.info(f"  Source N/A (skipped): {method_counts.get('source_not_applicable', 0)}")
    logger.info(f"  Assumed Not Applicable (no evidence found): {method_counts.get('evaluated_no_evidence', 0)}")
    logger.info(f"  True gap fills (no legacy pillar maps): {method_counts.get('true_gap_fill', 0)}")
    evidence_total = evidence_mask.sum()
    evidence_high = (evidence_mask & (transformed_df["confidence"] == "high")).sum()
    evidence_med = (evidence_mask & (transformed_df["confidence"] == "medium")).sum()
    logger.info(f"  Evidence-based matches: {evidence_total} (high: {evidence_high}, medium: {evidence_med})")
    logger.info(f"  Issue-confirmed applicable: {method_counts.get('issue_confirmed', 0)}")
    logger.info(f"  No evidence — all candidates (flagged for review): {method_counts.get('no_evidence_all_candidates', 0)}")
    logger.info(f"  Resolved via LLM overrides: {method_contains('llm_override').sum()}")
    logger.info(f"  Deduplicated (multiple sources -> same L2): {method_contains('dedup').sum()}")
    logger.info(f"  Dimensions parsed from rationale: {dims_parsed}")
    logger.info(f"  Overlay flags: {len(overlays_df)}")
    logger.info(f"  Applicability undetermined (team decision required): {needs_review}")
    logger.info("=" * 60)


def run_pipeline(
    legacy_df: pd.DataFrame,
    entity_id_col: str,
    ctx: TransformContext,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Run the full transformation pipeline across all entities."""
    all_transformed = []
    all_overlays = []

    for i, (_, row) in enumerate(legacy_df.iterrows(), start=1):
        entity_id = str(row[entity_id_col]).strip()
        logger.info(f"Processing entity {entity_id} ({i}/{len(legacy_df)})")
        transformed, overlays = transform_entity(entity_id, row, ctx)
        all_transformed.extend(transformed)
        all_overlays.extend(overlays)

    transformed_df = pd.DataFrame(all_transformed)
    overlays_df = pd.DataFrame(all_overlays) if all_overlays else pd.DataFrame()

    _log_transformation_summary(transformed_df, overlays_df)

    return transformed_df, overlays_df


def apply_overlay_flags(transformed_df: pd.DataFrame, overlays_df: pd.DataFrame) -> pd.DataFrame:
    """Join overlay flags onto the transformed data."""
    if overlays_df.empty:
        transformed_df["overlay_flag"] = False
        transformed_df["overlay_source"] = ""
        transformed_df["overlay_rating"] = None
        transformed_df["overlay_rationale"] = ""
        return transformed_df

    # Aggregate overlays per entity+L2
    overlay_agg = (
        overlays_df
        .groupby(["entity_id", "target_l2"])
        .agg({
            "overlay_source": lambda x: "; ".join(x),
            "overlay_rating": "max",
            "overlay_rationale": lambda x: " | ".join(x),
        })
        .reset_index()
    )

    merged = transformed_df.merge(
        overlay_agg,
        left_on=["entity_id", "new_l2"],
        right_on=["entity_id", "target_l2"],
        how="left",
        suffixes=("", "_overlay"),
    )
    merged["overlay_flag"] = merged["target_l2"].notna()
    merged.drop(columns=["target_l2"], errors="ignore", inplace=True)
    merged["overlay_source"] = merged["overlay_source"].fillna("")
    merged["overlay_rationale"] = merged["overlay_rationale"].fillna("")

    return merged


def flag_control_contradictions(transformed_df: pd.DataFrame, findings_index: dict) -> pd.DataFrame:
    """Flag rows where control rating contradicts open findings.

    Adds a 'control_flag' column with human-readable warnings.
    """
    if not findings_index:
        transformed_df["control_flag"] = ""
        return transformed_df

    flags = []
    for _, row in transformed_df.iterrows():
        eid = str(row.get("entity_id", ""))
        l2 = row.get("new_l2", "")
        control_eff = row.get("iag_control_effectiveness")

        entity_findings = findings_index.get(eid, {})
        l2_findings = entity_findings.get(l2, [])

        # Active findings — exclude Closed, Cancelled, and Not Started
        open_findings = [
            f for f in l2_findings
            if str(f.get("status", "")).strip().lower()
            in ("open", "in validation", "in sustainability")
        ]

        if not open_findings or control_eff is None:
            flags.append("")
            continue

        flag_parts = []
        for f in open_findings[:3]:
            sev = f.get("severity", "")
            iid = f.get("issue_id", "")
            title = f.get("issue_title", "")[:80]

            if control_eff == 1:
                flag_parts.append(
                    f"Open {sev} issue ({iid}: {title}) — "
                    f"review whether Well Controlled rating reflects current state"
                )
            elif control_eff == 2 and str(sev).strip().lower() in ("high", "critical"):
                flag_parts.append(
                    f"Open High issue ({iid}: {title}) — "
                    f"consider whether Moderately Controlled rating is appropriate"
                )

        flags.append(" | ".join(flag_parts))

    transformed_df["control_flag"] = flags
    return transformed_df


# Application/engagement columns in the legacy data
_APP_COLS = {
    "primary_it": "PRIMARY IT APPLICATIONS (MAPPED)",
    "secondary_it": "SECONDARY IT APPLICATIONS (RELATED OR RELIED ON)",
    "primary_tp": "PRIMARY TLM THIRD PARTY ENGAGEMENT",
    "secondary_tp": "SECONDARY TLM THIRD PARTY ENGAGEMENTS (RELATED OR RELIED ON)",
}

# Which L2s are flagged by which application/engagement columns
_APP_L2_MAP = {
    "Technology": ("primary_it", "secondary_it"),
    "Data": ("primary_it", "secondary_it"),
    "Information and Cyber Security": ("primary_it", "secondary_it"),
    "Third Party": ("primary_tp", "secondary_tp"),
}


def flag_application_applicability(
    transformed_df: pd.DataFrame,
    legacy_df: pd.DataFrame,
    entity_id_col: str,
) -> pd.DataFrame:
    """Flag L2 risks as potentially applicable when IT applications or
    third party engagements are tagged to the entity.

    Adds an 'app_flag' column with a recommendation message.
    """
    # Check which app columns exist in the legacy data
    available_cols = {key: col for key, col in _APP_COLS.items() if col in legacy_df.columns}
    if not available_cols:
        transformed_df["app_flag"] = ""
        return transformed_df

    # Build lookup: {entity_id: {col_key: [list of IDs]}}
    entity_apps = {}
    for _, row in legacy_df.iterrows():
        eid = str(row[entity_id_col]).strip()
        entity_apps[eid] = {}
        for key, col in available_cols.items():
            raw = str(row.get(col, ""))
            if raw and raw not in ("", "nan", "None"):
                # Split on newlines (alt+enter in Excel) and semicolons
                ids = [v.strip() for v in raw.replace("\r\n", "\n").replace("\r", "\n").split("\n")
                       if v.strip() and v.strip() != "nan"]
                entity_apps[eid][key] = ids
            else:
                entity_apps[eid][key] = []

    flags = []
    for _, row in transformed_df.iterrows():
        eid = str(row.get("entity_id", ""))
        l2 = row.get("new_l2", "")

        app_col_keys = _APP_L2_MAP.get(l2)
        if not app_col_keys:
            flags.append("")
            continue

        apps = entity_apps.get(eid, {})
        flag_parts = []

        for col_key in app_col_keys:
            ids = apps.get(col_key, [])
            if ids:
                id_list = ", ".join(ids[:5])
                if len(ids) > 5:
                    id_list += f" (+{len(ids) - 5} more)"

                if col_key == "primary_it":
                    flag_parts.append(
                        f"Primary application mapped to entity ({id_list}) — "
                        f"consider this risk may be applicable"
                    )
                elif col_key == "secondary_it":
                    flag_parts.append(
                        f"Secondary application related to entity ({id_list}) — "
                        f"consider this risk may be applicable"
                    )
                elif col_key == "primary_tp":
                    flag_parts.append(
                        f"Primary third party engagement mapped to entity ({id_list}) — "
                        f"consider this risk may be applicable"
                    )
                elif col_key == "secondary_tp":
                    flag_parts.append(
                        f"Secondary third party engagement related to entity ({id_list}) — "
                        f"consider this risk may be applicable"
                    )

        flags.append(" | ".join(flag_parts))

    transformed_df["app_flag"] = flags

    flagged = sum(1 for f in flags if f)
    logger.info(f"  Application/engagement flags: {flagged} rows flagged across "
                f"{len({row.get('entity_id') for _, row in transformed_df.iterrows() if row.get('entity_id') in entity_apps and any(entity_apps[row.get('entity_id')].values())})} entities")

    return transformed_df


# Auxiliary risk dimension columns in the legacy data
_AUX_COLS = [
    "AXP Auxiliary Risk Dimensions",
    "AENB Auxiliary Risk Dimensions",
]


def flag_auxiliary_risks(
    transformed_df: pd.DataFrame,
    legacy_df: pd.DataFrame,
    entity_id_col: str,
) -> pd.DataFrame:
    """Flag L2 risks as potentially applicable when they appear in the entity's
    auxiliary risk dimensions columns.

    Adds an 'aux_flag' column with a recommendation message.
    """
    # Check which aux columns exist
    available_cols = [c for c in _AUX_COLS if c in legacy_df.columns]
    if not available_cols:
        transformed_df["aux_flag"] = ""
        return transformed_df

    # Build lookup: {entity_id: set of normalized L2 names from auxiliary columns}
    entity_aux = {}
    for _, row in legacy_df.iterrows():
        eid = str(row[entity_id_col]).strip()
        aux_l2s = {}  # {l2_name: source_column}
        for col in available_cols:
            raw = str(row.get(col, ""))
            if raw and raw not in ("", "nan", "None"):
                # Split on newlines and commas
                entries = raw.replace("\r\n", "\n").replace("\r", "\n").split("\n")
                for entry in entries:
                    entry = entry.strip()
                    if not entry:
                        continue
                    normalized = normalize_l2_name(entry)
                    if normalized and normalized not in aux_l2s:
                        aux_l2s[normalized] = col
        entity_aux[eid] = aux_l2s

    flags = []
    for _, row in transformed_df.iterrows():
        eid = str(row.get("entity_id", ""))
        l2 = row.get("new_l2", "")

        aux = entity_aux.get(eid, {})
        if l2 in aux:
            source = aux[l2]
            short_source = "AXP" if "AXP" in source else "AENB"
            flags.append(
                f"Listed as auxiliary risk in legacy entity data ({short_source}) — "
                f"consider this risk may be applicable"
            )
        else:
            flags.append("")

    transformed_df["aux_flag"] = flags

    flagged = sum(1 for f in flags if f)
    entities_flagged = len({eid for eid, aux in entity_aux.items() if aux})
    logger.info(f"  Auxiliary risk flags: {flagged} rows flagged across {entities_flagged} entities")

    return transformed_df


# Inherent Risk Rating matrix: (likelihood, overall_impact) -> rating
_RISK_MATRIX = {
    (1, 1): 1, (1, 2): 1, (1, 3): 2, (1, 4): 2,
    (2, 1): 1, (2, 2): 2, (2, 3): 2, (2, 4): 3,
    (3, 1): 2, (3, 2): 2, (3, 3): 3, (3, 4): 4,
    (4, 1): 2, (4, 2): 3, (4, 3): 4, (4, 4): 4,
}

_RATING_LABELS = {1: "Low", 2: "Medium", 3: "High", 4: "Critical"}


def derive_inherent_risk_rating(transformed_df: pd.DataFrame) -> pd.DataFrame:
    """Derive the composite Inherent Risk Rating from Likelihood × max(Impact dimensions).

    Adds columns: overall_impact, inherent_risk_rating, inherent_risk_rating_label.
    """
    impact_cols = ["impact_financial", "impact_reputational",
                   "impact_consumer_harm", "impact_regulatory"]

    def _compute(row):
        method = str(row.get("method", ""))

        # Source N/A rows get an explicit "N/A" label — this is a real determination
        if "source_not_applicable" in method:
            return None, None, "Not Applicable"

        likelihood = row.get("likelihood")
        if pd.isna(likelihood) or likelihood is None:
            return None, None, None

        impacts = [row.get(c) for c in impact_cols]
        valid_impacts = [int(v) for v in impacts if v is not None and not pd.isna(v)]
        if not valid_impacts:
            return None, None, None

        overall_impact = max(valid_impacts)
        rating = _RISK_MATRIX.get((int(likelihood), overall_impact))
        label = _RATING_LABELS.get(rating, "") if rating else None
        return overall_impact, rating, label

    results = transformed_df.apply(_compute, axis=1, result_type="expand")
    results.columns = ["overall_impact", "inherent_risk_rating", "inherent_risk_rating_label"]
    transformed_df = pd.concat([transformed_df, results], axis=1)

    rated = transformed_df["inherent_risk_rating"].notna().sum()
    logger.info(f"  Inherent Risk Rating derived for {rated} of {len(transformed_df)} rows")

    return transformed_df


def _derive_decision_basis(row) -> str:
    """Plain-language explanation of mapping method for a transformed row.

    Checks base method substrings before the dedup suffix so the explanation
    reflects the original method, with a note about multiple sources if deduped.
    """
    method = str(row.get("method", ""))
    pillar = str(row.get("source_legacy_pillar", "")).split(" (also")[0].strip()
    evidence = str(row.get("sub_risk_evidence", ""))
    rating = str(row.get("source_risk_rating_raw", ""))
    if rating in ("", "nan", "None"):
        rating = "unknown"
    dedup_note = (" This L2 was also referenced by other legacy pillars; "
                  "the higher rating was kept.") if "dedup" in method else ""

    if "source_not_applicable" in method:
        return (f"The legacy {pillar} pillar was rated Not Applicable for this entity, "
                f"so this L2 risk is also marked as not applicable.{dedup_note}")
    if "evaluated_no_evidence" in method:
        return (f"The {pillar} pillar rationale was reviewed for relevance to this L2 risk. "
                f"No direct connection was found, so this L2 is marked as not applicable "
                f"for this entity. If your review of the rationale suggests otherwise, "
                f"this can be changed to applicable.{dedup_note}")
    if "no_evidence_all_candidates" in method:
        return (f"The {pillar} pillar (rated {rating}) covers multiple L2 risks. "
                f"The rationale didn't clearly indicate which ones apply, so all candidates "
                f"are shown with the original rating as a starting point. Review the rationale "
                f"below and determine which of these L2s are relevant to this entity.{dedup_note}")
    if "true_gap_fill" in method or "gap_fill" in method:
        return ("No legacy pillar maps to this L2 risk. This is a new risk category "
                "that will need to be assessed from scratch.")
    if "direct" in method:
        return (f"The legacy {pillar} pillar maps directly to this L2 risk. "
                f"The original rating ({rating}) is carried forward as a starting point.{dedup_note}")
    if "issue_confirmed" in method:
        return f"Confirmed applicable — open finding: {evidence}{dedup_note}"
    if "evidence_match" in method:
        if evidence:
            return (f"This L2 was mapped from the {pillar} pillar (rated {rating}) based on "
                    f"references found in the rationale and sub-risk descriptions. "
                    f"Matched references: {evidence}{dedup_note}")
        return (f"This L2 was mapped from the {pillar} pillar (rated {rating}) based on "
                f"keyword evidence in the rationale text.{dedup_note}")
    if "llm_override" in method:
        return (f"This L2 was classified based on an AI review of the {pillar} pillar "
                f"rationale and sub-risk descriptions.{dedup_note}")
    return method


def _derive_status(method) -> str:
    """Map a mapping method string to a human-readable status.

    Checks base method substrings before the dedup suffix, so a deduped
    evaluated_no_evidence stays "Not Applicable" rather than flipping to "Applicable".
    """
    method = str(method)
    if "source_not_applicable" in method:
        return "Not Applicable"
    if "evaluated_no_evidence" in method:
        return "Assumed Not Applicable"
    if "no_evidence_all_candidates" in method:
        return "Applicability Undetermined"
    if "true_gap_fill" in method or "gap_fill" in method:
        return "Not Assessed"
    if ("direct" in method or "evidence_match" in method
            or "llm_override" in method or "issue_confirmed" in method
            or "dedup" in method):
        return "Applicable"
    return "Needs Review"


def build_audit_review_df(transformed_df: pd.DataFrame,
                          legacy_df: pd.DataFrame = None,
                          entity_id_col: str = "Audit Entity") -> pd.DataFrame:
    """Build the auditor-facing Audit Review dataframe with plain-language columns."""
    df = transformed_df.copy()

    # Join organizational metadata from legacy data if available
    org_cols = ["Audit Leader", "PGA/ASL", "Core Audit Team"]
    if legacy_df is not None:
        available_org = [c for c in org_cols if c in legacy_df.columns]
        if available_org and entity_id_col in legacy_df.columns:
            org_data = legacy_df[[entity_id_col] + available_org].copy()
            org_data = org_data.rename(columns={entity_id_col: "entity_id"})
            org_data["entity_id"] = org_data["entity_id"].astype(str).str.strip()
            df = df.merge(org_data, on="entity_id", how="left")

    df["Status"] = df["method"].apply(_derive_status)
    df["Decision Basis"] = df.apply(_derive_decision_basis, axis=1)

    # Rating Source column
    def derive_rating_source(row):
        has_ratings = row.get("likelihood") is not None and not pd.isna(row.get("likelihood", None))
        if not has_ratings:
            return "No ratings — legacy source was N/A or not assessed"

        parts = []
        if row.get("dims_parsed_from_rationale"):
            parts.append("Inherent Risk: Parsed from rationale — likelihood and impact stated separately")
        else:
            raw = row.get("source_risk_rating_raw", "")
            parts.append(f"Inherent Risk: Carried from legacy pillar rating {raw}")

        ctrl = row.get("source_control_raw", "")
        if row.get("iag_control_effectiveness") is not None and not pd.isna(row.get("iag_control_effectiveness")):
            parts.append(f"Control: All 3 set from {ctrl}")
        else:
            parts.append("Control: Not assessed")

        return " | ".join(parts)

    df["Rating Source"] = df.apply(derive_rating_source, axis=1)

    # Consolidate all flags into a single "Additional Signals" column
    def consolidate_signals(row):
        signals = []
        for col in ("control_flag", "app_flag", "aux_flag"):
            val = row.get(col, "")
            if pd.isna(val):
                continue
            val = str(val).strip()
            if val:
                signals.append(val)
        return " | ".join(signals)

    df["Additional Signals"] = df.apply(consolidate_signals, axis=1)

    # Select and rename columns
    audit_cols = {
        "entity_id": "Entity ID",
        "Audit Leader": "Audit Leader",
        "PGA/ASL": "PGA",
        "Core Audit Team": "Core Audit Team",
        "new_l1": "New L1",
        "new_l2": "New L2",
        "Status": "Status",
        "inherent_risk_rating_label": "Inherent Risk Rating",
        "Decision Basis": "Decision Basis",
        "Rating Source": "Rating Source",
        "Additional Signals": "Additional Signals",
        "likelihood": "Likelihood",
        "overall_impact": "Overall Impact",
        "impact_financial": "Impact - Financial",
        "impact_reputational": "Impact - Reputational",
        "impact_consumer_harm": "Impact - Consumer Harm",
        "impact_regulatory": "Impact - Regulatory",
        "iag_control_effectiveness": "IAG Control Effectiveness",
        "aligned_assurance_rating": "Aligned Assurance Rating",
        "management_awareness_rating": "Management Awareness Rating",
        "source_legacy_pillar": "Legacy Source",
        "confidence": "Confidence",
    }

    available = {k: v for k, v in audit_cols.items() if k in df.columns}
    result = df[list(available.keys())].copy()
    result.columns = list(available.values())

    # Sort: Undetermined first, then Applicable, then assumed/confirmed N/A, then Not Assessed
    status_order = {"Applicability Undetermined": 0, "Applicable": 1, "Assumed Not Applicable": 2, "Not Applicable": 3, "Not Assessed": 4}
    result["_sort"] = result["Status"].map(status_order).fillna(4)
    result = result.sort_values(["Entity ID", "_sort"]).drop(columns=["_sort"])

    return result


def build_review_queue_df(transformed_df: pd.DataFrame) -> pd.DataFrame:
    """Build redesigned Review Queue including both defaults and evaluated-no-evidence."""
    mask = transformed_df["method"].isin(["no_evidence_all_candidates", "evaluated_no_evidence"])
    df = transformed_df[mask].copy()

    if df.empty:
        return df

    # Review Type column
    def derive_review_type(method):
        if method == "no_evidence_all_candidates":
            return "Determine Applicability — all candidate L2s populated, team decides which apply"
        if method == "evaluated_no_evidence":
            return "Assumed Not Applicable — no evidence found, override if relevant"
        return ""

    df["Review Type"] = df["method"].apply(derive_review_type)

    df["Decision Basis"] = df.apply(_derive_decision_basis, axis=1)

    review_cols = [
        "entity_id", "Review Type", "new_l1", "new_l2", "Decision Basis",
        "source_legacy_pillar", "source_risk_rating_raw", "source_rationale",
        "sub_risk_evidence",
    ]
    available = [c for c in review_cols if c in df.columns]
    result = df[available].copy()

    col_rename = {
        "entity_id": "Entity ID", "new_l1": "New L1", "new_l2": "New L2",
        "source_legacy_pillar": "Legacy Source", "source_risk_rating_raw": "Source Rating",
        "source_rationale": "Source Rationale", "sub_risk_evidence": "Sub-Risk Evidence",
    }
    result = result.rename(columns=col_rename)

    # Sort: Determine Applicability first, then Confirm Not Applicable
    result = result.sort_values(["Review Type", "Entity ID"])

    return result


def _find_header_column(ws, header_name: str) -> int | None:
    """Find the 1-based column index of a header by name, or None."""
    for cell in ws[1]:
        if cell.value == header_name:
            return cell.column
    return None


def _color_rows_by_column(ws, col_index: int, value_to_fill: dict,
                          match_contains: bool = False):
    """Color entire rows based on the value in a specific column.

    If match_contains is True, checks if any key is a substring of the cell value.
    """
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        raw_val = row[col_index - 1].value
        cell_val = str(raw_val or "")
        fill = None
        if match_contains:
            for key, f in value_to_fill.items():
                if key in cell_val:
                    fill = f
                    break
        else:
            # Try raw value first (for bool/int keys), then stringified
            fill = value_to_fill.get(raw_val) or value_to_fill.get(cell_val)
        if fill:
            for cell in row:
                cell.fill = fill


def style_header(ws, max_col: int):
    """Apply header styling to first row."""
    header_font = Font(bold=True, color="FFFFFF", size=10, name="Arial")
    header_fill = PatternFill("solid", fgColor="2F5496")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    for col_idx in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border


def _enrich_findings_source(
    findings_path: str,
    column_name_map: dict,
    transformed_df: pd.DataFrame,
) -> pd.DataFrame:
    """Build an enriched findings source tab showing what each finding mapped to.

    Reads the raw findings file (before filtering) and annotates each row with:
    - Disposition: what happened to this finding (mapped, filtered, unmapped)
    - Mapped L2(s): which L2 risk(s) this finding confirmed applicability for
    """
    df = pd.read_excel(findings_path)
    df.columns = [c.strip() for c in df.columns]

    # Rename to internal names for consistency
    rename = {}
    for internal, actual in column_name_map.items():
        if actual and actual in df.columns:
            rename[actual] = internal
    df = df.rename(columns=rename)
    df["entity_id"] = df["entity_id"].astype(str).str.strip()

    # Determine disposition for each row
    dispositions = []
    mapped_l2s_col = []

    # Build a set of (entity_id, l2) pairs that were issue_confirmed in the output
    confirmed = set()
    if transformed_df is not None:
        for _, row in transformed_df.iterrows():
            if "issue_confirmed" in str(row.get("method", "")):
                confirmed.add((str(row["entity_id"]), str(row["new_l2"])))

    for _, row in df.iterrows():
        # Check approval
        approval = str(row.get("Finding Approval Status", "")).strip()
        if approval and approval != "Approved":
            dispositions.append(f"Filtered — not approved ({approval})")
            mapped_l2s_col.append("")
            continue

        # Check severity
        sev = row.get("severity")
        if pd.isna(sev) or str(sev).strip() == "":
            dispositions.append("Filtered — blank severity")
            mapped_l2s_col.append("")
            continue

        # Check L2 mapping
        raw_l2 = str(row.get("l2_risk", ""))
        if not raw_l2 or raw_l2 == "nan":
            dispositions.append("Filtered — blank L2 risk category")
            mapped_l2s_col.append("")
            continue

        # Normalize and check each L2 value (could be multi-value)
        l2_parts = raw_l2.replace("\r\n", "\n").replace("\r", "\n").split("\n")
        mapped = []
        unmapped = []
        for part in l2_parts:
            normalized = normalize_l2_name(part.strip())
            if normalized:
                eid = str(row["entity_id"])
                if (eid, normalized) in confirmed:
                    mapped.append(normalized)
                else:
                    mapped.append(f"{normalized} (not active/applicable)")
            elif part.strip():
                unmapped.append(part.strip())

        if mapped:
            dispositions.append("Included")
            mapped_l2s_col.append("; ".join(mapped))
        elif unmapped:
            dispositions.append(f"Filtered — unmappable L2 ({'; '.join(unmapped)})")
            mapped_l2s_col.append("")
        else:
            dispositions.append("Filtered — L2 not resolved")
            mapped_l2s_col.append("")

    df["Disposition"] = dispositions
    df["Mapped To L2(s)"] = mapped_l2s_col

    return df


def _enrich_sub_risks_source(
    sub_risks_df: pd.DataFrame,
    transformed_df: pd.DataFrame,
) -> pd.DataFrame:
    """Build an enriched sub-risks source tab showing what each sub-risk contributed to.

    Annotates each row with which L2(s) it provided keyword evidence for.
    """
    if sub_risks_df is None or sub_risks_df.empty:
        return pd.DataFrame()

    df = sub_risks_df.copy()

    contributions = []
    for _, row in df.iterrows():
        eid = str(row.get("entity_id", ""))
        desc = str(row.get("risk_description", "")).lower()
        l1 = str(row.get("legacy_l1", ""))

        if not desc or desc == "nan":
            contributions.append("No description text")
            continue

        # Check which L2 keywords match this description
        matched_l2s = []
        for l2_name, keywords in KEYWORD_MAP.items():
            hits = [kw for kw in keywords if kw in desc]
            if hits:
                matched_l2s.append(f"{l2_name} ({', '.join(hits[:3])})")

        if matched_l2s:
            contributions.append("; ".join(matched_l2s))
        else:
            contributions.append("No keyword matches — did not contribute to any L2 mapping")

    df["Contributed To (keyword matches)"] = contributions

    return df


def export_results(
    transformed_df: pd.DataFrame,
    overlays_df: pd.DataFrame,
    legacy_df: pd.DataFrame,
    output_path: str,
    findings_df: pd.DataFrame = None,
    sub_risks_df: pd.DataFrame = None,
    findings_path: str = None,
    findings_cols: dict = None,
    entity_id_col: str = "Audit Entity",
):
    """Write multi-sheet Excel output."""
    logger.info(f"Writing output to {output_path}")

    # --- Sheet 1: Transformed Upload ---
    upload_cols = [
        "composite_key", "entity_id", "new_l1", "new_l2",
        "inherent_risk_rating_label", "overall_impact",
        "likelihood", "impact_financial", "impact_reputational",
        "impact_consumer_harm", "impact_regulatory",
        "iag_control_effectiveness", "aligned_assurance_rating",
        "management_awareness_rating",
    ]
    upload_df = transformed_df[upload_cols].copy()
    upload_df.columns = [
        "Risk-Entity Key", "Entity ID", "L1 Risk Pillar", "L2 Risk",
        "Inherent Risk Rating", "Overall Impact",
        "Likelihood", "Impact - Financial", "Impact - Reputational",
        "Impact - Consumer Harm", "Impact - Regulatory",
        "IAG Control Effectiveness", "Aligned Assurance Rating",
        "Management Awareness Rating",
    ]

    # --- Sheet 2: Audit Review (auditor-facing) ---
    audit_df = build_audit_review_df(transformed_df, legacy_df, entity_id_col)

    # --- Sheet 3: Review Queue (redesigned) ---
    review_df = build_review_queue_df(transformed_df)

    # --- Sheet 4: Side-by-side (debugging) ---
    trace_cols = [
        "composite_key", "entity_id", "new_l1", "new_l2",
        "inherent_risk_rating", "inherent_risk_rating_label", "overall_impact",
        "likelihood", "impact_financial", "impact_reputational",
        "impact_consumer_harm", "impact_regulatory",
        "iag_control_effectiveness", "aligned_assurance_rating",
        "management_awareness_rating",
        "source_legacy_pillar", "source_risk_rating_raw", "source_rationale",
        "source_control_raw", "source_control_rationale",
        "mapping_type", "confidence", "method",
        "dims_parsed_from_rationale", "sub_risk_evidence", "needs_review",
        "control_flag", "app_flag", "aux_flag",
        "overlay_flag", "overlay_source", "overlay_rating", "overlay_rationale",
    ]
    available_trace_cols = [c for c in trace_cols if c in transformed_df.columns]
    trace_df = transformed_df[available_trace_cols].copy()

    # --- Sheet 5: Legacy original (written as-is, no copy needed) ---

    # --- Sheet 6: Overlay flags ---
    overlay_out = overlays_df.copy() if not overlays_df.empty else pd.DataFrame()

    # Build Methodology tab
    methodology_data = [
        ["Risk Taxonomy Transformer — Methodology & Legend", ""],
        ["", ""],
        ["PURPOSE", ""],
        ["This workbook maps legacy 14-pillar risk assessments to the new 23 L2 risk taxonomy.", ""],
        ["For each audit entity, the tool determines which L2 risks are applicable and carries", ""],
        ["forward inherent risk ratings and control assessments as a starting point for teams.", ""],
        ["", ""],
        ["STATUS VALUES", ""],
        ["Status", "Meaning"],
        ["Applicable", "Evidence confirms this L2 risk applies to the entity. Sources include keyword matches in rationale text, sub-risk descriptions, and open findings."],
        ["Applicability Undetermined", "Could not determine applicability from available data. All candidate L2s from the legacy pillar are populated with the legacy rating. Team must decide which apply and mark the rest N/A."],
        ["Assumed Not Applicable", "Other L2s from the same legacy pillar had keyword evidence, but this one did not. Assumed not applicable — override if this L2 is relevant to the entity."],
        ["Not Applicable", "The legacy pillar was explicitly rated Not Applicable. Carried forward with high confidence."],
        ["Not Assessed", "No legacy pillar maps to this L2 risk. This is a structural gap in the crosswalk, not a team decision."],
        ["", ""],
        ["CONFIDENCE LEVELS", ""],
        ["Level", "Meaning"],
        ["high", "3+ keyword matches from rationale/sub-risks, or direct 1:1 mapping, or legacy source explicitly N/A, or confirmed by open finding."],
        ["medium", "1-2 keyword matches. Likely correct but should be verified."],
        ["low", "Zero keyword matches. All candidates populated for team review."],
        ["none", "L2 was evaluated as a candidate but had no evidence (Assumed Not Applicable rows)."],
        ["", ""],
        ["EVIDENCE SOURCES (in priority order)", ""],
        ["Source", "How it's used"],
        ["Open Findings", "If an approved finding with severity is tagged to an L2 risk for this entity, that L2 is confirmed applicable regardless of keyword matching."],
        ["Sub-Risk Descriptions", "Key Risk Descriptions tagged to the entity's legacy pillar are scored against L2 keyword lists. Evidence trail shows sub-risk ID and matched keywords."],
        ["Pillar Rationale Text", "Inherent Risk Rationale text from the legacy assessment is scored against L2 keyword lists. Evidence trail shows matched keywords."],
        ["Application/Engagement Tags", "If IT applications or third party engagements are tagged to the entity, Technology, Data, Information and Cyber Security, and/or Third Party are flagged in Additional Signals."],
        ["Auxiliary Risk Dimensions", "If L2 risks appear in the entity's AXP or AENB Auxiliary Risk Dimensions columns, they are flagged in Additional Signals."],
        ["", ""],
        ["ADDITIONAL SIGNALS COLUMN", ""],
        ["Signal", "Meaning"],
        ["Control Flag", "A control contradiction was detected — e.g., the control is rated Well Controlled but there is an open High or Critical finding for this L2."],
        ["Application / Engagement Flag", "IT applications or third party engagements are tagged to this entity. Distinguishes primary (tested here) vs secondary (used but tested elsewhere)."],
        ["Auxiliary Risk Flag", "This L2 was listed as an auxiliary risk in the entity's legacy data."],
        ["", ""],
        ["RATING SOURCE COLUMN", ""],
        ["Value", "Meaning"],
        ["Carried from legacy pillar rating", "The legacy pillar had a single composite rating (e.g., High). All five inherent risk dimensions are set to this value as a starting point."],
        ["Parsed from rationale", "The rationale text contained explicit likelihood and/or impact statements (e.g., 'Likelihood is low, impact is high'). Dimensions are set individually."],
        ["Control: All 3 set from", "The three control columns (IAG Control Effectiveness, Aligned Assurance Rating, Management Awareness Rating) are all set to the same legacy control assessment value."],
        ["", ""],
        ["TABS IN THIS WORKBOOK", ""],
        ["Tab", "Purpose"],
        ["Methodology", "This tab — explains the tool's approach, status values, and column definitions."],
        ["Transformed_Upload", "Final output formatted for upload to the target system."],
        ["Audit_Review", "Auditor-facing view with status, decision basis, rating source, and additional signals. Primary review tab."],
        ["Review_Queue", "Filtered view of rows requiring team action (Applicability Undetermined and Assumed Not Applicable)."],
        ["Side_by_Side", "Full traceability — every column including method, confidence, individual flags, and overlay data. For debugging and audit trail."],
        ["Legacy_Original", "Unmodified legacy risk data as ingested."],
        ["Findings_Source", "All findings from the source file with Disposition (included/filtered/reason) and which L2(s) each finding mapped to."],
        ["Sub_Risks_Source", "All sub-risk descriptions with which L2(s) each contributed keyword evidence to."],
        ["Overlay_Flags", "Country risk overlay flags showing which L2s are amplified by country risk."],
        ["", ""],
        ["FINDING FILTERS APPLIED", ""],
        ["Filter", "Rule"],
        ["Approval Status", "Only findings with Finding Approval Status = 'Approved' are included."],
        ["Blank Severity", "Findings with blank or missing severity are excluded."],
        ["Active Statuses", "Only Open, In Validation, and In Sustainability findings trigger control contradiction flags. Closed, Cancelled, and Not Started are excluded from contradiction checks."],
        ["", ""],
        ["DEDUPLICATION", ""],
        ["When multiple legacy pillars map to the same L2 for an entity, the tool keeps the", ""],
        ["row with the higher (more conservative) rating and logs both sources in the Legacy", ""],
        ["Source column. Findings-confirmed rows take priority over keyword matches.", ""],
    ]
    methodology_df = pd.DataFrame(methodology_data, columns=["Topic", "Detail"])

    # Write sheets
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        methodology_df.to_excel(writer, sheet_name="Methodology", index=False, header=False)
        upload_df.to_excel(writer, sheet_name="Transformed_Upload", index=False)
        audit_df.to_excel(writer, sheet_name="Audit_Review", index=False)
        review_df.to_excel(writer, sheet_name="Review_Queue", index=False)
        trace_df.to_excel(writer, sheet_name="Side_by_Side", index=False)
        legacy_df.to_excel(writer, sheet_name="Legacy_Original", index=False)
        if findings_path and findings_cols:
            enriched_findings = _enrich_findings_source(
                findings_path, findings_cols, transformed_df)
            enriched_findings.to_excel(writer, sheet_name="Findings_Source", index=False)
        elif findings_df is not None and not findings_df.empty:
            findings_df.to_excel(writer, sheet_name="Findings_Source", index=False)
        if sub_risks_df is not None and not sub_risks_df.empty:
            enriched_sub_risks = _enrich_sub_risks_source(sub_risks_df, transformed_df)
            enriched_sub_risks.to_excel(writer, sheet_name="Sub_Risks_Source", index=False)
        if not overlay_out.empty:
            overlay_out.to_excel(writer, sheet_name="Overlay_Flags", index=False)

    # Apply formatting
    wb = load_workbook(output_path)

    # Status color fills
    green_fill = PatternFill("solid", fgColor="C6EFCE")
    gray_fill = PatternFill("solid", fgColor="D9D9D9")
    yellow_fill = PatternFill("solid", fgColor="FFFF00")
    blue_fill = PatternFill("solid", fgColor="BDD7EE")
    orange_fill = PatternFill("solid", fgColor="FCE4D6")
    status_fills = {
        "Applicable": green_fill,
        "Not Applicable": gray_fill,
        "Assumed Not Applicable": orange_fill,
        "Applicability Undetermined": yellow_fill,
        "Not Assessed": blue_fill,
    }

    review_type_fills = {
        "Determine Applicability": yellow_fill,
        "Assumed Not Applicable": orange_fill,
    }

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        style_header(ws, ws.max_column)

        # Auto-width columns
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                except (TypeError, ValueError):
                    pass
            # Cap wider for text-heavy columns
            cap = 60 if sheet_name in ("Review_Queue", "Audit_Review") else 40
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), cap)

        # Color-code Audit_Review by Status
        if sheet_name == "Audit_Review":
            col = _find_header_column(ws, "Status")
            if col:
                _color_rows_by_column(ws, col, status_fills)

        # Color-code Review_Queue by Review Type
        if sheet_name == "Review_Queue":
            col = _find_header_column(ws, "Review Type")
            if col:
                _color_rows_by_column(ws, col, review_type_fills, match_contains=True)

        # Highlight needs_review rows in yellow on Side_by_Side
        if sheet_name == "Side_by_Side":
            col = _find_header_column(ws, "needs_review")
            if col:
                _color_rows_by_column(ws, col, {True: yellow_fill})

    # Format Methodology tab
    if "Methodology" in wb.sheetnames:
        ws = wb["Methodology"]
        bold_font = Font(bold=True, size=11, name="Arial")
        title_font = Font(bold=True, size=14, name="Arial", color="2F5496")
        sub_header_font = Font(bold=True, size=10, name="Arial", color="2F5496")
        ws.column_dimensions["A"].width = 45
        ws.column_dimensions["B"].width = 120

        # Bold section headers and title
        section_headers = {
            "PURPOSE", "STATUS VALUES", "CONFIDENCE LEVELS",
            "EVIDENCE SOURCES (in priority order)", "ADDITIONAL SIGNALS COLUMN",
            "RATING SOURCE COLUMN", "TABS IN THIS WORKBOOK",
            "FINDING FILTERS APPLIED", "DEDUPLICATION",
        }
        sub_headers = {"Status", "Level", "Source", "Signal", "Value", "Tab", "Filter"}

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            cell_val = str(row[0].value or "")
            if cell_val.startswith("Risk Taxonomy Transformer"):
                row[0].font = title_font
            elif cell_val in section_headers:
                row[0].font = bold_font
            elif cell_val in sub_headers:
                row[0].font = sub_header_font
                row[1].font = sub_header_font

    wb.save(output_path)
    logger.info(f"  Output saved: {output_path}")
    logger.info(f"  Sheets: {wb.sheetnames}")


# =============================================================================
# SECTION 6: MAIN — CONFIGURE AND RUN
# =============================================================================

def main():
    # -------------------------------------------------------------------------
    # CONFIGURE THESE PATHS AND COLUMN NAMES
    # -------------------------------------------------------------------------
    input_dir = _PROJECT_ROOT / "data" / "input"
    output_dir = _PROJECT_ROOT / "data" / "output"

    # Find the most recent legacy data file (filename includes variable datetime)
    legacy_files = sorted(input_dir.glob("legacy_risk_data_*.xlsx"), key=lambda f: f.stat().st_mtime)
    if not legacy_files:
        raise FileNotFoundError(f"No legacy_risk_data_*.xlsx found in {input_dir}")
    legacy_data_path = str(legacy_files[-1])  # most recent
    logger.info(f"Using legacy data file: {legacy_data_path}")

    crosswalk_path = None                             # Set path or None to use YAML config
    timestamp = datetime.now().strftime("%m%d%Y%I%M%p")
    output_path = str(output_dir / f"transformed_risk_taxonomy_{timestamp}.xlsx")
    entity_id_col = "Audit Entity ID"  # Change to "Audit Entity" for production data

    # Sub-risk descriptions file (optional but recommended for accuracy)
    # Set to None to skip sub-risk lookup and use keyword matching only.
    sub_risk_files = sorted(input_dir.glob("sub_risk_descriptions_*.xlsx"), key=lambda f: f.stat().st_mtime)
    sub_risk_path = str(sub_risk_files[-1]) if sub_risk_files else None
    if sub_risk_path:
        logger.info(f"Using sub-risk file: {sub_risk_path}")
    else:
        logger.info("No sub_risk_descriptions_*.xlsx found — skipping sub-risk lookup")
    sub_risk_cols = {
        "entity_id": "Audit Entity ID",
        "risk_id": "Key Risk ID",
        "risk_desc": "Key Risk Description",
        "legacy_l1": "Level 1 Risk Category",
        "rating": "Inherent Risk Rating",
    }
    # Note: "Key Risk Title" is also available but not currently used

    # LLM Override file (optional — produced by batching Review Queue through LLM)
    # Set to None on first run. After LLM classification, point to the output file.
    # Expected columns: entity_id, source_legacy_pillar, classified_l2, llm_confidence
    override_path = None  # e.g., str(input_dir / "llm_overrides.xlsx")

    # Findings/Issues file (optional — confirms L2 applicability and flags control contradictions)
    # Set to None to skip findings integration.
    findings_files = sorted(input_dir.glob("findings_data_*.xlsx"), key=lambda f: f.stat().st_mtime)
    findings_path = str(findings_files[-1]) if findings_files else None
    if findings_path:
        logger.info(f"Using findings file: {findings_path}")
    else:
        logger.info("No findings_data_*.xlsx found — skipping findings integration")
    findings_cols = {
        "entity_id": "Audit Entity ID",
        "issue_id": "Finding ID",
        "l2_risk": "Risk Dimension Categories",
        "severity": "Final Reportable Finding Risk Rating",
        "status": "Finding Status",
        "issue_title": "Finding Name",
        "remediation_date": "Actual Remediation Date",
    }
    # Also available but not currently used:
    # "Audit Leader", "Finding Approval Status", "Finding Description",
    # "Audit Engagement Name", "Source"

    # Legacy pillar column names — fixed set, these pillars are historical
    def _pillar(name):
        return {
            "rating":            f"{name} Inherent Risk",
            "rationale":         f"{name} Inherent Risk Rationale",
            "control":           f"{name} Control Assessment",
            "control_rationale": f"{name} Control Assessment Rationale",
        }

    def _pillar_no_rationale(name):
        return {
            "rating":            f"{name} Inherent Risk",
            "rationale":         None,
            "control":           f"{name} Control Assessment",
            "control_rationale": None,
        }

    pillar_columns = {
        "Credit":               _pillar("Credit"),
        "Market":               _pillar("Market"),
        "Strategic & Business":  _pillar("Strategic & Business"),
        "Funding & Liquidity":   _pillar("Funding & Liquidity"),
        "Reputational":         _pillar("Reputational"),
        "Model":                _pillar("Model"),
        "Financial Reporting":  _pillar("Financial Reporting"),
        "External Fraud":       _pillar("External Fraud"),
        "Operational":          _pillar("Operational"),
        "Compliance":           _pillar("Compliance"),
        "Country":              _pillar("Country"),
        "Information Technology": _pillar_no_rationale("Information Technology"),
        "Information Security":   _pillar_no_rationale("Information Security"),
        "Third Party":            _pillar_no_rationale("Third Party"),
    }

    # Application/engagement columns used by flag_application_applicability()
    # to flag Technology, Data, InfoSec, and Third Party as potentially applicable

    # -------------------------------------------------------------------------
    # RUN
    # -------------------------------------------------------------------------
    crosswalk = ingest_crosswalk(crosswalk_path)
    legacy_df = ingest_legacy_data(legacy_data_path)

    # Load sub-risk descriptions if configured
    sub_risk_index = None
    sub_risks_df = None
    if sub_risk_path:
        sub_risks_df = ingest_sub_risks(
            sub_risk_path,
            entity_id_col=sub_risk_cols["entity_id"],
            legacy_l1_col=sub_risk_cols["legacy_l1"],
            risk_desc_col=sub_risk_cols["risk_desc"],
            risk_id_col=sub_risk_cols.get("risk_id"),
            rating_col=sub_risk_cols.get("rating"),
        )
        sub_risk_index = build_sub_risk_index(sub_risks_df)
        logger.info(f"  Sub-risk index built: {len(sub_risk_index)} entities with sub-risks")

    # Load LLM overrides if configured
    overrides = None
    if override_path is not None:
        overrides = load_overrides(override_path)
        logger.info(f"  Override index built: {len(overrides)} entity-pillar overrides")

    # Load findings if configured
    findings_index = None
    findings_df = None
    if findings_path is not None:
        findings_df = ingest_findings(findings_path, findings_cols)
        findings_index = build_findings_index(findings_df)

    ctx = TransformContext(
        crosswalk=crosswalk,
        pillar_columns=pillar_columns,
        sub_risk_index=sub_risk_index,
        overrides=overrides,
        findings_index=findings_index,
    )

    transformed_df, overlays_df = run_pipeline(legacy_df, entity_id_col, ctx)

    transformed_df = apply_overlay_flags(transformed_df, overlays_df)
    transformed_df = flag_control_contradictions(transformed_df, findings_index)
    transformed_df = flag_application_applicability(transformed_df, legacy_df, entity_id_col)
    transformed_df = flag_auxiliary_risks(transformed_df, legacy_df, entity_id_col)
    transformed_df = derive_inherent_risk_rating(transformed_df)

    export_results(
        transformed_df, overlays_df, legacy_df, output_path,
        findings_df=findings_df,
        sub_risks_df=sub_risks_df,
        findings_path=findings_path,
        findings_cols=findings_cols,
        entity_id_col=entity_id_col,
    )

    print(f"\nDone! Output: {output_path}")
    print(f"Applicability undetermined: {transformed_df['needs_review'].sum()} items require team decision")


if __name__ == "__main__":
    main()
