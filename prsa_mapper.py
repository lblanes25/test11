"""
PRSA-to-L2 Risk Mapper
======================
Maps PRSA (Process Risk Self Assessment) issues to new L2 risk categories
using spaCy semantic similarity (en_core_web_md word vectors). Each issue's
text (Issue Description + Control Title + Process Title) is compared against
each L2 definition.

Each issue can map to multiple L2s when the text legitimately spans more
than one risk category. Raw scores are replaced with plain-language
mapping statuses (Suggested Match / Needs Review / No Match).

Usage:
    python prsa_mapper.py

Input:
    - data/input/L2_Risk_Taxonomy.xlsx (L2 definitions)
    - data/input/prsa_report_*.xlsx (AE ID, Issue ID, Issue Description,
      Control Title, Process Title)

Output:
    - data/output/prsa_mapping_{timestamp}.xlsx
"""

import pandas as pd
import numpy as np
import logging
from datetime import datetime
from pathlib import Path
import spacy
import yaml

from risk_taxonomy_transformer.config import L2_TO_L1
from risk_taxonomy_transformer.normalization import normalize_l2_name

_PROJECT_ROOT = Path(__file__).parent

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(_PROJECT_ROOT / "logs" / "prsa_mapping_log.txt", mode="w"),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# =============================================================================
# CONFIGURATION (loaded from taxonomy_config.yaml)
# =============================================================================

_CONFIG_PATH = _PROJECT_ROOT / "config" / "taxonomy_config.yaml"
with open(_CONFIG_PATH, "r", encoding="utf-8") as _f:
    _cfg = yaml.safe_load(_f)

_prsa_cfg = _cfg.get("columns", {}).get("prsa_mapper", {})

# spaCy model — medium model has 300-dimensional word vectors
SPACY_MODEL = _prsa_cfg.get("spacy_model", "en_core_web_md")

# Margin threshold between top matches; None = auto-detect from data
AMBIGUITY_MARGIN_THRESHOLD = None

# Minimum similarity score for a match to be considered valid
MIN_SIMILARITY_SCORE = _prsa_cfg.get("min_similarity_score", 0.50)

# High similarity threshold for "Strong" confidence band
HIGH_SIMILARITY_SCORE = _prsa_cfg.get("high_similarity_score", 0.75)

# PRSA file pattern
PRSA_FILE_PATTERN = _prsa_cfg.get("prsa_file_pattern", "prsa_report_*.xlsx")

# PRSA column names
PRSA_ID_COL = _prsa_cfg.get("issue_id", "Issue ID")
PRSA_ENTITY_COL = _prsa_cfg.get("ae_id", "AE ID")
PRSA_TITLE_COL = _prsa_cfg.get("issue_title", "Issue Title")
PRSA_DESC_COL = _prsa_cfg.get("issue_description", "Issue Description")
PRSA_CONTROL_COL = _prsa_cfg.get("control_title", "Control Title")
PRSA_PROCESS_COL = _prsa_cfg.get("process_title", "Process Title")
PRSA_RATING_COL = _prsa_cfg.get("issue_rating", "Issue Rating")
PRSA_STATUS_COL = _prsa_cfg.get("issue_status", "Issue Status")

# L2 taxonomy file
L2_TAXONOMY_FILE = _prsa_cfg.get("l2_taxonomy_file", "L2_Risk_Taxonomy.xlsx")


# =============================================================================
# CORE FUNCTIONS
# =============================================================================

def load_l2_definitions(input_dir: Path) -> pd.DataFrame:
    """Load L2 risk definitions from taxonomy file."""
    filepath = input_dir / L2_TAXONOMY_FILE
    logger.info(f"Loading L2 definitions from {filepath}")
    df = pd.read_excel(filepath)
    logger.info(f"  Loaded {len(df)} L2 definitions")
    return df


def load_prsa_data(input_dir: Path) -> pd.DataFrame:
    """Load PRSA data from the most recent matching file."""
    prsa_files = sorted(input_dir.glob(PRSA_FILE_PATTERN),
                        key=lambda f: f.stat().st_mtime)
    if not prsa_files:
        raise FileNotFoundError(
            f"No files matching '{PRSA_FILE_PATTERN}' found in {input_dir}")

    filepath = prsa_files[-1]
    logger.info(f"Loading PRSA data from {filepath}")
    df = pd.read_excel(filepath)
    df.columns = [c.strip() for c in df.columns]

    required = [PRSA_ID_COL, PRSA_DESC_COL]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"PRSA file missing required columns: {missing}. "
                         f"Found: {list(df.columns)}")

    pre_count = len(df)

    # Clean text columns
    for col in [PRSA_ID_COL, PRSA_TITLE_COL, PRSA_DESC_COL,
                PRSA_CONTROL_COL, PRSA_PROCESS_COL]:
        if col in df.columns:
            df[col] = df[col].astype(str).fillna("").str.strip()
    if PRSA_STATUS_COL in df.columns:
        df[PRSA_STATUS_COL] = df[PRSA_STATUS_COL].astype(str).fillna("").str.strip()
    if PRSA_RATING_COL in df.columns:
        df[PRSA_RATING_COL] = df[PRSA_RATING_COL].astype(str).fillna("").str.strip()

    # Exclude closed issues — no need to map issues no longer actionable
    _CLOSED_STATUSES = {"closed", "canceled", "cancelled"}
    if PRSA_STATUS_COL in df.columns:
        closed_mask = df[PRSA_STATUS_COL].str.lower().isin(_CLOSED_STATUSES)
        if closed_mask.any():
            logger.info(f"  Excluded {closed_mask.sum()} closed PRSA issues")
            df = df[~closed_mask]

    # Drop rows with blank Issue ID
    df = df[~df[PRSA_ID_COL].isin(["", "nan"])]

    # Drop rows with blank entity — can't place without an AE
    if PRSA_ENTITY_COL in df.columns:
        df[PRSA_ENTITY_COL] = df[PRSA_ENTITY_COL].astype(str).str.strip()
        no_entity = df[PRSA_ENTITY_COL].isin(["", "nan"])
        if no_entity.any():
            logger.info(f"  Dropped {no_entity.sum()} PRSA rows with blank AE ID")
            df = df[~no_entity]
    else:
        logger.warning(f"  Column '{PRSA_ENTITY_COL}' not found — cannot filter by entity")

    # Build combined text: description + control title + process title
    def _combine(row):
        parts = []
        for col in [PRSA_DESC_COL, PRSA_CONTROL_COL, PRSA_PROCESS_COL]:
            if col in df.columns:
                val = str(row.get(col, "")).strip()
                if val and val.lower() not in ("", "nan", "none"):
                    parts.append(val)
        return ". ".join(parts)

    df["_combined_text"] = df.apply(_combine, axis=1)
    # Drop rows whose combined text is empty after concat
    blank_text = df["_combined_text"].str.strip() == ""
    if blank_text.any():
        logger.info(f"  Dropped {blank_text.sum()} PRSA rows with no text content")
        df = df[~blank_text]

    # Deduplicate by (AE, Issue ID) so we don't double-map the same issue
    # when a PRSA report lists one issue under multiple PRSA controls
    pre_dedup = len(df)
    df = df.drop_duplicates(subset=[PRSA_ENTITY_COL, PRSA_ID_COL], keep="first")
    if len(df) < pre_dedup:
        logger.info(f"  Deduplicated {pre_dedup} -> {len(df)} issue rows "
                    f"(one per AE+Issue ID)")

    logger.info(f"  Loaded {len(df)} PRSA issues with text content "
                f"(of {pre_count} total rows)")
    return df


def build_reference_vectors(
    nlp: spacy.language.Language,
    l2_df: pd.DataFrame,
) -> tuple[np.ndarray, list[str], list[str]]:
    """Build document vectors for L2 risk definitions.

    Uses L2 name + definition for richer semantic representation.
    Returns (vectors array, l2_names list, l2_definitions list).
    """
    # Filter to evaluated L2s only. Not-assessed L2s (Earnings,
    # Reputation, Country after 2026-04-21 Matt update) can remain in
    # the xlsx file; they're excluded here so they don't compete in
    # top-3 ranking and hijack events that should land on real L2s.
    _evaluated = set(L2_TO_L1.keys())
    def _is_evaluated(raw):
        canonical = normalize_l2_name(raw)
        return canonical in _evaluated
    before = len(l2_df)
    l2_df = l2_df[l2_df["L2"].apply(lambda x: _is_evaluated(str(x)))]
    if len(l2_df) < before:
        excluded = before - len(l2_df)
        logger.info(f"  Filtered out {excluded} rows with not-assessed L2s")

    # Aggregate by L2 name. Enterprise L2_Risk_Taxonomy files are often
    # at L4 grain (one row per L4 leaf) with L2 + L2 Definition repeated
    # across rows. Building one vector per row would produce duplicate
    # vectors and tied top-3 rankings; aggregating pulls L3/L4 text into
    # each L2's semantic vector and ensures one unique vector per L2.
    def _clean(v):
        s = str(v if v is not None else "").strip()
        return "" if s.lower() in ("nan", "none") else s

    sub_cols = [c for c in ["L3", "L3 Definition", "L4", "L4 Definition"]
                if c in l2_df.columns]

    l2_aggregate = {}  # l2_name -> {"def": str, "text_parts": list[str]}
    for _, row in l2_df.iterrows():
        l2_name = _clean(row.get("L2"))
        if not l2_name:
            continue
        if l2_name not in l2_aggregate:
            l2_aggregate[l2_name] = {"def": "", "text_parts": [l2_name]}
            l2_def = _clean(row.get("L2 Definition"))
            if l2_def:
                l2_aggregate[l2_name]["def"] = l2_def
                l2_aggregate[l2_name]["text_parts"].append(l2_def)
        # Fold L3/L4 text into this L2's vector for richer signal.
        for col in sub_cols:
            val = _clean(row.get(col))
            if val and val not in l2_aggregate[l2_name]["text_parts"]:
                l2_aggregate[l2_name]["text_parts"].append(val)

    l2_names = list(l2_aggregate.keys())
    l2_definitions = [l2_aggregate[n]["def"] for n in l2_names]
    l2_texts = [". ".join(l2_aggregate[n]["text_parts"]) for n in l2_names]

    logger.info(
        f"Computing vectors for {len(l2_texts)} unique L2s "
        f"(aggregated from {len(l2_df)} rows)..."
    )
    vectors = []
    for text in l2_texts:
        doc = nlp(text)
        vectors.append(doc.vector)
    vectors = np.array(vectors)

    # Normalize to unit vectors for cosine similarity via dot product
    norms = np.linalg.norm(vectors, axis=1, keepdims=True)
    norms[norms == 0] = 1
    vectors = vectors / norms

    logger.info(f"  Reference vectors shape: {vectors.shape}")
    return vectors, l2_names, l2_definitions


def compute_mappings(
    nlp: spacy.language.Language,
    prsa_df: pd.DataFrame,
    ref_vectors: np.ndarray,
    l2_names: list[str],
    l2_definitions: list[str],
) -> pd.DataFrame:
    """Compute semantic similarity and produce top-3 mappings per PRSA issue."""
    total = len(prsa_df)
    logger.info(f"Computing vectors for {total} PRSA issues...")

    results = []
    log_interval = max(1, total // 10)

    for i, (_, issue_row) in enumerate(prsa_df.iterrows()):
        if i > 0 and i % log_interval == 0:
            logger.info(f"  Processed {i}/{total} PRSA issues ({i/total*100:.0f}%)")

        combined = str(issue_row["_combined_text"])
        doc = nlp(combined)
        issue_vector = doc.vector

        norm = np.linalg.norm(issue_vector)
        if norm > 0:
            issue_vector = issue_vector / norm

        # Cosine similarity via dot product (both vectors are unit-normalized)
        scores = ref_vectors @ issue_vector

        top_indices = np.argsort(scores)[::-1][:3]

        top1_idx = top_indices[0]
        top2_idx = top_indices[1]
        top3_idx = top_indices[2]

        top1_score = float(scores[top1_idx])
        top2_score = float(scores[top2_idx])
        top3_score = float(scores[top3_idx])

        margin_1_2 = top1_score - top2_score
        margin_2_3 = top2_score - top3_score

        full_desc = str(issue_row.get(PRSA_DESC_COL, ""))
        full_desc = "" if full_desc == "nan" else full_desc

        rating = str(issue_row.get(PRSA_RATING_COL, "")) if PRSA_RATING_COL in issue_row.index else ""
        rating = "" if rating in ("", "nan", "none") else rating

        status = str(issue_row.get(PRSA_STATUS_COL, "")) if PRSA_STATUS_COL in issue_row.index else ""
        status = "" if status in ("", "nan", "none") else status

        results.append({
            "Issue ID": issue_row[PRSA_ID_COL],
            "AE ID": issue_row.get(PRSA_ENTITY_COL, ""),
            "Issue Title": issue_row.get(PRSA_TITLE_COL, ""),
            "Issue Description": full_desc[:200],
            "Issue Description Full": full_desc,
            "Issue Rating": rating,
            "Issue Status": status,
            "Match 1 - L2": l2_names[top1_idx],
            "Match 1 - Score": round(top1_score, 4),
            "Match 1 - Definition": l2_definitions[top1_idx],
            "Match 2 - L2": l2_names[top2_idx],
            "Match 2 - Score": round(top2_score, 4),
            "Match 2 - Definition": l2_definitions[top2_idx],
            "Match 3 - L2": l2_names[top3_idx],
            "Match 3 - Score": round(top3_score, 4),
            "Match 3 - Definition": l2_definitions[top3_idx],
            "Margin 1-2": round(margin_1_2, 4),
            "Margin 2-3": round(margin_2_3, 4),
            "Match 1 Valid": top1_score >= MIN_SIMILARITY_SCORE,
        })

    logger.info(f"  Computed mappings for {len(results)} PRSA issues")
    return pd.DataFrame(results)


def determine_ambiguity_threshold(mapping_df: pd.DataFrame) -> float:
    """Determine margin threshold from data distribution (P25, floored/capped)."""
    valid = mapping_df[mapping_df["Match 1 Valid"]]
    margins = valid["Margin 1-2"]
    margins = margins[margins > 0]

    if len(margins) == 0:
        return 0.02

    p25 = margins.quantile(0.25)
    median = margins.quantile(0.50)
    threshold = max(0.01, min(p25, 0.05))
    logger.info(f"  Margin distribution (valid) — P25: {p25:.4f}, median: {median:.4f}")
    logger.info(f"  Ambiguity threshold set to: {threshold:.4f}")
    return threshold


def classify_mappings(mapping_df: pd.DataFrame, threshold: float) -> pd.DataFrame:
    """Classify each PRSA issue and determine which L2s it maps to."""
    df = mapping_df.copy()

    statuses = []
    confidence_bands = []
    mapped_l2s_list = []
    mapped_l2_counts = []
    mapped_l2_defs_list = []

    for _, row in df.iterrows():
        if not row["Match 1 Valid"]:
            statuses.append("No Match")
            confidence_bands.append("Weak")
            mapped_l2s_list.append("")
            mapped_l2_counts.append(0)
            mapped_l2_defs_list.append("")
            continue

        margin_1_2 = row["Margin 1-2"]

        if margin_1_2 < threshold:
            candidates = []
            candidate_defs = []
            for n in [1, 2, 3]:
                if row[f"Match {n} - Score"] >= MIN_SIMILARITY_SCORE:
                    candidates.append(row[f"Match {n} - L2"])
                    candidate_defs.append(row[f"Match {n} - Definition"])
            statuses.append("Needs Review")
            confidence_bands.append("Review Required")
            mapped_l2s_list.append("; ".join(candidates))
            mapped_l2_counts.append(len(candidates))
            mapped_l2_defs_list.append("; ".join(candidate_defs))
        else:
            top_score = row["Match 1 - Score"]
            l2s = [row["Match 1 - L2"]]
            defs = [row["Match 1 - Definition"]]
            for n in [2, 3]:
                score = row[f"Match {n} - Score"]
                if (score >= MIN_SIMILARITY_SCORE
                        and (top_score - score) < threshold * 2):
                    l2s.append(row[f"Match {n} - L2"])
                    defs.append(row[f"Match {n} - Definition"])
            statuses.append("Suggested Match")
            if top_score >= HIGH_SIMILARITY_SCORE:
                confidence_bands.append("Strong")
            else:
                confidence_bands.append("Moderate")
            mapped_l2s_list.append("; ".join(l2s))
            mapped_l2_counts.append(len(l2s))
            mapped_l2_defs_list.append("; ".join(defs))

    df["Mapping Status"] = statuses
    df["Match Confidence"] = confidence_bands
    df["Mapped L2s"] = mapped_l2s_list
    df["Mapped L2 Count"] = mapped_l2_counts
    df["Mapped L2 Definitions"] = mapped_l2_defs_list

    return df


def export_results(
    mapping_df: pd.DataFrame,
    threshold: float,
    output_dir: Path,
) -> Path:
    """Write results to multi-sheet Excel with formatting."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    timestamp = datetime.now().strftime("%m%d%Y%I%M%p")
    output_path = output_dir / f"prsa_mapping_{timestamp}.xlsx"

    header_font = Font(bold=True, color="FFFFFF", size=10, name="Arial")
    header_fill = PatternFill("solid", fgColor="2F5496")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    wrap_align = Alignment(vertical="top", wrap_text=True)

    green_fill = PatternFill("solid", fgColor="C6EFCE")
    yellow_fill = PatternFill("solid", fgColor="FFFF00")
    gray_fill = PatternFill("solid", fgColor="D9D9D9")

    status_fills = {
        "Suggested Match": green_fill,
        "Needs Review": yellow_fill,
        "No Match": gray_fill,
    }
    confidence_fills = {
        "Strong": PatternFill("solid", fgColor="C6EFCE"),
        "Moderate": PatternFill("solid", fgColor="FCE4D6"),
        "Weak": PatternFill("solid", fgColor="D9D9D9"),
        "Review Required": PatternFill("solid", fgColor="FFFF00"),
    }

    def style_header(ws, max_col: int):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

    def auto_fit_columns(ws, overrides: dict | None = None, cap: int = 25):
        overrides = overrides or {}
        for col in ws.iter_cols(min_row=1, max_row=1):
            col_letter = get_column_letter(col[0].column)
            header_val = str(col[0].value or "")
            if header_val in overrides:
                ws.column_dimensions[col_letter].width = overrides[header_val]
            else:
                ws.column_dimensions[col_letter].width = min(
                    max(len(header_val) + 4, 12), cap
                )

    def apply_wrap(ws, columns: list[str]):
        header_map = {}
        for col in ws.iter_cols(min_row=1, max_row=1):
            header_map[str(col[0].value)] = col[0].column
        for col_name in columns:
            if col_name in header_map:
                col_idx = header_map[col_name]
                for row in range(2, ws.max_row + 1):
                    ws.cell(row=row, column=col_idx).alignment = wrap_align

    def color_column(ws, col_name: str, fills: dict):
        header_map = {}
        for col in ws.iter_cols(min_row=1, max_row=1):
            header_map[str(col[0].value)] = col[0].column
        if col_name not in header_map:
            return
        col_idx = header_map[col_name]
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_idx)
            fill = fills.get(str(cell.value))
            if fill:
                cell.fill = fill

    # Sheet 1: All Mappings
    all_cols = [
        "Issue ID", "AE ID", "Issue Title", "Issue Description",
        "Issue Rating", "Issue Status",
        "Mapping Status", "Match Confidence", "Mapped L2s", "Mapped L2 Count",
        "Mapped L2 Definitions",
    ]
    all_cols = [c for c in all_cols if c in mapping_df.columns]
    all_mappings = mapping_df[all_cols].copy()

    # Sheet 2: Needs Review
    needs_review_rows = mapping_df[mapping_df["Mapping Status"] == "Needs Review"].copy()
    review_records = []
    for _, row in needs_review_rows.iterrows():
        record = {
            "Issue ID": row["Issue ID"],
            "AE ID": row["AE ID"],
            "Issue Title": row["Issue Title"],
            "Issue Description": row["Issue Description Full"],
            "Match Confidence": row["Match Confidence"],
        }
        for n in [1, 2, 3]:
            score = row[f"Match {n} - Score"]
            if score >= MIN_SIMILARITY_SCORE:
                record[f"Candidate {n} L2"] = row[f"Match {n} - L2"]
                record[f"Candidate {n} Definition"] = row[f"Match {n} - Definition"]
            else:
                record[f"Candidate {n} L2"] = ""
                record[f"Candidate {n} Definition"] = ""
            record[f"Candidate {n} Applies"] = ""
        record["Reviewer Notes"] = ""
        review_records.append(record)
    review_df = pd.DataFrame(review_records)

    # Sheet 3: Summary
    total = len(mapping_df)
    suggested = (mapping_df["Mapping Status"] == "Suggested Match").sum()
    needs_review = (mapping_df["Mapping Status"] == "Needs Review").sum()
    no_match = (mapping_df["Mapping Status"] == "No Match").sum()

    def pct(n):
        return f"{n} ({n/total*100:.1f}%)" if total > 0 else "0"

    summary_df = pd.DataFrame({
        "Metric": ["Total PRSA Issues", "Suggested Match", "Needs Review", "No Match",
                   "", "Ambiguity Threshold", "Min Similarity Score"],
        "Value": [total, pct(suggested), pct(needs_review), pct(no_match),
                  "", f"{threshold:.4f}", MIN_SIMILARITY_SCORE],
    })

    # Sheet 4: L2 Distribution
    suggested_rows = mapping_df[mapping_df["Mapping Status"] == "Suggested Match"].copy()
    exploded_l2s = (
        suggested_rows["Mapped L2s"].str.split("; ").explode().str.strip()
    )
    exploded_l2s = exploded_l2s[exploded_l2s != ""]
    l2_dist = exploded_l2s.value_counts().reset_index()
    l2_dist.columns = ["L2 Risk", "Issue Count (Suggested Match)"]

    # Sheet 5: Raw Scores (hidden)
    raw_cols = [
        "Issue ID", "AE ID", "Issue Title", "Issue Description Full",
        "Match 1 - L2", "Match 1 - Score",
        "Match 2 - L2", "Match 2 - Score",
        "Match 3 - L2", "Match 3 - Score",
        "Margin 1-2", "Margin 2-3",
        "Mapping Status", "Match Confidence", "Match 1 Valid",
    ]
    raw_scores = mapping_df[raw_cols].copy().rename(
        columns={"Issue Description Full": "Issue Description"})

    logger.info(f"Writing output to {output_path}")
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        all_mappings.to_excel(writer, sheet_name="All Mappings", index=False)
        review_df.to_excel(writer, sheet_name="Needs Review", index=False)
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        l2_dist.to_excel(writer, sheet_name="L2 Distribution", index=False)
        raw_scores.to_excel(writer, sheet_name="Raw Scores", index=False)

        wb = writer.book

        ws_all = wb["All Mappings"]
        style_header(ws_all, ws_all.max_column)
        auto_fit_columns(ws_all, overrides={
            "Issue Description": 60, "Issue Title": 30,
            "Mapped L2s": 50, "Mapped L2 Definitions": 60,
        })
        apply_wrap(ws_all, ["Issue Description", "Mapped L2s", "Mapped L2 Definitions"])
        color_column(ws_all, "Mapping Status", status_fills)
        color_column(ws_all, "Match Confidence", confidence_fills)
        ws_all.freeze_panes = "C2"

        ws_review = wb["Needs Review"]
        style_header(ws_review, ws_review.max_column)
        auto_fit_columns(ws_review, overrides={
            "Issue Description": 60, "Issue Title": 30,
            "Candidate 1 Definition": 60, "Candidate 2 Definition": 60,
            "Candidate 3 Definition": 60,
            "Reviewer Notes": 30,
        })
        apply_wrap(ws_review, [
            "Issue Description",
            "Candidate 1 Definition", "Candidate 2 Definition", "Candidate 3 Definition",
        ])
        color_column(ws_review, "Match Confidence", confidence_fills)
        ws_review.freeze_panes = "A2"

        ws_summary = wb["Summary"]
        style_header(ws_summary, ws_summary.max_column)
        ws_summary.column_dimensions["A"].width = 40
        ws_summary.column_dimensions["B"].width = 25

        ws_l2 = wb["L2 Distribution"]
        style_header(ws_l2, ws_l2.max_column)
        auto_fit_columns(ws_l2, overrides={"L2 Risk": 45})

        ws_raw = wb["Raw Scores"]
        style_header(ws_raw, ws_raw.max_column)
        auto_fit_columns(ws_raw, overrides={"Issue Description": 60, "Issue Title": 30})
        apply_wrap(ws_raw, ["Issue Description"])
        ws_raw.sheet_state = "hidden"

    logger.info(f"  Output saved: {output_path}")
    return output_path


# =============================================================================
# MAIN
# =============================================================================

def main():
    global AMBIGUITY_MARGIN_THRESHOLD

    input_dir = _PROJECT_ROOT / "data" / "input"
    output_dir = _PROJECT_ROOT / "data" / "output"
    output_dir.mkdir(parents=True, exist_ok=True)
    (_PROJECT_ROOT / "logs").mkdir(parents=True, exist_ok=True)

    l2_df = load_l2_definitions(input_dir)
    prsa_df = load_prsa_data(input_dir)

    logger.info(f"Loading spaCy model: {SPACY_MODEL}")
    nlp = spacy.load(SPACY_MODEL)
    logger.info(f"  Model loaded ({len(nlp.vocab.vectors)} vectors, "
                f"{nlp.vocab.vectors.shape[1]} dimensions)")

    ref_vectors, l2_names, l2_definitions = build_reference_vectors(nlp, l2_df)

    mapping_df = compute_mappings(nlp, prsa_df, ref_vectors, l2_names, l2_definitions)

    if AMBIGUITY_MARGIN_THRESHOLD is None:
        AMBIGUITY_MARGIN_THRESHOLD = determine_ambiguity_threshold(mapping_df)

    mapping_df = classify_mappings(mapping_df, AMBIGUITY_MARGIN_THRESHOLD)

    total = len(mapping_df)
    suggested = (mapping_df["Mapping Status"] == "Suggested Match").sum()
    needs_review = (mapping_df["Mapping Status"] == "Needs Review").sum()
    no_match = (mapping_df["Mapping Status"] == "No Match").sum()

    logger.info("=" * 60)
    logger.info("PRSA MAPPING COMPLETE")
    logger.info(f"  Total PRSA issues: {total}")
    if total:
        logger.info(f"  Suggested Match: {suggested} ({suggested/total*100:.1f}%)")
        logger.info(f"  Needs Review: {needs_review} ({needs_review/total*100:.1f}%)")
        logger.info(f"  No Match: {no_match} ({no_match/total*100:.1f}%)")
    logger.info(f"  Ambiguity threshold: {AMBIGUITY_MARGIN_THRESHOLD:.4f}")
    logger.info("=" * 60)

    output_path = export_results(mapping_df, AMBIGUITY_MARGIN_THRESHOLD, output_dir)

    print(f"\nDone! Output: {output_path}")
    print(f"  Suggested Match: {suggested} | Needs Review: {needs_review} | No Match: {no_match}")


if __name__ == "__main__":
    main()
