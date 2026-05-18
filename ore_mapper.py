"""
ORE-to-L2 Risk Mapper
=====================
Maps Operational Risk Events (OREs) to new L2 risk categories using
spaCy semantic similarity (en_core_web_lg word vectors).

Each ORE can map to multiple L2s when the event legitimately spans more
than one risk category. Raw scores are replaced with plain-language
mapping statuses. Every item that passes the similarity floor is presented
as Needs Review (the tool does not assert a positive-confidence band);
items below the floor are No Match (excluded). Scores are retained in the
hidden Raw Scores sheet for traceability.

Usage:
    python ore_mapper.py

Input:
    - data/input/L2_Risk_Taxonomy.xlsx (L2 definitions)
    - data/input/ORE_*.xlsx (Event ID, Event Title, Event Description / Summary)

Output:
    - data/output/ore_mapping_{timestamp}.xlsx
      Sheet 1: All Mappings (one row per ORE, reviewer-friendly)
      Sheet 2: Needs Review (side-by-side comparison for ambiguous OREs)
      Sheet 3: Summary (counts + plain-language explanation)
      Sheet 4: L2 Distribution (ORE counts per L2, multi-L2 exploded)
      Sheet 5: Raw Scores (hidden — development/threshold tuning only)
"""

import argparse
import pandas as pd
import numpy as np
import logging
from datetime import datetime
from pathlib import Path
import spacy
import yaml

from risk_taxonomy_transformer.config import L2_TO_L1
from risk_taxonomy_transformer.normalization import normalize_l2_name
from risk_taxonomy_transformer.utils import log_run_provenance, spacy_model_label

_PROJECT_ROOT = Path(__file__).parent

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(_PROJECT_ROOT / "logs" / "ore_mapping_log.txt", mode="w"),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# =============================================================================
# CONFIGURATION (loaded from taxonomy_config.yaml)
# =============================================================================

_CONFIG_PATH = _PROJECT_ROOT / "config" / "taxonomy_config.yaml"
with open(_CONFIG_PATH, "r", encoding="utf-8") as _f:
    _cfg = yaml.safe_load(_f)

# Margin threshold: if the score gap between 1st and 2nd match is below this,
# the ORE is flagged as ambiguous. Set to None to auto-detect from data.
AMBIGUITY_MARGIN_THRESHOLD = None

# Minimum similarity score for a match to be considered valid
MIN_SIMILARITY_SCORE = 0.50

# Retained for Raw Scores traceability only — no longer drives a user-facing
# confidence band (all floor-passing items are uniformly Needs Review).
HIGH_SIMILARITY_SCORE = 0.75

# Source-specific config — populated by set_active_source() at startup.
SPACY_MODEL = "en_core_web_lg"
# Resolved "name version" of the actually-loaded model; set in main() after
# spacy.load. Falls back to the configured name until the model loads.
RESOLVED_SPACY_MODEL = SPACY_MODEL
SOURCE_NAME = "ore"
ORE_FILE_PATTERN = ""
ORE_ID_COL = ""
ORE_TITLE_COL = ""
ORE_DESC_COL = ""
ORE_ENTITY_COL = ""
ORE_CLASS_COL = ""
ORE_STATUS_COL = ""
ORE_RISK_L2_COL = ""        # only ore_irm — empty for ore
ORE_LEGACY_EVENT_ID_COL = ""  # only ore_irm
L2_TAXONOMY_FILE = ""
OUTPUT_FILENAME_PREFIX = "ore_mapping"


def set_active_source(name: str):
    """Bind module-level mapper config to the named source (ore | ore_irm).

    Loads the corresponding YAML block from columns.{ore_mapper, ore_irm_mapper}
    and sets ORE_*_COL, file pattern, output filename prefix.
    """
    global SOURCE_NAME, ORE_FILE_PATTERN, ORE_ID_COL, ORE_TITLE_COL
    global ORE_DESC_COL, ORE_ENTITY_COL, ORE_CLASS_COL, ORE_STATUS_COL
    global ORE_RISK_L2_COL, ORE_LEGACY_EVENT_ID_COL
    global L2_TAXONOMY_FILE, SPACY_MODEL, OUTPUT_FILENAME_PREFIX

    SOURCE_NAME = name
    if name == "ore":
        cfg = _cfg.get("columns", {}).get("ore_mapper", {})
        ORE_FILE_PATTERN = cfg.get("ore_file_pattern", "ORE_*.xlsx")
        ORE_ID_COL = cfg.get("event_id", "Event ID")
        ORE_TITLE_COL = cfg.get("event_title", "Event Title")
        ORE_DESC_COL = cfg.get("event_description", "Event Description / Summary")
        ORE_ENTITY_COL = cfg.get("entity_id", "Audit Entity (Operational Risk Events)")
        ORE_CLASS_COL = cfg.get("event_classification", "Final Event Classification")
        ORE_STATUS_COL = cfg.get("event_status", "Event Status")
        ORE_RISK_L2_COL = ""
        ORE_LEGACY_EVENT_ID_COL = ""
        OUTPUT_FILENAME_PREFIX = "ore_mapping"
    elif name == "ore_irm":
        cfg = _cfg.get("columns", {}).get("ore_irm_mapper", {})
        ORE_FILE_PATTERN = cfg.get("ore_irm_file_pattern", "ORE_IRM_*.xlsx")
        ORE_ID_COL = cfg.get("ore_id", "ORE ID")
        ORE_TITLE_COL = cfg.get("ore_title", "ORE Title")
        ORE_DESC_COL = cfg.get("ore_description", "ORE Description")
        ORE_ENTITY_COL = ""  # no AE column — AE join happens at ingestion time
        ORE_CLASS_COL = ""    # IRM has Capture Status (display-only) — no severity class
        ORE_STATUS_COL = cfg.get("capture_status", "Capture Status")
        ORE_RISK_L2_COL = cfg.get("risk_level_2", "Risk Level 2")
        ORE_LEGACY_EVENT_ID_COL = cfg.get("legacy_event_id", "Legacy Event ID")
        OUTPUT_FILENAME_PREFIX = "ore_irm_mapping"
    else:
        raise ValueError(f"Unknown source: {name!r}; expected 'ore' or 'ore_irm'")

    SPACY_MODEL = cfg.get("spacy_model", "en_core_web_lg")
    L2_TAXONOMY_FILE = cfg.get("l2_taxonomy_file", "L2_Risk_Taxonomy.xlsx")


# Default to legacy ORE source so that imports of this module before main()
# runs (e.g., tests that import constants) still see populated globals.
set_active_source("ore")


# =============================================================================
# CORE FUNCTIONS
# =============================================================================

def load_l2_definitions(input_dir: Path) -> pd.DataFrame:
    """Load L2 risk definitions from taxonomy file.

    Real enterprise files merge L1/L2/L3 cells across multiple rows; pandas
    reads the continuation rows as NaN. Forward-fill so every row has a
    populated L1/L2/L3 — otherwise the bucketing loop in
    build_reference_vectors skips continuation rows and silently drops their
    L3/L4 definitions from the per-L2 reference vector.
    """
    filepath = input_dir / L2_TAXONOMY_FILE
    logger.info(f"Loading L2 definitions from {filepath}")
    df = pd.read_excel(filepath)
    ffill_cols = [c for c in ("L1", "L2", "L3") if c in df.columns]
    if ffill_cols:
        df[ffill_cols] = df[ffill_cols].ffill()
    logger.info(f"  Loaded {len(df)} L2 definitions")
    return df


def load_ore_data(input_dir: Path) -> tuple[pd.DataFrame, pd.DataFrame, str]:
    """Load ORE data from the most recent matching file.

    Returns (df, orphans_df, source_filename) where orphans_df captures rows
    dropped for blank Audit Entity ID. Empty orphans_df for ore_irm (IRM has
    no AE column at all; AE attribution happens at ingestion time via the
    legacy_risk_data 'IRM ORE ID' bridge).
    """
    ore_files = sorted(input_dir.glob(ORE_FILE_PATTERN),
                       key=lambda f: f.stat().st_mtime)
    # The legacy ORE pattern (ORE_*.xlsx) also matches ORE_IRM_*.xlsx — filter
    # those out when running the legacy source so the IRM file doesn't shadow
    # the legacy file. The IRM mapper has its own dedicated pattern.
    if SOURCE_NAME == "ore":
        ore_files = [f for f in ore_files if not f.name.upper().startswith("ORE_IRM_")]
    if not ore_files:
        raise FileNotFoundError(
            f"No files matching '{ORE_FILE_PATTERN}' found in {input_dir}")

    filepath = ore_files[-1]
    source_filename = filepath.name
    logger.info(f"Loading ORE data from {filepath}")
    df = pd.read_excel(filepath)
    # Strip whitespace and any leading "*" prefix some ORE exports use on
    # Event Title / Event Description / Summary headers.
    df.columns = [str(c).strip().lstrip("*").strip() for c in df.columns]

    # Validate required columns
    required = [ORE_ID_COL, ORE_TITLE_COL, ORE_DESC_COL]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"ORE file missing required columns: {missing}. "
                         f"Found: {list(df.columns)}")

    pre_count = len(df)

    # Clean data
    df[ORE_ID_COL] = df[ORE_ID_COL].astype(str).str.strip()
    df[ORE_TITLE_COL] = df[ORE_TITLE_COL].astype(str).fillna("").str.strip()
    df[ORE_DESC_COL] = df[ORE_DESC_COL].astype(str).fillna("").str.strip()
    if ORE_CLASS_COL in df.columns:
        df[ORE_CLASS_COL] = df[ORE_CLASS_COL].astype(str).fillna("").str.strip()
    if ORE_STATUS_COL in df.columns:
        df[ORE_STATUS_COL] = df[ORE_STATUS_COL].astype(str).fillna("").str.strip()

    # Exclude closed OREs — no need to map events that are no longer active.
    # For IRM source: per Lu's spec, no status filter (Capture Status is
    # display-only; banner discloses).
    _CLOSED_STATUSES = {"closed", "canceled", "draft canceled", "draft expired",
                        "draft", "pending cancelation by event admin"}
    if SOURCE_NAME != "ore_irm" and ORE_STATUS_COL in df.columns:
        closed_mask = df[ORE_STATUS_COL].str.lower().isin(_CLOSED_STATUSES)
        if closed_mask.any():
            logger.info(f"  Excluded {closed_mask.sum()} closed OREs (statuses: "
                        f"{df.loc[closed_mask, ORE_STATUS_COL].unique().tolist()})")
            df = df[~closed_mask]

    # Drop rows with no meaningful text
    df = df[~((df[ORE_TITLE_COL].isin(["", "nan"])) &
              (df[ORE_DESC_COL].isin(["", "nan"])))]
    df = df[~df[ORE_ID_COL].isin(["", "nan"])]

    # Drop OREs with no Audit Entity ID — can't place in entity evidence briefs.
    # IRM source has no AE column at all; AE attribution is established at
    # ingestion time via the legacy_risk_data 'IRM ORE ID' bridge.
    orphans = pd.DataFrame()
    if SOURCE_NAME != "ore_irm":
        if ORE_ENTITY_COL in df.columns:
            df[ORE_ENTITY_COL] = df[ORE_ENTITY_COL].astype(str).str.strip()
            no_entity = df[ORE_ENTITY_COL].isin(["", "nan"])
            if no_entity.any():
                orphans = df[no_entity].copy()
                logger.info(f"  Dropped {no_entity.sum()} OREs with blank Audit Entity ID "
                            f"(captured to orphans sidecar)")
                df = df[~no_entity]
        else:
            logger.warning(f"  Column '{ORE_ENTITY_COL}' not found — cannot filter by entity")

    logger.info(f"  Loaded {len(df)} OREs with text content (of {pre_count} total rows)")
    return df, orphans, source_filename


def build_reference_vectors(
    nlp: spacy.language.Language,
    l2_df: pd.DataFrame,
) -> tuple[np.ndarray, list[str], list[str]]:
    """Build document vectors for L2 risk definitions.

    Uses L2 name + definition for richer semantic representation.
    Returns (vectors array, l2_names list, l2_definitions list).
    """
    # Aggregate by CANONICAL evaluated L2, bucketing at the finer
    # grain (L3 over L2) when L3 normalizes to an evaluated L2.
    # Rationale: enterprise files often split L2s at L3 grain -- e.g.
    # "Fraud (External and Internal)" at L2 with Internal Fraud /
    # External Fraud - First Party / - Victim Fraud at L3. We want
    # Internal Fraud and External Fraud as distinct vectors, not
    # merged under the parent L2.
    def _clean(v):
        s = str(v if v is not None else "").strip()
        return "" if s.lower() in ("nan", "none") else s

    _evaluated = set(L2_TO_L1.keys())
    has_l3 = "L3" in l2_df.columns
    # Fold child-level (L3, L4) text into the per-L2 reference vector for
    # richer semantic signal. L1 / L1 Definition are intentionally NOT
    # included: L1 is the parent category, more generic than L2, and would
    # dilute the L2's vector with broader concepts rather than sharpen it.
    # L3 and L4 narrow the L2's scope (sub-categories) — exactly what helps
    # spaCy match items more precisely.
    sub_cols = [c for c in ["L3", "L3 Definition", "L4", "L4 Definition"]
                if c in l2_df.columns]

    def _bucket_for(l2_name, l3_name):
        """Return canonical evaluated L2 for this row, or None to skip."""
        if l3_name:
            c = normalize_l2_name(l3_name)
            if c in _evaluated:
                return c
        c = normalize_l2_name(l2_name)
        if c in _evaluated:
            return c
        return None

    l2_aggregate = {}  # canonical -> {"def": str, "text_parts": list[str]}
    skipped = 0
    for _, row in l2_df.iterrows():
        l2_name = _clean(row.get("L2"))
        if not l2_name:
            skipped += 1
            continue
        l3_name = _clean(row.get("L3")) if has_l3 else ""
        bucket = _bucket_for(l2_name, l3_name)
        if bucket is None:
            skipped += 1
            continue
        if bucket not in l2_aggregate:
            l2_aggregate[bucket] = {"def": "", "text_parts": [bucket]}
            l2_def = _clean(row.get("L2 Definition"))
            if l2_def:
                l2_aggregate[bucket]["def"] = l2_def
                l2_aggregate[bucket]["text_parts"].append(l2_def)
        # Fold L3/L4 text into this bucket's vector for richer signal.
        for col in sub_cols:
            val = _clean(row.get(col))
            if val and val not in l2_aggregate[bucket]["text_parts"]:
                l2_aggregate[bucket]["text_parts"].append(val)

    if skipped:
        logger.info(f"  Skipped {skipped} rows that did not resolve to an evaluated L2")

    l2_names = list(l2_aggregate.keys())
    l2_definitions = [l2_aggregate[n]["def"] for n in l2_names]
    l2_texts = [". ".join(l2_aggregate[n]["text_parts"]) for n in l2_names]

    logger.info(
        f"Computing vectors for {len(l2_texts)} unique L2s "
        f"(aggregated from {len(l2_df)} rows)..."
    )
    vectors = []
    for text in l2_texts:
        doc = nlp(text)
        vectors.append(doc.vector)
    vectors = np.array(vectors)

    # Normalize to unit vectors for cosine similarity via dot product
    norms = np.linalg.norm(vectors, axis=1, keepdims=True)
    norms[norms == 0] = 1  # avoid division by zero
    vectors = vectors / norms

    logger.info(f"  Reference vectors shape: {vectors.shape}")
    return vectors, l2_names, l2_definitions


def compute_mappings(
    nlp: spacy.language.Language,
    ore_df: pd.DataFrame,
    ref_vectors: np.ndarray,
    l2_names: list[str],
    l2_definitions: list[str],
) -> pd.DataFrame:
    """Compute semantic similarity and produce top-3 mappings per ORE.

    Stores both truncated (200-char) and full event descriptions.
    Includes L2 definitions for each match for reviewer side-by-side comparison.
    """

    total = len(ore_df)
    logger.info(f"Computing vectors for {total} OREs...")

    results = []
    log_interval = max(1, total // 10)

    for i, (_, ore_row) in enumerate(ore_df.iterrows()):
        if i > 0 and i % log_interval == 0:
            logger.info(f"  Processed {i}/{total} OREs ({i/total*100:.0f}%)")

        # Build ORE text: title + description
        title = str(ore_row[ORE_TITLE_COL])
        desc = str(ore_row[ORE_DESC_COL])
        title = "" if title == "nan" else title
        desc = "" if desc == "nan" else desc
        combined = f"{title}. {desc}" if desc else title

        # Get document vector
        doc = nlp(combined)
        ore_vector = doc.vector

        # Normalize
        norm = np.linalg.norm(ore_vector)
        if norm > 0:
            ore_vector = ore_vector / norm

        # Cosine similarity via dot product (both vectors are unit-normalized)
        scores = ref_vectors @ ore_vector

        # Top 3
        top_indices = np.argsort(scores)[::-1][:3]

        top1_idx = top_indices[0]
        top2_idx = top_indices[1]
        top3_idx = top_indices[2]

        top1_score = float(scores[top1_idx])
        top2_score = float(scores[top2_idx])
        top3_score = float(scores[top3_idx])

        margin_1_2 = top1_score - top2_score
        margin_2_3 = top2_score - top3_score

        full_desc = str(ore_row[ORE_DESC_COL])
        full_desc = "" if full_desc == "nan" else full_desc

        # Classification is optional — may not exist in older ORE files
        cls_raw = str(ore_row.get(ORE_CLASS_COL, "")) if ORE_CLASS_COL in ore_row.index else ""
        cls_val = "" if cls_raw in ("", "nan", "none") else cls_raw

        # Event Status is optional — may not exist in older ORE files
        status_raw = str(ore_row.get(ORE_STATUS_COL, "")) if ORE_STATUS_COL in ore_row.index else ""
        status_val = "" if status_raw in ("", "nan", "none") else status_raw

        results.append({
            "Event ID": ore_row[ORE_ID_COL],
            "Audit Entity ID": (ore_row.get(ORE_ENTITY_COL, "") if ORE_ENTITY_COL else ""),
            "Event Title": ore_row[ORE_TITLE_COL],
            "Event Description": full_desc[:200],
            "Event Description Full": full_desc,
            "Final Event Classification": cls_val,
            "Event Status": status_val,
            "Match 1 - L2": l2_names[top1_idx],
            "Match 1 - Score": round(top1_score, 4),
            "Match 1 - Definition": l2_definitions[top1_idx],
            "Match 2 - L2": l2_names[top2_idx],
            "Match 2 - Score": round(top2_score, 4),
            "Match 2 - Definition": l2_definitions[top2_idx],
            "Match 3 - L2": l2_names[top3_idx],
            "Match 3 - Score": round(top3_score, 4),
            "Match 3 - Definition": l2_definitions[top3_idx],
            "Margin 1-2": round(margin_1_2, 4),
            "Margin 2-3": round(margin_2_3, 4),
            "Match 1 Valid": top1_score >= MIN_SIMILARITY_SCORE,
        })

    logger.info(f"  Computed mappings for {len(results)} OREs")
    return pd.DataFrame(results)


def determine_ambiguity_threshold(mapping_df: pd.DataFrame) -> float:
    """Determine the margin threshold from data distribution.

    Uses the 25th percentile of margins for valid matches,
    floored at 0.01 and capped at 0.05.
    """
    valid = mapping_df[mapping_df["Match 1 Valid"]]
    margins = valid["Margin 1-2"]
    margins = margins[margins > 0]

    if len(margins) == 0:
        return 0.02  # fallback

    p25 = margins.quantile(0.25)
    median = margins.quantile(0.50)

    # SpaCy scores are more compressed than TF-IDF, so tighter thresholds
    threshold = max(0.01, min(p25, 0.05))

    logger.info(f"  Margin distribution (valid matches) — P25: {p25:.4f}, median: {median:.4f}")
    logger.info(f"  Ambiguity threshold set to: {threshold:.4f}")
    return threshold


def classify_mappings(mapping_df: pd.DataFrame, threshold: float) -> pd.DataFrame:
    """Classify each ORE and determine which L2s it maps to.

    Status logic:
      - No Match: Match 1 below MIN_SIMILARITY_SCORE (excluded).
      - Needs Review: Match 1 valid (>= MIN_SIMILARITY_SCORE). Every item
        that passes the inclusion floor is presented as Needs Review — the
        tool does not assert a positive-confidence band on NLP matches.
        The similarity scores, margins, and threshold remain in the hidden
        Raw Scores sheet for traceability. The single-vs-multi L2 selection
        logic (margin to Match 2/3) is retained so reviewers see the right
        candidate set, but the user-facing band is uniformly Needs Review.

    Produces:
      - Mapping Status (Needs Review / No Match)
      - Match Confidence (Review Required / Weak)
      - Mapped L2s (semicolon-separated list)
      - Mapped L2 Count (integer)
      - Mapped L2 Definitions (semicolon-separated, matching order)
    """
    df = mapping_df.copy()

    statuses = []
    confidence_bands = []
    mapped_l2s_list = []
    mapped_l2_counts = []
    mapped_l2_defs_list = []

    for _, row in df.iterrows():
        if not row["Match 1 Valid"]:
            statuses.append("No Match")
            confidence_bands.append("Weak")
            mapped_l2s_list.append("")
            mapped_l2_counts.append(0)
            mapped_l2_defs_list.append("")
            continue

        margin_1_2 = row["Margin 1-2"]

        if margin_1_2 < threshold:
            # Needs Review — can't confidently separate top matches
            # Show all candidates above MIN_SIMILARITY_SCORE for reviewer
            candidates = []
            candidate_defs = []
            for n in [1, 2, 3]:
                if row[f"Match {n} - Score"] >= MIN_SIMILARITY_SCORE:
                    candidates.append(row[f"Match {n} - L2"])
                    candidate_defs.append(row[f"Match {n} - Definition"])
            statuses.append("Needs Review")
            confidence_bands.append("Review Required")
            mapped_l2s_list.append("; ".join(candidates))
            mapped_l2_counts.append(len(candidates))
            mapped_l2_defs_list.append("; ".join(candidate_defs))
        else:
            # Primary (Match 1) plus any Match 2/3 that qualify as additional
            # L2s: must be above MIN_SIMILARITY_SCORE AND within 2x the
            # ambiguity threshold of Match 1's score. The band is NOT asserted
            # as a positive match — every floor-passing item is Needs Review.
            top_score = row["Match 1 - Score"]
            l2s = [row["Match 1 - L2"]]
            defs = [row["Match 1 - Definition"]]
            for n in [2, 3]:
                score = row[f"Match {n} - Score"]
                if (score >= MIN_SIMILARITY_SCORE
                        and (top_score - score) < threshold * 2):
                    l2s.append(row[f"Match {n} - L2"])
                    defs.append(row[f"Match {n} - Definition"])
            statuses.append("Needs Review")
            confidence_bands.append("Review Required")
            mapped_l2s_list.append("; ".join(l2s))
            mapped_l2_counts.append(len(l2s))
            mapped_l2_defs_list.append("; ".join(defs))

    df["Mapping Status"] = statuses
    df["Match Confidence"] = confidence_bands
    df["Mapped L2s"] = mapped_l2s_list
    df["Mapped L2 Count"] = mapped_l2_counts
    df["Mapped L2 Definitions"] = mapped_l2_defs_list

    return df


def export_results(
    mapping_df: pd.DataFrame,
    threshold: float,
    output_dir: Path,
) -> Path:
    """Write results to multi-sheet Excel with formatting.

    Sheets:
      1. All Mappings — one row per ORE, reviewer-friendly (no raw scores)
      2. Needs Review — side-by-side comparison for ambiguous OREs
      3. Summary — counts and plain-language explanation
      4. L2 Distribution — ORE count per L2 (exploded for multi-L2)
      5. Raw Scores — hidden, for development and threshold tuning
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    timestamp = datetime.now().strftime("%m%d%Y%I%M%p")
    output_path = output_dir / f"{OUTPUT_FILENAME_PREFIX}_{timestamp}.xlsx"

    # -- Shared styles --
    header_font = Font(bold=True, color="FFFFFF", size=10, name="Arial")
    header_fill = PatternFill("solid", fgColor="2F5496")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    wrap_align = Alignment(vertical="top", wrap_text=True)

    green_fill = PatternFill("solid", fgColor="C6EFCE")
    yellow_fill = PatternFill("solid", fgColor="FFFF00")
    gray_fill = PatternFill("solid", fgColor="D9D9D9")
    reviewer_fill = PatternFill("solid", fgColor="E2EFDA")

    status_fills = {
        "Suggested Match": green_fill,
        "Needs Review": yellow_fill,
        "No Match": gray_fill,
    }

    def style_header(ws, max_col: int):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

    def auto_fit_columns(ws, overrides: dict | None = None, cap: int = 25):
        """Set column widths with optional overrides and a max cap."""
        overrides = overrides or {}
        for col in ws.iter_cols(min_row=1, max_row=1):
            col_letter = get_column_letter(col[0].column)
            header_val = str(col[0].value or "")
            if header_val in overrides:
                ws.column_dimensions[col_letter].width = overrides[header_val]
            else:
                ws.column_dimensions[col_letter].width = min(
                    max(len(header_val) + 4, 12), cap
                )

    def apply_wrap(ws, columns: list[str]):
        """Apply text wrap to data cells in named columns."""
        header_map = {}
        for col in ws.iter_cols(min_row=1, max_row=1):
            header_map[str(col[0].value)] = col[0].column
        for col_name in columns:
            if col_name in header_map:
                col_idx = header_map[col_name]
                for row in range(2, ws.max_row + 1):
                    ws.cell(row=row, column=col_idx).alignment = wrap_align

    confidence_fills = {
        "Strong": PatternFill("solid", fgColor="C6EFCE"),       # Green
        "Moderate": PatternFill("solid", fgColor="FCE4D6"),      # Light orange
        "Weak": PatternFill("solid", fgColor="D9D9D9"),          # Gray
        "Review Required": PatternFill("solid", fgColor="FFFF00"),  # Yellow
    }

    def color_status_column(ws, col_name: str = "Mapping Status"):
        """Apply conditional fill to Status column cells."""
        header_map = {}
        for col in ws.iter_cols(min_row=1, max_row=1):
            header_map[str(col[0].value)] = col[0].column
        if col_name not in header_map:
            return
        col_idx = header_map[col_name]
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_idx)
            fill = status_fills.get(str(cell.value))
            if fill:
                cell.fill = fill

    def color_confidence_column(ws, col_name: str = "Match Confidence"):
        """Apply conditional fill to Match Confidence column cells."""
        header_map = {}
        for col in ws.iter_cols(min_row=1, max_row=1):
            header_map[str(col[0].value)] = col[0].column
        if col_name not in header_map:
            return
        col_idx = header_map[col_name]
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_idx)
            fill = confidence_fills.get(str(cell.value))
            if fill:
                cell.fill = fill

    # =========================================================================
    # Sheet 1: All Mappings (visible) — one row per ORE, no raw scores
    # Downstream consumers (control effectiveness pipeline) explode the
    # semicolon-separated Mapped L2s column into per-L2 rows when building
    # their indexes, same pattern used for multi-value findings.
    # =========================================================================
    all_cols = [
        "Event ID", "Audit Entity ID", "Event Title", "Event Description",
        "Final Event Classification", "Event Status",
        "Mapping Status", "Match Confidence", "Mapped L2s", "Mapped L2 Count",
        "Mapped L2 Definitions",
    ]
    # Drop classification column if not present in data (optional column)
    all_cols = [c for c in all_cols if c in mapping_df.columns]
    all_mappings = mapping_df[all_cols].copy()

    # =========================================================================
    # Sheet 2: Needs Review (visible) — side-by-side comparison workspace
    # =========================================================================
    needs_review_rows = mapping_df[mapping_df["Mapping Status"] == "Needs Review"].copy()
    review_records = []
    for _, row in needs_review_rows.iterrows():
        record = {
            "Event ID": row["Event ID"],
            "Audit Entity ID": row["Audit Entity ID"],
            "Event Title": row["Event Title"],
            "Event Description": row["Event Description Full"],
            "Match Confidence": row["Match Confidence"],
        }
        for n in [1, 2, 3]:
            score = row[f"Match {n} - Score"]
            if score >= MIN_SIMILARITY_SCORE:
                record[f"Candidate {n} L2"] = row[f"Match {n} - L2"]
                record[f"Candidate {n} Definition"] = row[f"Match {n} - Definition"]
            else:
                record[f"Candidate {n} L2"] = ""
                record[f"Candidate {n} Definition"] = ""
            record[f"Candidate {n} Applies"] = ""
        record["Reviewer Notes"] = ""
        review_records.append(record)
    review_df = pd.DataFrame(review_records, columns=[
        "Event ID", "Audit Entity ID", "Event Title", "Event Description",
        "Match Confidence",
        "Candidate 1 L2", "Candidate 1 Definition", "Candidate 1 Applies",
        "Candidate 2 L2", "Candidate 2 Definition", "Candidate 2 Applies",
        "Candidate 3 L2", "Candidate 3 Definition", "Candidate 3 Applies",
        "Reviewer Notes",
    ])

    # =========================================================================
    # Sheet 3: Summary (visible)
    # =========================================================================
    total = len(mapping_df)
    needs_review_count = (mapping_df["Mapping Status"] == "Needs Review").sum()
    no_match_count = (mapping_df["Mapping Status"] == "No Match").sum()

    def pct(n):
        return f"{n} ({n/total*100:.1f}%)" if total > 0 else "0"

    nr_single = ((mapping_df["Mapping Status"] == "Needs Review") & (mapping_df["Mapped L2 Count"] == 1)).sum()
    nr_multi = ((mapping_df["Mapping Status"] == "Needs Review") & (mapping_df["Mapped L2 Count"] > 1)).sum()

    summary_data = {
        "Metric": [
            "Total OREs",
            "",
            "Needs Review",
            "  Mapped to single L2",
            "  Mapped to multiple L2s",
            "No Match",
            "",
            "",
            "HOW THIS WORKS",
            "",
            (f"This tool uses NLP word-vector similarity (spaCy {RESOLVED_SPACY_MODEL}) to suggest\n"
             "which L2 risk categories each ORE relates to. This is classical NLP — not\n"
             "generative AI / LLMs. Each ORE description is compared against the 23 L2\n"
             "definitions; matches are based on word-vector cosine similarity. These are\n"
             "starting points, not confirmed lookups, and must be reviewer-validated."),
            "",
            ("A single ORE can be suggested for more than one L2. For example, \"unauthorized\n"
             "payment processed due to system access control failure\" relates to both Fraud\n"
             "and Information and Cyber Security. When the tool detects this, it lists\n"
             "all L2s that fit."),
            "",
            ("Needs Review — Every ORE that passes the minimum similarity floor is marked\n"
             "Needs Review by design. NLP text similarity can be wrong (generic wording,\n"
             "L2 definitions that read similarly), so the tool does not assert a positive-\n"
             "confidence band. Open the Needs Review tab and confirm the L2 attribution\n"
             "for each ORE before relying on it. Similarity scores remain in the hidden\n"
             "Raw Scores tab for traceability."),
            "",
            ("No Match — Nothing fit well enough. The similarity scores were all below the\n"
             "minimum threshold. These are excluded from the pipeline. A reviewer can\n"
             "manually assign an L2 if needed."),
        ],
        "Value": [
            total,
            "",
            pct(needs_review_count),
            nr_single,
            nr_multi,
            pct(no_match_count),
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
        ],
    }
    summary_df = pd.DataFrame(summary_data)

    # =========================================================================
    # Sheet 4: L2 Distribution (visible)
    # Explode multi-L2 mappings so each L2 is counted separately. All
    # floor-passing items are Needs Review and feed downstream Impact of
    # Issues — reviewers want to see real volume per L2.
    # =========================================================================
    def _explode_band(status_value: str) -> pd.Series:
        rows = mapping_df[mapping_df["Mapping Status"] == status_value].copy()
        if rows.empty:
            return pd.Series(dtype=int)
        s = rows["Mapped L2s"].str.split("; ").explode().str.strip()
        return s[s != ""].value_counts()

    nr_counts = _explode_band("Needs Review")
    all_l2s = sorted(set(nr_counts.index))
    l2_dist = pd.DataFrame({
        "L2 Risk": all_l2s,
        "Needs Review": [int(nr_counts.get(l2, 0)) for l2 in all_l2s],
    })
    l2_dist["Total"] = l2_dist["Needs Review"]
    l2_dist = l2_dist.sort_values("Total", ascending=False).reset_index(drop=True)

    # =========================================================================
    # Sheet 5: Raw Scores (HIDDEN) — development and threshold tuning
    # =========================================================================
    raw_cols = [
        "Event ID", "Audit Entity ID", "Event Title", "Event Description Full",
        "Match 1 - L2", "Match 1 - Score",
        "Match 2 - L2", "Match 2 - Score",
        "Match 3 - L2", "Match 3 - Score",
        "Margin 1-2", "Margin 2-3",
        "Mapping Status", "Match Confidence", "Match 1 Valid",
    ]
    raw_scores = mapping_df[raw_cols].copy()
    raw_scores = raw_scores.rename(columns={"Event Description Full": "Event Description"})

    # Score distribution stats for the raw sheet
    valid_scores = mapping_df[mapping_df["Match 1 Valid"]]["Match 1 - Score"]
    valid_margins = mapping_df[mapping_df["Match 1 Valid"]]["Margin 1-2"]
    valid_margins = valid_margins[valid_margins > 0]

    raw_stats = {
        "Metric": [
            "Score Distribution (valid Match 1)",
            "  Mean",
            "  Median",
            "  Min",
            "  Max",
            "",
            "Margin Distribution (valid, non-zero)",
            "  P25",
            "  P50 (Median)",
            "  P75",
            "",
            "Settings",
            "  Ambiguity Threshold",
            "  Min Similarity Score",
            "  spaCy Model",
        ],
        "Value": [
            "",
            f"{valid_scores.mean():.4f}" if len(valid_scores) > 0 else "N/A",
            f"{valid_scores.median():.4f}" if len(valid_scores) > 0 else "N/A",
            f"{valid_scores.min():.4f}" if len(valid_scores) > 0 else "N/A",
            f"{valid_scores.max():.4f}" if len(valid_scores) > 0 else "N/A",
            "",
            "",
            f"{valid_margins.quantile(0.25):.4f}" if len(valid_margins) > 0 else "N/A",
            f"{valid_margins.quantile(0.50):.4f}" if len(valid_margins) > 0 else "N/A",
            f"{valid_margins.quantile(0.75):.4f}" if len(valid_margins) > 0 else "N/A",
            "",
            "",
            f"{threshold:.4f}",
            MIN_SIMILARITY_SCORE,
            RESOLVED_SPACY_MODEL,
        ],
    }
    raw_stats_df = pd.DataFrame(raw_stats)

    # =========================================================================
    # Write all sheets
    # =========================================================================
    logger.info(f"Writing output to {output_path}")
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        all_mappings.to_excel(writer, sheet_name="All Mappings", index=False)
        review_df.to_excel(writer, sheet_name="Needs Review", index=False)
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        l2_dist.to_excel(writer, sheet_name="L2 Distribution", index=False)
        raw_scores.to_excel(writer, sheet_name="Raw Scores", index=False)

        wb = writer.book

        # -- Format: All Mappings --
        ws_all = wb["All Mappings"]
        style_header(ws_all, ws_all.max_column)
        auto_fit_columns(ws_all, overrides={
            "Event Description": 60,
            "Event Title": 30,
            "Mapped L2s": 50,
            "Mapped L2 Definitions": 60,
        })
        apply_wrap(ws_all, ["Event Description", "Mapped L2s", "Mapped L2 Definitions"])
        color_status_column(ws_all)
        color_confidence_column(ws_all)
        ws_all.freeze_panes = "C2"  # Freeze header row + first 2 columns

        # -- Format: Needs Review --
        ws_review = wb["Needs Review"]
        style_header(ws_review, ws_review.max_column)
        review_width_overrides = {
            "Event Description": 60,
            "Event Title": 30,
            "Candidate 1 Definition": 60,
            "Candidate 2 Definition": 60,
            "Candidate 3 Definition": 60,
            "Candidate 1 L2": 25,
            "Candidate 2 L2": 25,
            "Candidate 3 L2": 25,
            "Candidate 1 Applies": 15,
            "Candidate 2 Applies": 15,
            "Candidate 3 Applies": 15,
            "Reviewer Notes": 30,
        }
        auto_fit_columns(ws_review, overrides=review_width_overrides)
        apply_wrap(ws_review, [
            "Event Description",
            "Candidate 1 Definition", "Candidate 2 Definition", "Candidate 3 Definition",
        ])
        color_confidence_column(ws_review)
        ws_review.freeze_panes = "A2"  # Freeze header row only
        # Set row heights for readability with full-length descriptions
        for row in range(2, ws_review.max_row + 1):
            ws_review.row_dimensions[row].height = 60
        # Highlight reviewer input column headers with green fill
        reviewer_input_cols = [
            "Candidate 1 Applies", "Candidate 2 Applies", "Candidate 3 Applies",
            "Reviewer Notes",
        ]
        for col in ws_review.iter_cols(min_row=1, max_row=1):
            if str(col[0].value) in reviewer_input_cols:
                col[0].fill = reviewer_fill

        # -- Format: Summary --
        ws_summary = wb["Summary"]
        style_header(ws_summary, ws_summary.max_column)
        ws_summary.column_dimensions["A"].width = 80
        ws_summary.column_dimensions["B"].width = 25
        # Wrap the explanation text cells
        for row in range(2, ws_summary.max_row + 1):
            ws_summary.cell(row=row, column=1).alignment = wrap_align

        # -- Format: L2 Distribution --
        ws_l2 = wb["L2 Distribution"]
        style_header(ws_l2, ws_l2.max_column)
        auto_fit_columns(ws_l2, overrides={"L2 Risk": 45})

        # -- Format: Raw Scores (then hide) --
        ws_raw = wb["Raw Scores"]
        style_header(ws_raw, ws_raw.max_column)
        auto_fit_columns(ws_raw, overrides={"Event Description": 60, "Event Title": 30})
        apply_wrap(ws_raw, ["Event Description"])

        # Write stats below the data
        stats_start_row = ws_raw.max_row + 3
        for r_idx, row_data in raw_stats_df.iterrows():
            ws_raw.cell(row=stats_start_row + r_idx, column=1, value=row_data["Metric"])
            ws_raw.cell(row=stats_start_row + r_idx, column=2, value=row_data["Value"])

        ws_raw.sheet_state = "hidden"

    logger.info(f"  Output saved: {output_path}")
    return output_path


# =============================================================================
# ORPHAN SIDECAR
# =============================================================================

def _write_orphans_sidecar(
    orphans_df: pd.DataFrame,
    mapping_output_path: Path,
    source_filename: str,
    *,
    id_col: str,
    title_col: str,
    status_col: str,
    source_label: str,
) -> Path:
    """Write a sidecar orphans file next to the mapping output.

    Filename: same prefix + timestamp as the mapping file, with `_orphans`
    inserted before the extension. Schema matches the Upstream Tagging Gaps
    tab so the main pipeline can read and pass through verbatim.
    """
    mapping_output_path = Path(mapping_output_path)
    sidecar_name = mapping_output_path.stem + "_orphans" + mapping_output_path.suffix
    sidecar_path = mapping_output_path.parent / sidecar_name

    n = len(orphans_df)

    def _col_as_list(df, col):
        if not col or col not in df.columns:
            return [""] * n
        return df[col].astype(str).tolist()

    out = pd.DataFrame({
        "Source": [source_label] * n,
        "Item ID": _col_as_list(orphans_df, id_col),
        "Title": _col_as_list(orphans_df, title_col),
        "Status": _col_as_list(orphans_df, status_col),
        "Drop Reason": ["Blank AE upstream"] * n,
        "Source File": [source_filename] * n,
    })
    out.to_excel(sidecar_path, index=False)
    return sidecar_path


# =============================================================================
# MAIN
# =============================================================================

def main():
    global AMBIGUITY_MARGIN_THRESHOLD

    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--source", choices=["ore", "ore_irm"], default="ore",
        help="Source dataset to map (ore = legacy ORE_*.xlsx; ore_irm = ORE_IRM_*.xlsx)",
    )
    args = parser.parse_args()
    set_active_source(args.source)
    logger.info(f"ORE mapper running with source: {SOURCE_NAME}")

    input_dir = _PROJECT_ROOT / "data" / "input"
    output_dir = _PROJECT_ROOT / "data" / "output"

    # Load data
    l2_df = load_l2_definitions(input_dir)
    ore_df, orphans_df, source_filename = load_ore_data(input_dir)

    # Load spaCy model
    global RESOLVED_SPACY_MODEL
    logger.info(f"Loading spaCy model: {SPACY_MODEL}")
    nlp = spacy.load(SPACY_MODEL)
    RESOLVED_SPACY_MODEL = spacy_model_label(nlp)
    logger.info(f"  Model loaded: {RESOLVED_SPACY_MODEL} "
                f"({len(nlp.vocab.vectors)} vectors, "
                f"{nlp.vocab.vectors.shape[1]} dimensions)")
    log_run_provenance(logger, SPACY_MODEL)

    # Build reference vectors from L2 definitions
    ref_vectors, l2_names, l2_definitions = build_reference_vectors(nlp, l2_df)

    # Compute similarity and get top-3 mappings
    mapping_df = compute_mappings(nlp, ore_df, ref_vectors, l2_names, l2_definitions)

    # Determine ambiguity threshold from data
    if AMBIGUITY_MARGIN_THRESHOLD is None:
        AMBIGUITY_MARGIN_THRESHOLD = determine_ambiguity_threshold(mapping_df)

    # Classify mappings
    mapping_df = classify_mappings(mapping_df, AMBIGUITY_MARGIN_THRESHOLD)

    # Summary stats
    total = len(mapping_df)
    needs_review = (mapping_df["Mapping Status"] == "Needs Review").sum()
    nr_single = ((mapping_df["Mapping Status"] == "Needs Review") & (mapping_df["Mapped L2 Count"] == 1)).sum()
    nr_multi = ((mapping_df["Mapping Status"] == "Needs Review") & (mapping_df["Mapped L2 Count"] > 1)).sum()
    no_match = (mapping_df["Mapping Status"] == "No Match").sum()

    logger.info("=" * 60)
    logger.info("ORE MAPPING COMPLETE")
    logger.info(f"  Total OREs: {total}")
    logger.info(f"  Needs Review: {needs_review} ({needs_review/total*100:.1f}%) — single: {nr_single}, multi: {nr_multi}")
    logger.info(f"  No Match: {no_match} ({no_match/total*100:.1f}%)")
    logger.info(f"  Ambiguity threshold: {AMBIGUITY_MARGIN_THRESHOLD:.4f}")
    logger.info("=" * 60)

    # Export
    output_path = export_results(mapping_df, AMBIGUITY_MARGIN_THRESHOLD, output_dir)

    # Write orphans sidecar — same timestamp as the mapping file. Picked up by
    # the main pipeline and surfaced in the Upstream Tagging Gaps tab.
    if not orphans_df.empty:
        sidecar_path = _write_orphans_sidecar(
            orphans_df, output_path, source_filename,
            id_col=ORE_ID_COL, title_col=ORE_TITLE_COL,
            status_col=ORE_STATUS_COL,
            source_label="OREs (legacy)" if SOURCE_NAME == "ore" else "ORE IRM",
        )
        logger.info(f"  Orphans sidecar saved: {sidecar_path} ({len(orphans_df)} rows)")

    print(f"\nDone! Output: {output_path}")
    print(f"  Needs Review: {needs_review} (single: {nr_single}, multi: {nr_multi}) | No Match: {no_match}")


if __name__ == "__main__":
    main()
