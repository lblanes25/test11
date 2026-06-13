"""Shared core for the NLP source-to-L2 mappers (ORE / PRSA / RAP).

Holds the source-agnostic mapping pipeline: L2 taxonomy loading, reference
vector construction, top-3 similarity scoring, ambiguity threshold detection,
classification (uniform Needs Review band), the 5-sheet Excel export
(All Mappings / Needs Review / Summary / L2 Distribution / hidden Raw Scores),
and the orphans sidecar writer.

Per-source deltas (column names, sheet column orders, formatting extras,
output filename prefix) are carried by a MapperSpec built by each mapper
script. The mappers keep their own CLI, config binding, source loading, and
filter stacks.

Advisory only — these mappers never set status authority in the pipeline.
"""

from __future__ import annotations

import logging
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Callable

import numpy as np
import pandas as pd
import spacy

from risk_taxonomy_transformer.config import L2_TO_L1
from risk_taxonomy_transformer.normalization import normalize_l2_name

logger = logging.getLogger(__name__)

# Ambiguity threshold auto-detection parameters. Margin threshold = P25 of
# valid non-zero margins, floored / capped. SpaCy scores are more compressed
# than TF-IDF, so tighter thresholds. (Config-resident move is a later item.)
AMBIGUITY_MARGIN_QUANTILE = 0.25
AMBIGUITY_MARGIN_FLOOR = 0.01
AMBIGUITY_MARGIN_CAP = 0.05
AMBIGUITY_MARGIN_FALLBACK = 0.02


@dataclass
class MapperSpec:
    """Per-source schema and formatting spec for the shared mapper core.

    Field semantics:
    - item_label: plural noun for log lines ("OREs", "PRSA issues", "RAPs")
    - output_prefix: output filename prefix (e.g. "ore_mapping")
    - min_similarity_score: inclusion floor for a valid Match 1
    - all_cols: All Mappings sheet column order (filtered to present columns)
    - review_fields: Needs Review base fields as (output header, mapping_df
      column) pairs — lets the full-description column feed the sheet
    - review_columns: explicit Needs Review column order, or None to rely on
      record insertion order (None yields a headerless sheet when empty)
    - raw_cols / raw_rename: Raw Scores sheet columns and header renames
    - *_width_overrides / *_wrap_cols: per-sheet formatting
    - summary_a_width / summary_wrap: Summary column A width and text wrap
    - review_row_height: fixed Needs Review data-row height, or None
    - reviewer_input_cols: Needs Review header cells to highlight as reviewer
      input, or None
    """
    item_label: str
    output_prefix: str
    min_similarity_score: float
    all_cols: list[str]
    review_fields: list[tuple[str, str]]
    review_columns: list[str] | None
    raw_cols: list[str]
    raw_rename: dict[str, str]
    all_width_overrides: dict[str, int] = field(default_factory=dict)
    all_wrap_cols: list[str] = field(default_factory=list)
    review_width_overrides: dict[str, int] = field(default_factory=dict)
    review_wrap_cols: list[str] = field(default_factory=list)
    raw_width_overrides: dict[str, int] = field(default_factory=dict)
    raw_wrap_cols: list[str] = field(default_factory=list)
    summary_a_width: int = 40
    summary_wrap: bool = False
    review_row_height: int | None = None
    reviewer_input_cols: list[str] | None = None


# =============================================================================
# TAXONOMY LOADING + REFERENCE VECTORS
# =============================================================================

def load_l2_definitions(input_dir: Path, taxonomy_file: str) -> pd.DataFrame:
    """Load L2 risk definitions from taxonomy file.

    Real enterprise files merge L1/L2/L3 cells across multiple rows; pandas
    reads the continuation rows as NaN. Forward-fill so every row has a
    populated L1/L2/L3 — otherwise the bucketing loop in
    build_reference_vectors skips continuation rows and silently drops their
    L3/L4 definitions from the per-L2 reference vector.
    """
    filepath = input_dir / taxonomy_file
    logger.info(f"Loading L2 definitions from {filepath}")
    df = pd.read_excel(filepath)
    ffill_cols = [c for c in ("L1", "L2", "L3") if c in df.columns]
    if ffill_cols:
        df[ffill_cols] = df[ffill_cols].ffill()
    logger.info(f"  Loaded {len(df)} L2 definitions")
    return df


def build_reference_vectors(
    nlp: spacy.language.Language,
    l2_df: pd.DataFrame,
) -> tuple[np.ndarray, list[str], list[str]]:
    """Build document vectors for L2 risk definitions.

    Uses L2 name + definition for richer semantic representation.
    Returns (vectors array, l2_names list, l2_definitions list).
    """
    # Aggregate by CANONICAL evaluated L2, bucketing at the finer
    # grain (L3 over L2) when L3 normalizes to an evaluated L2.
    # Rationale: enterprise files often split L2s at L3 grain -- e.g.
    # "Fraud (External and Internal)" at L2 with Internal Fraud /
    # External Fraud - First Party / - Victim Fraud at L3. We want
    # Internal Fraud and External Fraud as distinct vectors, not
    # merged under the parent L2.
    def _clean(v):
        s = str(v if v is not None else "").strip()
        return "" if s.lower() in ("nan", "none") else s

    _evaluated = set(L2_TO_L1.keys())
    has_l3 = "L3" in l2_df.columns
    # Fold child-level (L3, L4) text into the per-L2 reference vector.
    # L1 / L1 Definition intentionally excluded — L1 is the parent (more
    # generic), would dilute the L2's vector with broader concepts. L3/L4
    # narrow the L2's scope and sharpen matches.
    sub_cols = [c for c in ["L3", "L3 Definition", "L4", "L4 Definition"]
                if c in l2_df.columns]

    def _bucket_for(l2_name, l3_name):
        """Return canonical evaluated L2 for this row, or None to skip."""
        if l3_name:
            c = normalize_l2_name(l3_name)
            if c in _evaluated:
                return c
        c = normalize_l2_name(l2_name)
        if c in _evaluated:
            return c
        return None

    l2_aggregate = {}  # canonical -> {"def": str, "text_parts": list[str]}
    skipped = 0
    for _, row in l2_df.iterrows():
        l2_name = _clean(row.get("L2"))
        if not l2_name:
            skipped += 1
            continue
        l3_name = _clean(row.get("L3")) if has_l3 else ""
        bucket = _bucket_for(l2_name, l3_name)
        if bucket is None:
            skipped += 1
            continue
        if bucket not in l2_aggregate:
            l2_aggregate[bucket] = {"def": "", "text_parts": [bucket]}
            l2_def = _clean(row.get("L2 Definition"))
            if l2_def:
                l2_aggregate[bucket]["def"] = l2_def
                l2_aggregate[bucket]["text_parts"].append(l2_def)
        # Fold L3/L4 text into this bucket's vector for richer signal.
        for col in sub_cols:
            val = _clean(row.get(col))
            if val and val not in l2_aggregate[bucket]["text_parts"]:
                l2_aggregate[bucket]["text_parts"].append(val)

    if skipped:
        logger.info(f"  Skipped {skipped} rows that did not resolve to an evaluated L2")

    l2_names = list(l2_aggregate.keys())
    l2_definitions = [l2_aggregate[n]["def"] for n in l2_names]
    l2_texts = [". ".join(l2_aggregate[n]["text_parts"]) for n in l2_names]

    logger.info(
        f"Computing vectors for {len(l2_texts)} unique L2s "
        f"(aggregated from {len(l2_df)} rows)..."
    )
    vectors = []
    for text in l2_texts:
        doc = nlp(text)
        vectors.append(doc.vector)
    vectors = np.array(vectors)

    # Normalize to unit vectors for cosine similarity via dot product
    norms = np.linalg.norm(vectors, axis=1, keepdims=True)
    norms[norms == 0] = 1
    vectors = vectors / norms

    logger.info(f"  Reference vectors shape: {vectors.shape}")
    return vectors, l2_names, l2_definitions


# =============================================================================
# SCORING + CLASSIFICATION
# =============================================================================

def compute_mappings(
    nlp: spacy.language.Language,
    source_df: pd.DataFrame,
    ref_vectors: np.ndarray,
    l2_names: list[str],
    l2_definitions: list[str],
    *,
    text_fn: Callable[[pd.Series], str],
    record_fn: Callable[[pd.Series], dict],
    item_label: str,
    min_similarity_score: float,
) -> pd.DataFrame:
    """Compute semantic similarity and produce top-3 mappings per item.

    text_fn builds the text to vectorize from a source row; record_fn returns
    the source-specific leading fields (IDs, titles, descriptions, statuses)
    for the result record. Match/Margin/Valid fields are appended here.
    """
    total = len(source_df)
    logger.info(f"Computing vectors for {total} {item_label}...")

    results = []
    log_interval = max(1, total // 10)

    for i, (_, row) in enumerate(source_df.iterrows()):
        if i > 0 and i % log_interval == 0:
            logger.info(f"  Processed {i}/{total} {item_label} ({i/total*100:.0f}%)")

        combined = text_fn(row)
        doc = nlp(combined)
        item_vector = doc.vector

        norm = np.linalg.norm(item_vector)
        if norm > 0:
            item_vector = item_vector / norm

        # Cosine similarity via dot product (both vectors are unit-normalized)
        scores = ref_vectors @ item_vector

        top_indices = np.argsort(scores)[::-1][:3]

        top1_idx = top_indices[0]
        top2_idx = top_indices[1]
        top3_idx = top_indices[2]

        top1_score = float(scores[top1_idx])
        top2_score = float(scores[top2_idx])
        top3_score = float(scores[top3_idx])

        margin_1_2 = top1_score - top2_score
        margin_2_3 = top2_score - top3_score

        record = record_fn(row)
        record.update({
            "Match 1 - L2": l2_names[top1_idx],
            "Match 1 - Score": round(top1_score, 4),
            "Match 1 - Definition": l2_definitions[top1_idx],
            "Match 2 - L2": l2_names[top2_idx],
            "Match 2 - Score": round(top2_score, 4),
            "Match 2 - Definition": l2_definitions[top2_idx],
            "Match 3 - L2": l2_names[top3_idx],
            "Match 3 - Score": round(top3_score, 4),
            "Match 3 - Definition": l2_definitions[top3_idx],
            "Margin 1-2": round(margin_1_2, 4),
            "Margin 2-3": round(margin_2_3, 4),
            "Match 1 Valid": top1_score >= min_similarity_score,
        })
        results.append(record)

    logger.info(f"  Computed mappings for {len(results)} {item_label}")
    return pd.DataFrame(results)


def determine_ambiguity_threshold(mapping_df: pd.DataFrame) -> float:
    """Determine the margin threshold from data distribution.

    Uses the AMBIGUITY_MARGIN_QUANTILE of margins for valid matches,
    floored at AMBIGUITY_MARGIN_FLOOR and capped at AMBIGUITY_MARGIN_CAP.
    """
    valid = mapping_df[mapping_df["Match 1 Valid"]]
    margins = valid["Margin 1-2"]
    margins = margins[margins > 0]

    if len(margins) == 0:
        return AMBIGUITY_MARGIN_FALLBACK

    p25 = margins.quantile(AMBIGUITY_MARGIN_QUANTILE)
    median = margins.quantile(0.50)
    threshold = max(AMBIGUITY_MARGIN_FLOOR, min(p25, AMBIGUITY_MARGIN_CAP))

    logger.info(f"  Margin distribution (valid matches) — P25: {p25:.4f}, median: {median:.4f}")
    logger.info(f"  Ambiguity threshold set to: {threshold:.4f}")
    return threshold


def classify_mappings(
    mapping_df: pd.DataFrame,
    threshold: float,
    *,
    min_similarity_score: float,
) -> pd.DataFrame:
    """Classify each item and determine which L2s it maps to.

    Status logic:
      - No Match: Match 1 below the similarity floor (excluded).
      - Needs Review: Match 1 valid. Every item that passes the inclusion
        floor is presented as Needs Review — the tool does not assert a
        positive-confidence band on NLP matches. The similarity scores,
        margins, and threshold remain in the hidden Raw Scores sheet for
        traceability. The single-vs-multi L2 selection logic (margin to
        Match 2/3) is retained so reviewers see the right candidate set.

    Produces:
      - Mapping Status (Needs Review / No Match)
      - Match Confidence (Review Required / Weak)
      - Mapped L2s (semicolon-separated list)
      - Mapped L2 Count (integer)
      - Mapped L2 Definitions (semicolon-separated, matching order)
    """
    df = mapping_df.copy()

    statuses = []
    confidence_bands = []
    mapped_l2s_list = []
    mapped_l2_counts = []
    mapped_l2_defs_list = []

    for _, row in df.iterrows():
        if not row["Match 1 Valid"]:
            statuses.append("No Match")
            confidence_bands.append("Weak")
            mapped_l2s_list.append("")
            mapped_l2_counts.append(0)
            mapped_l2_defs_list.append("")
            continue

        margin_1_2 = row["Margin 1-2"]

        if margin_1_2 < threshold:
            # Can't confidently separate top matches — show all candidates
            # above the floor for the reviewer.
            candidates = []
            candidate_defs = []
            for n in [1, 2, 3]:
                if row[f"Match {n} - Score"] >= min_similarity_score:
                    candidates.append(row[f"Match {n} - L2"])
                    candidate_defs.append(row[f"Match {n} - Definition"])
            statuses.append("Needs Review")
            confidence_bands.append("Review Required")
            mapped_l2s_list.append("; ".join(candidates))
            mapped_l2_counts.append(len(candidates))
            mapped_l2_defs_list.append("; ".join(candidate_defs))
        else:
            # Primary (Match 1) plus any Match 2/3 that qualify as additional
            # L2s: must be above the floor AND within 2x the ambiguity
            # threshold of Match 1's score. The band is NOT asserted as a
            # positive match — every floor-passing item is Needs Review.
            top_score = row["Match 1 - Score"]
            l2s = [row["Match 1 - L2"]]
            defs = [row["Match 1 - Definition"]]
            for n in [2, 3]:
                score = row[f"Match {n} - Score"]
                if (score >= min_similarity_score
                        and (top_score - score) < threshold * 2):
                    l2s.append(row[f"Match {n} - L2"])
                    defs.append(row[f"Match {n} - Definition"])
            statuses.append("Needs Review")
            confidence_bands.append("Review Required")
            mapped_l2s_list.append("; ".join(l2s))
            mapped_l2_counts.append(len(l2s))
            mapped_l2_defs_list.append("; ".join(defs))

    df["Mapping Status"] = statuses
    df["Match Confidence"] = confidence_bands
    df["Mapped L2s"] = mapped_l2s_list
    df["Mapped L2 Count"] = mapped_l2_counts
    df["Mapped L2 Definitions"] = mapped_l2_defs_list

    return df


# =============================================================================
# SUMMARY HELPERS
# =============================================================================

def build_basic_summary_df(
    mapping_df: pd.DataFrame,
    threshold: float,
    min_similarity_score: float,
    *,
    total_label: str,
    item_word: str,
) -> pd.DataFrame:
    """Build the short-form Summary sheet used by the PRSA and RAP mappers."""
    total = len(mapping_df)
    needs_review = (mapping_df["Mapping Status"] == "Needs Review").sum()
    no_match = (mapping_df["Mapping Status"] == "No Match").sum()

    def pct(n):
        return f"{n} ({n/total*100:.1f}%)" if total > 0 else "0"

    return pd.DataFrame({
        "Metric": [total_label, "Needs Review", "No Match",
                   "", "Ambiguity Threshold", "Min Similarity Score",
                   "", "Note"],
        "Value": [total, pct(needs_review), pct(no_match),
                  "", f"{threshold:.4f}", min_similarity_score,
                  "",
                  (f"Every {item_word} above the similarity floor is marked Needs Review by "
                   "design. NLP text similarity can be wrong; confirm the L2 attribution "
                   "before relying on it. Scores remain in the hidden Raw Scores tab.")],
    })


# =============================================================================
# EXCEL EXPORT
# =============================================================================

def style_header(ws, max_col: int):
    """Apply the standard mapper header style to row 1."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    header_font = Font(bold=True, color="FFFFFF", size=10, name="Arial")
    header_fill = PatternFill("solid", fgColor="2F5496")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    for col in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border


def auto_fit_columns(ws, overrides: dict | None = None, cap: int = 25):
    """Set column widths with optional overrides and a max cap."""
    from openpyxl.utils import get_column_letter
    overrides = overrides or {}
    for col in ws.iter_cols(min_row=1, max_row=1):
        col_letter = get_column_letter(col[0].column)
        header_val = str(col[0].value or "")
        if header_val in overrides:
            ws.column_dimensions[col_letter].width = overrides[header_val]
        else:
            ws.column_dimensions[col_letter].width = min(
                max(len(header_val) + 4, 12), cap
            )


def apply_wrap(ws, columns: list[str]):
    """Apply text wrap to data cells in named columns."""
    from openpyxl.styles import Alignment
    wrap_align = Alignment(vertical="top", wrap_text=True)
    header_map = {}
    for col in ws.iter_cols(min_row=1, max_row=1):
        header_map[str(col[0].value)] = col[0].column
    for col_name in columns:
        if col_name in header_map:
            col_idx = header_map[col_name]
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=col_idx).alignment = wrap_align


def color_column(ws, col_name: str, fills: dict):
    """Apply conditional fill to a named column's data cells."""
    header_map = {}
    for col in ws.iter_cols(min_row=1, max_row=1):
        header_map[str(col[0].value)] = col[0].column
    if col_name not in header_map:
        return
    col_idx = header_map[col_name]
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=col_idx)
        fill = fills.get(str(cell.value))
        if fill:
            cell.fill = fill


def export_results(
    mapping_df: pd.DataFrame,
    threshold: float,
    output_dir: Path,
    spec: MapperSpec,
    summary_df: pd.DataFrame,
    raw_stats_df: pd.DataFrame | None = None,
) -> Path:
    """Write results to multi-sheet Excel with formatting.

    Sheets:
      1. All Mappings — one row per item, reviewer-friendly (no raw scores)
      2. Needs Review — side-by-side comparison for ambiguous items
      3. Summary — counts and plain-language explanation (caller-built)
      4. L2 Distribution — item count per L2 (exploded for multi-L2)
      5. Raw Scores — hidden, for development and threshold tuning;
         raw_stats_df (if given) is written below the data
    """
    from openpyxl.styles import PatternFill, Alignment

    timestamp = datetime.now().strftime("%m%d%Y%I%M%p")
    output_path = output_dir / f"{spec.output_prefix}_{timestamp}.xlsx"

    wrap_align = Alignment(vertical="top", wrap_text=True)
    reviewer_fill = PatternFill("solid", fgColor="E2EFDA")

    status_fills = {
        "Suggested Match": PatternFill("solid", fgColor="C6EFCE"),
        "Needs Review": PatternFill("solid", fgColor="FFFF00"),
        "No Match": PatternFill("solid", fgColor="D9D9D9"),
    }
    confidence_fills = {
        "Strong": PatternFill("solid", fgColor="C6EFCE"),
        "Moderate": PatternFill("solid", fgColor="FCE4D6"),
        "Weak": PatternFill("solid", fgColor="D9D9D9"),
        "Review Required": PatternFill("solid", fgColor="FFFF00"),
    }

    # Sheet 1: All Mappings — downstream consumers (control effectiveness
    # pipeline) explode the semicolon-separated Mapped L2s column into per-L2
    # rows when building their indexes.
    all_cols = [c for c in spec.all_cols if c in mapping_df.columns]
    all_mappings = mapping_df[all_cols].copy()

    # Sheet 2: Needs Review — side-by-side comparison workspace
    needs_review_rows = mapping_df[mapping_df["Mapping Status"] == "Needs Review"].copy()
    review_records = []
    for _, row in needs_review_rows.iterrows():
        record = {out_name: row[src_col] for out_name, src_col in spec.review_fields}
        record["Match Confidence"] = row["Match Confidence"]
        for n in [1, 2, 3]:
            score = row[f"Match {n} - Score"]
            if score >= spec.min_similarity_score:
                record[f"Candidate {n} L2"] = row[f"Match {n} - L2"]
                record[f"Candidate {n} Definition"] = row[f"Match {n} - Definition"]
            else:
                record[f"Candidate {n} L2"] = ""
                record[f"Candidate {n} Definition"] = ""
            record[f"Candidate {n} Applies"] = ""
        record["Reviewer Notes"] = ""
        review_records.append(record)
    if spec.review_columns is not None:
        review_df = pd.DataFrame(review_records, columns=spec.review_columns)
    else:
        review_df = pd.DataFrame(review_records)

    # Sheet 4: L2 Distribution — explode multi-L2 mappings so each L2 is
    # counted separately. All floor-passing items are Needs Review and feed
    # downstream Impact of Issues; reviewers want real volume per L2.
    def _explode_band(status_value: str) -> pd.Series:
        rows = mapping_df[mapping_df["Mapping Status"] == status_value].copy()
        if rows.empty:
            return pd.Series(dtype=int)
        s = rows["Mapped L2s"].str.split("; ").explode().str.strip()
        return s[s != ""].value_counts()

    nr_counts = _explode_band("Needs Review")
    all_l2s = sorted(set(nr_counts.index))
    l2_dist = pd.DataFrame({
        "L2 Risk": all_l2s,
        "Needs Review": [int(nr_counts.get(l2, 0)) for l2 in all_l2s],
    })
    l2_dist["Total"] = l2_dist["Needs Review"]
    l2_dist = l2_dist.sort_values("Total", ascending=False).reset_index(drop=True)

    # Sheet 5: Raw Scores (hidden)
    raw_scores = mapping_df[spec.raw_cols].copy().rename(columns=spec.raw_rename)

    logger.info(f"Writing output to {output_path}")
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        all_mappings.to_excel(writer, sheet_name="All Mappings", index=False)
        review_df.to_excel(writer, sheet_name="Needs Review", index=False)
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        l2_dist.to_excel(writer, sheet_name="L2 Distribution", index=False)
        raw_scores.to_excel(writer, sheet_name="Raw Scores", index=False)

        wb = writer.book

        # -- Format: All Mappings --
        ws_all = wb["All Mappings"]
        style_header(ws_all, ws_all.max_column)
        auto_fit_columns(ws_all, overrides=spec.all_width_overrides)
        apply_wrap(ws_all, spec.all_wrap_cols)
        color_column(ws_all, "Mapping Status", status_fills)
        color_column(ws_all, "Match Confidence", confidence_fills)
        ws_all.freeze_panes = "C2"  # Freeze header row + first 2 columns

        # -- Format: Needs Review --
        ws_review = wb["Needs Review"]
        style_header(ws_review, ws_review.max_column)
        auto_fit_columns(ws_review, overrides=spec.review_width_overrides)
        apply_wrap(ws_review, spec.review_wrap_cols)
        color_column(ws_review, "Match Confidence", confidence_fills)
        ws_review.freeze_panes = "A2"  # Freeze header row only
        if spec.review_row_height is not None:
            for row in range(2, ws_review.max_row + 1):
                ws_review.row_dimensions[row].height = spec.review_row_height
        if spec.reviewer_input_cols:
            for col in ws_review.iter_cols(min_row=1, max_row=1):
                if str(col[0].value) in spec.reviewer_input_cols:
                    col[0].fill = reviewer_fill

        # -- Format: Summary --
        ws_summary = wb["Summary"]
        style_header(ws_summary, ws_summary.max_column)
        ws_summary.column_dimensions["A"].width = spec.summary_a_width
        ws_summary.column_dimensions["B"].width = 25
        if spec.summary_wrap:
            for row in range(2, ws_summary.max_row + 1):
                ws_summary.cell(row=row, column=1).alignment = wrap_align

        # -- Format: L2 Distribution --
        ws_l2 = wb["L2 Distribution"]
        style_header(ws_l2, ws_l2.max_column)
        auto_fit_columns(ws_l2, overrides={"L2 Risk": 45})

        # -- Format: Raw Scores (then hide) --
        ws_raw = wb["Raw Scores"]
        style_header(ws_raw, ws_raw.max_column)
        auto_fit_columns(ws_raw, overrides=spec.raw_width_overrides)
        apply_wrap(ws_raw, spec.raw_wrap_cols)

        if raw_stats_df is not None:
            stats_start_row = ws_raw.max_row + 3
            for r_idx, row_data in raw_stats_df.iterrows():
                ws_raw.cell(row=stats_start_row + r_idx, column=1, value=row_data["Metric"])
                ws_raw.cell(row=stats_start_row + r_idx, column=2, value=row_data["Value"])

        ws_raw.sheet_state = "hidden"

    logger.info(f"  Output saved: {output_path}")
    return output_path


# =============================================================================
# ORPHAN SIDECAR
# =============================================================================

def write_orphans_sidecar(
    orphans_df: pd.DataFrame,
    mapping_output_path: Path,
    source_filename: str,
    *,
    id_col: str,
    title_col: str,
    status_col: str,
    source_label: str,
) -> Path:
    """Write a sidecar orphans file next to the mapping output.

    Filename: same prefix + timestamp as the mapping file, with `_orphans`
    inserted before the extension. Schema matches the Upstream Tagging Gaps
    tab so the main pipeline can read and pass through verbatim.
    """
    mapping_output_path = Path(mapping_output_path)
    sidecar_name = mapping_output_path.stem + "_orphans" + mapping_output_path.suffix
    sidecar_path = mapping_output_path.parent / sidecar_name

    n = len(orphans_df)

    def _col_as_list(df, col):
        if not col or col not in df.columns:
            return [""] * n
        return df[col].astype(str).tolist()

    out = pd.DataFrame({
        "Source": [source_label] * n,
        "Item ID": _col_as_list(orphans_df, id_col),
        "Title": _col_as_list(orphans_df, title_col),
        "Status": _col_as_list(orphans_df, status_col),
        "Drop Reason": ["Blank AE upstream"] * n,
        "Source File": [source_filename] * n,
    })
    out.to_excel(sidecar_path, index=False)
    return sidecar_path
