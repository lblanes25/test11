"""
Excel export for the Risk Taxonomy Transformer.

Writes the multi-sheet output workbook, enriches source tabs, and applies
all formatting via the formatting module.
"""

from __future__ import annotations

import logging
import re
from pathlib import Path

import pandas as pd
import yaml
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

from risk_taxonomy_transformer.config import KEYWORD_MAP, get_config
from risk_taxonomy_transformer.constants import Status
from risk_taxonomy_transformer.formatting import (
    _build_dashboard_sheet,
    _color_rows_by_column,
    _find_header_column,
    _format_audit_review_sheet,
    _format_risk_owner_review_sheet,
    _format_risk_owner_summary_sheet,
    style_header,
)
from risk_taxonomy_transformer.normalization import normalize_l2_name
from risk_taxonomy_transformer.review_builders import (
    build_audit_review_df,
    build_review_queue_df,
    build_risk_owner_review_df,
    build_ro_summary_df,
)

logger = logging.getLogger(__name__)


# Source tab → banners.yaml key. Drives the merged disclosure row written
# at Excel row 1 on each source tab. Sheets here are also written with
# startrow=1 so row 1 is reserved for the banner.
_SOURCE_TAB_BANNER_KEYS = {
    "Source - Findings": "iag",
    "Source - OREs": "ore",
    "Source - ORE IRM": "ore_irm",
    "Source - PRSA Issues": "prsa",
    "Source - PG Gaps": "pg_gap",
    "Source - GRA RAPs": "gra_rap",
    "Source - BM Activities": "bma",
    "Source - Key Risks": "key_risks",
    "Source - L2 Taxonomy": "l2_taxonomy",
    "Source - Models": "models",
    "Source - Legacy Data": "legacy_data",
    "Upstream Tagging Gaps": "upstream_tagging_gaps",
}

_BANNERS_PATH = Path(__file__).resolve().parent.parent / "config" / "banners.yaml"


def _load_source_banners() -> dict:
    """Load the ``source_banners`` map from config/banners.yaml.

    Returns a dict {key: {style, body}}. Empty dict on missing file.
    """
    try:
        with open(_BANNERS_PATH, "r", encoding="utf-8") as f:
            cfg = yaml.safe_load(f) or {}
        return cfg.get("source_banners", {})
    except FileNotFoundError:
        logger.warning(f"banners.yaml not found at {_BANNERS_PATH}; source-tab banners will be skipped.")
        return {}


_HTML_TAG_RE = re.compile(r"<[^>]+>")


def _strip_html(text: str) -> str:
    """Remove HTML tags from banner body text for Excel cell display."""
    return _HTML_TAG_RE.sub("", text or "").strip()


# ---------------------------------------------------------------------------
# Source enrichment helpers
# ---------------------------------------------------------------------------

def _enrich_findings_source(
    findings_path: str,
    column_name_map: dict,
    transformed_df: pd.DataFrame,
) -> pd.DataFrame:
    """Build an enriched findings source tab showing what each finding mapped to.

    Reads the raw findings file (before filtering) and annotates each row with:
    - Mapping Status: what happened to this finding (Included, Filtered, Closed)
    - Mapped L2(s): which L2 risk(s) this finding confirmed applicability for
    """
    if findings_path.endswith(".csv"):
        df = pd.read_csv(findings_path)
    else:
        df = pd.read_excel(findings_path)
    df.columns = [c.strip() for c in df.columns]

    # Rename to internal names for consistency
    rename = {}
    for internal, actual in column_name_map.items():
        if actual and actual in df.columns:
            rename[actual] = internal
    df = df.rename(columns=rename)
    df["entity_id"] = df["entity_id"].astype(str).str.strip()

    # Determine disposition for each row
    dispositions = []
    mapped_l2s_col = []

    # Build a set of (entity_id, l2) pairs that were issue_confirmed in the output
    confirmed = set()
    if transformed_df is not None:
        for _, row in transformed_df.iterrows():
            if "issue_confirmed" in str(row.get("method", "")):
                confirmed.add((str(row["entity_id"]), str(row["new_l2"])))

    # Active finding statuses -- only these are actionable
    _active_statuses = {"open", "in validation", "in sustainability"}

    for _, row in df.iterrows():
        # If finding is closed/inactive, suppress filter-reason noise
        status = str(row.get("status", row.get("Finding Status", ""))).strip()
        if status and status.lower() not in _active_statuses:
            dispositions.append("Closed")
            mapped_l2s_col.append("")
            continue

        # Check approval -- try internal name first (renamed), then original column name
        approval = str(row.get("approval_status", row.get("Finding Approval Status", ""))).strip()
        if approval and approval != "Approved":
            dispositions.append(f"Filtered \u2014 not approved ({approval})")
            mapped_l2s_col.append("")
            continue

        # Check severity
        sev = row.get("severity")
        if pd.isna(sev) or str(sev).strip() == "":
            dispositions.append("Filtered \u2014 blank severity")
            mapped_l2s_col.append("")
            continue

        # Check L2 mapping
        raw_l2 = str(row.get("l2_risk", ""))
        if not raw_l2 or raw_l2 == "nan":
            dispositions.append("Filtered \u2014 blank L2 risk category")
            mapped_l2s_col.append("")
            continue

        # Normalize and check each L2 value (could be multi-value)
        l2_parts = raw_l2.replace("\r\n", "\n").replace("\r", "\n").split("\n")
        mapped = []
        unmapped = []
        for part in l2_parts:
            normalized = normalize_l2_name(part.strip())
            if normalized:
                eid = str(row["entity_id"])
                if (eid, normalized) in confirmed:
                    mapped.append(normalized)
                else:
                    mapped.append(f"{normalized} (not active/applicable)")
            elif part.strip():
                unmapped.append(part.strip())

        if mapped:
            dispositions.append("Included")
            mapped_l2s_col.append("; ".join(mapped))
        elif unmapped:
            dispositions.append(f"Filtered \u2014 unmappable L2 ({'; '.join(unmapped)})")
            mapped_l2s_col.append("")
        else:
            dispositions.append("Filtered \u2014 L2 not resolved")
            mapped_l2s_col.append("")

    df["Mapping Status"] = dispositions
    df["Mapped To L2(s)"] = mapped_l2s_col

    return df


_ACRONYM_DISPLAY = {
    "aml": "AML", "bcp": "BCP", "bsa": "BSA", "ccar": "CCAR", "ccpa": "CCPA",
    "cra": "CRA", "ddos": "DDoS", "dr": "DR", "fx": "FX", "gaap": "GAAP",
    "gdpr": "GDPR", "hr": "HR", "it": "IT", "kyc": "KYC", "mrm": "MRM",
    "nii": "NII", "ofac": "OFAC", "pii": "PII", "sar": "SAR", "sec": "SEC",
    "udaap": "UDAAP",
}


def _format_keyword_for_display(kw: str) -> str:
    tokens = kw.strip().split()
    return " ".join(_ACRONYM_DISPLAY.get(t.lower(), t) for t in tokens)


def _enrich_key_risks_source(
    key_risks_df: pd.DataFrame,
    transformed_df: pd.DataFrame,
) -> pd.DataFrame:
    """Build an enriched key risks source tab showing what each key risk contributed to.

    Annotates each row with which L2(s) it provided keyword evidence for.
    """
    if key_risks_df is None or key_risks_df.empty:
        return pd.DataFrame()

    df = key_risks_df.copy()

    contributions = []
    for _, row in df.iterrows():
        eid = str(row.get("entity_id", ""))
        desc = str(row.get("risk_description", "")).lower()
        l1 = str(row.get("legacy_l1", ""))

        if not desc or desc == "nan":
            contributions.append("No description text")
            continue

        # Check which L2 keywords match this description
        matched_l2s = []
        for l2_name, keywords in KEYWORD_MAP.items():
            hits = [kw for kw in keywords if kw in desc]
            if hits:
                rendered = ", ".join(_format_keyword_for_display(h) for h in hits[:3])
                matched_l2s.append(f"{l2_name} ({rendered})")

        if matched_l2s:
            contributions.append("; ".join(matched_l2s))
        else:
            contributions.append("No keyword matches \u2014 did not contribute to any L2 mapping")

    df["L2 Keyword Matches"] = contributions

    return df


# ---------------------------------------------------------------------------
# Legacy ratings lookup builder
# ---------------------------------------------------------------------------

def _build_legacy_lookup(
    legacy_df: pd.DataFrame,
    pillar_columns: dict,
    entity_id_col: str,
) -> pd.DataFrame:
    """Unpivot legacy data into a clean lookup: one row per entity-pillar.

    Columns: Entity ID | Risk Pillar | Inherent Risk Rating |
             Inherent Risk Rationale | Control Assessment |
             Control Assessment Rationale
    """
    rows = []
    for _, entity_row in legacy_df.iterrows():
        eid = str(entity_row.get(entity_id_col, "")).strip()
        if not eid or eid == "nan":
            continue
        for pillar_name, cols in pillar_columns.items():
            rating = entity_row.get(cols["rating"], "")
            rationale = entity_row.get(cols.get("rationale") or "", "")
            control = entity_row.get(cols["control"], "")
            control_rationale = entity_row.get(cols.get("control_rationale") or "", "")
            # Convert NaN to empty string
            rating = "" if pd.isna(rating) else str(rating).strip()
            rationale = "" if pd.isna(rationale) else str(rationale).strip()
            control = "" if pd.isna(control) else str(control).strip()
            control_rationale = "" if pd.isna(control_rationale) else str(control_rationale).strip()
            rows.append({
                "Entity ID": eid,
                "Risk Pillar": pillar_name,
                "Inherent Risk Rating": rating,
                "Inherent Risk Rationale": rationale,
                "Control Assessment": control,
                "Control Assessment Rationale": control_rationale,
            })
    return pd.DataFrame(rows)


# ---------------------------------------------------------------------------
# Models inventory source-tab builder
# ---------------------------------------------------------------------------

def _build_models_source_df(
    legacy_df: pd.DataFrame,
    legacy_models_col: str,
    model_inv_cfg: dict,
    inventory_pattern: str,
) -> pd.DataFrame:
    """Load model inventory, filter to IDs referenced in legacy data, return
    the 5-column Source - Models DataFrame.

    Columns (in order): Model ID, Model Name, Model Class, Markets,
    Model Impact Category. Returns an empty DataFrame if no inventory file
    is found or no model IDs are referenced in legacy data.
    """
    project_root = Path(__file__).resolve().parent.parent
    input_dir = project_root / "data" / "input"
    matches = sorted(input_dir.glob(inventory_pattern))
    if not matches:
        logger.info(
            f"  No model inventory file matching '{inventory_pattern}' — "
            f"Source - Models tab will not be written"
        )
        return pd.DataFrame()
    latest = max(matches, key=lambda p: p.stat().st_mtime)
    try:
        inv_df = pd.read_excel(latest)
    except Exception as exc:
        logger.warning(f"  Failed to read model inventory '{latest}': {exc}")
        return pd.DataFrame()

    inv_df.columns = [c.strip() for c in inv_df.columns]
    id_col = model_inv_cfg.get("id", "Model ID")
    name_col = model_inv_cfg.get("name", "Model Name")
    class_col = model_inv_cfg.get("model_class", "Model Class")
    markets_col = model_inv_cfg.get("markets", "Markets")
    impact_col = model_inv_cfg.get("impact", "Model Impact Category")

    referenced: set[str] = set()
    if legacy_models_col in legacy_df.columns:
        import re
        for val in legacy_df[legacy_models_col].dropna().tolist():
            s = str(val).strip()
            if not s or s.lower() in ("nan", "none"):
                continue
            for part in re.split(r"[;\r\n]+", s):
                part = part.strip()
                if not part or part.lower() in ("nan", "none", "n/a", "not applicable", "not available"):
                    continue
                for tok in re.findall(r"\d{2,}", part):
                    referenced.add(tok)

    if not referenced or id_col not in inv_df.columns:
        return pd.DataFrame()

    # Normalize to strings; strip trailing ".0" so Excel-loaded float IDs
    # (1178.0) match legacy-derived "1178". Same issue as
    # export_html_report._norm_id_series.
    inv_ids_norm = (
        inv_df[id_col].astype(str).str.strip().str.replace(r"\.0+$", "", regex=True)
    )
    mask = inv_ids_norm.isin(referenced)
    filtered = inv_df[mask].copy()
    if filtered.empty:
        logger.warning(
            f"  Source - Models filtered to 0 rows. Referenced IDs: "
            f"{sorted(referenced)[:5]}... Inventory sample: "
            f"{inv_ids_norm.head(5).tolist()}"
        )
        return pd.DataFrame()

    desired_order = [id_col, name_col, class_col, markets_col, impact_col]
    available = [c for c in desired_order if c in filtered.columns]
    return filtered[available]


# ---------------------------------------------------------------------------
# PG team only PG gap rows builder (Track C2)
# ---------------------------------------------------------------------------

def _build_pg_team_only_pg_gap_rows(
    pg_team_df: pd.DataFrame | None,
    pg_team_cols: dict | None,
    prsa_df: pd.DataFrame | None,
    prsa_cols: dict | None,
    findings_df: pd.DataFrame | None,
) -> pd.DataFrame | None:
    """Synthesize Source - PG Gaps rows for PG-team-only Issue IDs.

    A PG-team-only issue is one whose Issue ID appears in the PG team inputs
    file but not in the prsa_df PG-flagged rows. AE + L2 are resolved by
    joining the row's Archer eGRC FND ID to findings_df. When a FND_ID
    resolves to multiple (AE, L2) pairs, the AE IDs and L2s are comma-joined
    into a single row (mirrors the existing tab's per-issue grain).

    Returns None when ``pg_team_df`` is None or empty. Returns an empty
    DataFrame when there are no PG-team-only Issue IDs to synthesize.
    """
    if pg_team_df is None or pg_team_df.empty:
        return None

    pg_team_cols = pg_team_cols or {}
    prsa_cols = prsa_cols or {}
    pg_issue_id_col = pg_team_cols.get("issue_id", "Issue ID (Archer IRM)")
    pg_finding_id_col = pg_team_cols.get("finding_id", "Archer eGRC FND ID")
    pg_impact_rating_col = pg_team_cols.get("impact_rating", "Impact Rating")

    issue_id_out = prsa_cols.get("issue_id", "Issue ID")
    issue_title_out = prsa_cols.get("issue_title", "Issue Title")
    issue_desc_out = prsa_cols.get("issue_description", "Issue Description")
    issue_status_out = prsa_cols.get("issue_status", "Issue Status")
    issue_rating_out = prsa_cols.get("issue_rating", "Issue Rating")
    risk_l2_out = prsa_cols.get("risk_level_2", "Risk Level 2")
    is_pg_col_out = prsa_cols.get("is_pg_gap", "Is PG Gap")

    # Build the set of Issue IDs present in prsa_df PG-flagged rows.
    prsa_issue_id_col = prsa_cols.get("issue_id", "Issue ID")
    is_pg_col = prsa_cols.get("is_pg_gap", "Is PG Gap")
    prsa_pg_issue_ids: set[str] = set()
    if prsa_df is not None and not prsa_df.empty \
            and is_pg_col in prsa_df.columns \
            and prsa_issue_id_col in prsa_df.columns:
        pg_mask = prsa_df[is_pg_col].map(
            lambda v: bool(v) if isinstance(v, bool)
            else str(v).strip().lower() in ("yes", "true", "1")
        )
        prsa_pg_issue_ids = set(
            prsa_df.loc[pg_mask, prsa_issue_id_col].astype(str).str.strip().tolist()
        )

    # FND_ID -> [(entity_id, l2_risk), ...] from the exploded findings_df.
    fnd_to_pairs: dict[str, list[tuple[str, str]]] = {}
    if findings_df is not None and not findings_df.empty \
            and "issue_id" in findings_df.columns:
        for _, frow in findings_df.iterrows():
            fid = str(frow.get("issue_id", "")).strip()
            if not fid or fid.lower() in ("nan", "none"):
                continue
            eid = str(frow.get("entity_id", "")).strip()
            l2 = str(frow.get("l2_risk", "")).strip()
            if not eid or not l2:
                continue
            fnd_to_pairs.setdefault(fid, []).append((eid, l2))

    rows: list[dict] = []
    seen_issue_ids: set[str] = set()
    for _, row in pg_team_df.iterrows():
        iid = str(row.get(pg_issue_id_col, "")).strip()
        if not iid or iid.lower() in ("nan", "none"):
            continue
        if iid in prsa_pg_issue_ids:
            continue
        if iid in seen_issue_ids:
            continue
        seen_issue_ids.add(iid)
        fid = str(row.get(pg_finding_id_col, "")).strip()
        pairs = fnd_to_pairs.get(fid, [])
        ae_ids = sorted({p[0] for p in pairs})
        l2s = sorted({p[1] for p in pairs})
        rating = str(row.get(pg_impact_rating_col, "")).strip()
        if rating.lower() in ("nan", "none"):
            rating = ""
        rows.append({
            issue_id_out: iid,
            issue_rating_out: rating,
            issue_status_out: "",
            issue_title_out: "(PG team gap — no PRSA record)",
            issue_desc_out: "",
            risk_l2_out: ", ".join(l2s),
            is_pg_col_out: "Yes",
            "AE ID": ", ".join(ae_ids),
        })

    if not rows:
        return pd.DataFrame()

    out = pd.DataFrame(rows)
    return out


# ---------------------------------------------------------------------------
# Methodology tab builder
# ---------------------------------------------------------------------------

def _build_methodology_data() -> dict[str, list[list[str]]]:
    """Load methodology content from YAML, route sections to their target tab.

    Returns a dict keyed by tab name. Each section's `tab:` field controls
    where it lands; missing field defaults to 'Methodology'. Tabs with no
    sections are not emitted.

    Sections may carry either `rows:` (legacy 2-col table) or `body:`
    (multi-paragraph prose). Body paragraphs are split on blank lines;
    bullet lines (leading "- ") are emitted as separate rows prefixed
    with a bullet character so Excel/HTML render them as a list.
    """
    yaml_path = Path(__file__).parent / "methodology.yaml"
    with open(yaml_path, "r", encoding="utf-8") as f:
        content = yaml.safe_load(f)

    by_tab: dict[str, list[list[str]]] = {}
    for section in content["sections"]:
        tab = section.get("tab", "Methodology")
        title = section.get("title", "")
        header = section.get("header")
        rows = section.get("rows", [])
        body = section.get("body")

        bucket = by_tab.setdefault(tab, [])
        bucket.append([title, ""])
        if header:
            bucket.append(header)
        for row in rows:
            bucket.append(row)
        if body:
            for para in _split_body_paragraphs(body):
                bucket.append(["", para])
        bucket.append(["", ""])

    return by_tab


_LABELED_PARA_PREFIXES = (
    "Scope.", "Attribution.", "Interpretation.", "Use.", "Caveats.",
    "Failure modes", "Source-specific failure mode",
)


def _split_body_paragraphs(body: str) -> list[str]:
    """Split a methodology `body:` block into renderable paragraph chunks.

    Each line starting with "- " (after stripping) becomes its own bullet
    paragraph (prefixed with the Unicode bullet so Excel and the HTML reader
    can render them uniformly). Lines starting with a labeled prefix
    (Scope., Attribution., Interpretation., Use., Caveats., Failure modes,
    Source-specific failure mode) start a new paragraph even without a blank
    line above. Other consecutive non-blank lines are joined into a single
    paragraph; blank lines separate paragraphs.
    """
    paragraphs: list[str] = []
    current: list[str] = []

    def _flush() -> None:
        if current:
            paragraphs.append(" ".join(current))
            current.clear()

    for raw_line in body.splitlines():
        line = raw_line.strip()
        if not line:
            _flush()
            continue
        if line.startswith("- "):
            _flush()
            paragraphs.append("• " + line[2:].strip())
            continue
        if any(line.startswith(p) for p in _LABELED_PARA_PREFIXES):
            _flush()
            current.append(line)
            continue
        current.append(line)
    _flush()
    return paragraphs


# ---------------------------------------------------------------------------
# Main export function
# ---------------------------------------------------------------------------

def export_results(
    transformed_df: pd.DataFrame,
    legacy_df: pd.DataFrame,
    output_path: str,
    findings_df: pd.DataFrame = None,
    key_risks_df: pd.DataFrame = None,
    findings_path: str = None,
    findings_cols: dict = None,
    entity_id_col: str = "Audit Entity",
    findings_index: dict | None = None,
    rco_overrides: dict | None = None,
    ore_df: pd.DataFrame = None,
    ore_irm_source_df: pd.DataFrame = None,
    ore_irm_index: dict | None = None,
    pillar_columns: dict | None = None,
    prsa_df: pd.DataFrame = None,
    prsa_cols: dict | None = None,
    pg_team_df: pd.DataFrame = None,
    pg_team_cols: dict | None = None,
    pg_team_diagnostics: dict | None = None,
    bma_df: pd.DataFrame = None,
    bma_cols: dict | None = None,
    gra_raps_df: pd.DataFrame = None,
    gra_raps_cols: dict | None = None,
    unmapped_findings: dict | None = None,
    unmapped_mapper_items: dict | None = None,
    key_inventory: dict | None = None,
    l2_taxonomy_df: pd.DataFrame = None,
    upstream_orphans_df: pd.DataFrame | None = None,
    provenance: dict | None = None,
):
    """Write multi-sheet Excel output."""
    logger.info(f"Writing output to {output_path}")

    # --- Audit Review (primary workspace) ---
    audit_df = build_audit_review_df(
        transformed_df, legacy_df, entity_id_col,
        unmapped_findings=unmapped_findings,
        unmapped_mapper_items=unmapped_mapper_items,
    )

    # --- Sheet 3: Review Queue (redesigned) ---
    review_df = build_review_queue_df(transformed_df)

    # --- Sheet 4: Side-by-side (debugging) ---
    # Diagnostic columns Confidence / Decision Type / Method previously lived
    # in Audit_Review (hidden) so the HTML report could read them; now they're
    # Side_by_Side-only. The HTML report merges Method/Decision Type back onto
    # the audit data via (entity_id, new_l2) before JSON serialization.
    from risk_taxonomy_transformer.review_builders import _derive_decision_type
    trace_cols = [
        "composite_key", "entity_id", "new_l1", "new_l2",
        "inherent_risk_rating", "inherent_risk_rating_label", "overall_impact",
        "likelihood", "impact_financial", "impact_reputational",
        "impact_consumer_harm", "impact_regulatory",
        "control_effectiveness_baseline", "impact_of_issues",
        "source_legacy_pillar", "source_risk_rating_raw", "source_rationale",
        "source_control_raw", "source_control_rationale",
        "mapping_type", "confidence", "method", "decision_type",
        "dims_parsed_from_rationale", "key_risk_evidence", "needs_review",
        "control_flag", "app_flag", "tp_flag", "model_flag", "core_flag", "aux_flag", "cross_boundary_flag",
    ]
    transformed_df = transformed_df.copy()
    if "method" in transformed_df.columns:
        transformed_df["decision_type"] = transformed_df["method"].apply(_derive_decision_type)
    available_trace_cols = [c for c in trace_cols if c in transformed_df.columns]
    trace_df = transformed_df[available_trace_cols].copy()

    # Reorder Side_by_Side rows to match Audit_Review's row order. The
    # Dashboard's cross-sheet COUNTIFS formulas (status range from Audit_Review
    # × method range from Side_by_Side) require row-by-row alignment to
    # correlate the two columns on the same logical record. Audit_Review is
    # sorted by review priority in build_audit_review_df; Side_by_Side
    # otherwise inherits transformed_df's per-entity per-pillar order.
    if "Entity ID" in audit_df.columns and "New L2" in audit_df.columns \
            and "entity_id" in trace_df.columns and "new_l2" in trace_df.columns:
        audit_order = audit_df[["Entity ID", "New L2"]].rename(
            columns={"Entity ID": "entity_id", "New L2": "new_l2"}
        ).reset_index(drop=True)
        audit_order["_audit_order"] = range(len(audit_order))
        trace_df = trace_df.merge(audit_order, on=["entity_id", "new_l2"], how="left")
        trace_df = trace_df.sort_values(
            "_audit_order", kind="mergesort"
        ).drop(columns=["_audit_order"]).reset_index(drop=True)

    # Build Methodology tabs (one section per tab key returned by the loader).
    # Visible audit-leader tab is "Methodology"; "Methodology Detail" and
    # "RCO Methodology" are emitted but hidden by default below.
    methodology_by_tab = _build_methodology_data()
    if provenance:
        prov_rows = [["Run Provenance", ""]]
        prov_rows += [
            ["Tool commit", provenance.get("tool_commit", "unknown")],
            ["Run timestamp", provenance.get("run_timestamp", "")],
            ["spaCy model",
             f"{provenance.get('spacy_model', 'n/a')} "
             f"({provenance.get('spacy_model_version', 'unknown')})"],
            ["Library versions",
             f"Python {provenance.get('python_version', '')} · "
             f"pandas {provenance.get('pandas_version', '')} · "
             f"openpyxl {provenance.get('openpyxl_version', '')} · "
             f"PyYAML {provenance.get('pyyaml_version', '')} · "
             f"spaCy {provenance.get('spacy_version', '')}"],
            ["", ""],
        ]
        methodology_by_tab.setdefault("Methodology", [])
        methodology_by_tab["Methodology"] = prov_rows + methodology_by_tab["Methodology"]
    methodology_dfs = {
        tab: pd.DataFrame(rows, columns=["Topic", "Detail"])
        for tab, rows in methodology_by_tab.items()
    }

    # Write sheets
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # Visible tabs first
        audit_df.to_excel(writer, sheet_name="Audit_Review", index=False)
        for tab_name, mdf in methodology_dfs.items():
            mdf.to_excel(writer, sheet_name=tab_name, index=False, header=False)
        # Hidden tabs
        review_df.to_excel(writer, sheet_name="Review_Queue", index=False)
        trace_df.to_excel(writer, sheet_name="Side_by_Side", index=False)
        legacy_df.to_excel(writer, sheet_name="Source - Legacy Data", index=False, startrow=1)
        if findings_path and findings_cols:
            enriched_findings = _enrich_findings_source(
                findings_path, findings_cols, transformed_df)
            enriched_findings.to_excel(writer, sheet_name="Source - Findings", index=False, startrow=1)
        elif findings_df is not None and not findings_df.empty:
            findings_df.to_excel(writer, sheet_name="Source - Findings", index=False, startrow=1)
        if key_risks_df is not None and not key_risks_df.empty:
            enriched_key_risks = _enrich_key_risks_source(key_risks_df, transformed_df)
            enriched_key_risks.to_excel(writer, sheet_name="Source - Key Risks", index=False, startrow=1)
        if ore_df is not None and not ore_df.empty:
            # Rename internal lowercase columns back to user-friendly display names.
            # NB: ore_df is exploded one row per (ORE × L2). The original mapper
            # output already has a "Mapped L2s" column (semicolon-joined list of
            # all L2s for the ORE) — we rename the per-row exploded `l2_risk`
            # column to "Canonical L2" so it doesn't collide.
            _ore_cfg = get_config().get("columns", {}).get("ore_mappings", {})
            _ore_rename = {
                "event_id": _ore_cfg.get("event_id", "Event ID"),
                "entity_id": _ore_cfg.get("entity_id", "Audit Entity (Operational Risk Events)"),
                "l2_risk": "Canonical L2",
            }
            ore_out = ore_df.rename(columns={
                k: v for k, v in _ore_rename.items() if k in ore_df.columns
            })
            ore_out.to_excel(writer, sheet_name="Source - OREs", index=False, startrow=1)

        # --- Source - ORE IRM tab (Track B for IRM OREs) ---
        # Per Lu's architecture: separate tab next to Source - OREs, top-row
        # disclosure (row 1 merged), header at row 2, data starts row 3.
        if ore_irm_source_df is not None and not ore_irm_source_df.empty:
            _ore_irm_yaml = get_config().get("columns", {}).get("ore_irm", {})
            irm_out_cols_in_order = [
                _ore_irm_yaml.get("ore_id", "ORE ID"),
                _ore_irm_yaml.get("ore_title", "ORE Title"),
                _ore_irm_yaml.get("capture_status", "Capture Status"),
                _ore_irm_yaml.get("rca_status", "RCA Status"),
                _ore_irm_yaml.get("impact_assessment_status", "Impact Assessment Status"),
                _ore_irm_yaml.get("stop_ongoing_impact_status", "Stop Ongoing Impact Status"),
                _ore_irm_yaml.get("ore_category", "ORE Category"),
                "ORE Status",
                _ore_irm_yaml.get("identified_by", "Identified By"),
                _ore_irm_yaml.get("identified_by_subgroup", "Identified By Sub-Group"),
                _ore_irm_yaml.get("ore_description", "ORE Description"),
                _ore_irm_yaml.get("ore_root_cause", "ORE Root Cause"),
                _ore_irm_yaml.get("root_cause_description", "Root Cause Description"),
                _ore_irm_yaml.get("root_cause_level_1", "Root Cause Level 1"),
                _ore_irm_yaml.get("root_cause_level_2", "Root Cause Level 2"),
                _ore_irm_yaml.get("risk_level_2", "Risk Level 2"),
                _ore_irm_yaml.get("risk_level_4", "Risk Level 4"),
                _ore_irm_yaml.get("remediation_id", "Remediation ID"),
                _ore_irm_yaml.get("legacy_event_id", "Legacy Event ID"),
            ]
            irm_out = ore_irm_source_df.copy()
            # Tool-added L2 Source column from L2 Provenance (matches PRSA convention).
            if "L2 Provenance" in irm_out.columns:
                _irm_l2_source_label = {"source": "IRM Archer", "mapper": "Inferred"}
                irm_out["L2 Source"] = irm_out["L2 Provenance"].map(
                    lambda v: _irm_l2_source_label.get(str(v).strip().lower(), "")
                )

            # Synthesize Mapped L2s + Mapping Status per ORE from the index so
            # the HTML drill-down can resolve which L2s an IRM ORE maps to
            # without re-reading the mapper output xlsx.
            ore_to_l2s: dict[str, set[str]] = {}
            ore_to_mstatus: dict[str, str] = {}
            for _eid, by_l2 in (ore_irm_index or {}).items():
                for l2_name, items in by_l2.items():
                    for item in items:
                        oid = str(item.get("event_id", "")).strip()
                        if oid:
                            ore_to_l2s.setdefault(oid, set()).add(l2_name)
                            mstat = str(item.get("mapping_status", "")).strip()
                            if mstat and oid not in ore_to_mstatus:
                                ore_to_mstatus[oid] = mstat
            ore_id_col_irm = _ore_irm_yaml.get("ore_id", "ORE ID")
            if ore_id_col_irm in irm_out.columns:
                irm_out["Mapped L2s"] = irm_out[ore_id_col_irm].map(
                    lambda oid: "; ".join(sorted(ore_to_l2s.get(str(oid).strip(), set())))
                )
                irm_out["Mapping Status"] = irm_out[ore_id_col_irm].map(
                    lambda oid: ore_to_mstatus.get(str(oid).strip(), "")
                )

            available = [c for c in irm_out_cols_in_order if c in irm_out.columns]
            for tool_col in ("L2 Source", "Mapped L2s", "Mapping Status"):
                if tool_col in irm_out.columns:
                    available.append(tool_col)
            irm_out = irm_out[available]
            # startrow=1 leaves row 1 free for the disclosure banner.
            irm_out.to_excel(
                writer, sheet_name="Source - ORE IRM", index=False, startrow=1
            )
        if prsa_df is not None and not prsa_df.empty:
            # Surface the Track B provenance column as "L2 Source" next to the
            # L2 columns. Recase the internal sentinels ('source' / 'mapper')
            # to user-facing values ('IRM Archer' / 'Inferred'). Blank stays
            # blank so rows with no L2 don't get a label.
            prsa_out = prsa_df.copy()
            if "L2 Provenance" in prsa_out.columns:
                prsa_out = prsa_out.rename(columns={"L2 Provenance": "L2 Source"})
                _l2_source_label = {"source": "IRM Archer", "mapper": "Inferred"}
                prsa_out["L2 Source"] = prsa_out["L2 Source"].map(
                    lambda v: _l2_source_label.get(str(v).strip().lower(), "")
                )
                # Reposition next to Mapped L2s when present
                cols = list(prsa_out.columns)
                cols.remove("L2 Source")
                if "Mapped L2s" in cols:
                    insert_at = cols.index("Mapped L2s") + 1
                else:
                    insert_at = len(cols)
                cols.insert(insert_at, "L2 Source")
                prsa_out = prsa_out[cols]
            # Track B left this tab as one row per (mapped) AE × Issue × Control.
            # Track C does NOT add unmapped PG rows here — those go to the new
            # Source - PG Gaps tab. Filter explicitly to AE-populated rows so
            # the synthesized blank-AE PG-unmapped rows from the Frankenstein
            # don't leak into PRSA's tab.
            is_pg_col = (prsa_cols or {}).get("is_pg_gap", "Is PG Gap")
            ae_id_col_local = (prsa_cols or {}).get("ae_id", "AE ID")
            if ae_id_col_local in prsa_out.columns:
                prsa_out = prsa_out[
                    prsa_out[ae_id_col_local].astype(str).str.strip() != ""
                ]
            # Display header rename: "Is PG Gap" -> "PG Gap" (parallels other
            # source labels). Value scheme: "Yes" for PG-flagged rows, blank
            # otherwise — audit teams filter "non-blank" to isolate flagged
            # rows in two clicks. Reposition the column to the rightmost slot
            # so it's easy to find via filter dropdown without scrolling.
            if is_pg_col in prsa_out.columns:
                prsa_out[is_pg_col] = prsa_out[is_pg_col].map(
                    lambda v: "Yes" if (
                        v is True
                        or str(v).strip().lower() in ("yes", "true", "1")
                    ) else ""
                )
                prsa_out = prsa_out.rename(columns={is_pg_col: "PG Gap"})
                cols = list(prsa_out.columns)
                cols.remove("PG Gap")
                cols.append("PG Gap")
                prsa_out = prsa_out[cols]
            prsa_out.to_excel(writer, sheet_name="Source - PRSA Issues", index=False, startrow=1)

            # Track C: Source - PG Gaps tab. Filtered to PG-flagged rows
            # (mapped + unmapped). Mapped PG rows duplicate into both this tab
            # and the PRSA tab — Lu-confirmed acceptable trade-off so PG gaps
            # surface as their own evidence type.
            if is_pg_col in prsa_df.columns:
                pg_mask = prsa_df[is_pg_col].map(
                    lambda v: bool(v) if isinstance(v, bool)
                    else str(v).strip().lower() in ("yes", "true", "1")
                )
                pg_only = prsa_df.loc[pg_mask].copy()
                issue_id_col_local = (prsa_cols or {}).get("issue_id", "Issue ID")
                issue_title_col_local = (prsa_cols or {}).get("issue_title", "Issue Title")
                issue_desc_col_local = (prsa_cols or {}).get("issue_description", "Issue Description")
                issue_status_col_local = (prsa_cols or {}).get("issue_status", "Issue Status")
                issue_rating_col_local = (prsa_cols or {}).get("issue_rating", "Issue Rating")
                risk_l2_col_local = (prsa_cols or {}).get("risk_level_2", "Risk Level 2")
                desired = [
                    issue_id_col_local,
                    issue_rating_col_local,
                    issue_status_col_local,
                    issue_title_col_local,
                    issue_desc_col_local,
                    risk_l2_col_local,
                    is_pg_col,
                ]
                if not pg_only.empty:
                    # Every row on this tab is a PG gap, so all rows get "Yes".
                    # The column stays as a sanity-check for copy-paste workflows.
                    pg_only[is_pg_col] = "Yes"
                    # Per-issue dedup — Frankenstein grain is AE × Issue × Control,
                    # but the PG Gaps tab is per-issue.
                    if issue_id_col_local in pg_only.columns:
                        pg_only = pg_only.drop_duplicates(
                            subset=[issue_id_col_local], keep="first"
                        )
                    pg_cols = [c for c in desired if c in pg_only.columns]
                    pg_only = pg_only[pg_cols]

                # Track C2: append synthesized rows for PG-team-only Issue IDs
                # (Issue IDs in pg_team_df absent from prsa_df PG-flagged rows).
                # Each synthesized row resolves AE + L2 via findings_df by FND_ID.
                pg_team_only_rows = _build_pg_team_only_pg_gap_rows(
                    pg_team_df, pg_team_cols, prsa_df, prsa_cols, findings_df,
                )
                if pg_team_only_rows is not None and not pg_team_only_rows.empty:
                    if not pg_only.empty:
                        pg_only = pd.concat(
                            [pg_only, pg_team_only_rows], ignore_index=True, sort=False
                        )
                    else:
                        pg_only = pg_team_only_rows

                if not pg_only.empty:
                    # Display header rename: "Is PG Gap" -> "PG Gap" so it
                    # matches the PRSA tab and the chip label.
                    if is_pg_col in pg_only.columns:
                        pg_only = pg_only.rename(columns={is_pg_col: "PG Gap"})
                    # Write starting at row 1 (zero-indexed) so row 0 is
                    # available for the merged disclosure row added below.
                    pg_only.to_excel(
                        writer, sheet_name="Source - PG Gaps", index=False, startrow=1
                    )
        if bma_df is not None and not bma_df.empty:
            bma_df.to_excel(writer, sheet_name="Source - BM Activities", index=False, startrow=1)
        if gra_raps_df is not None and not gra_raps_df.empty:
            gra_raps_df.to_excel(writer, sheet_name="Source - GRA RAPs", index=False, startrow=1)
        if l2_taxonomy_df is not None and not l2_taxonomy_df.empty:
            l2_taxonomy_df.to_excel(writer, sheet_name="Source - L2 Taxonomy", index=False, startrow=1)

        # --- Source - Models tab ---
        # Loaded from the inventory file (data/input/model_inventory_*.xlsx)
        # and filtered to IDs referenced in the legacy `Models`
        # column. Mirrors the HTML report drill-down inventory section.
        _cfg_columns = get_config().get("columns", {})
        _legacy_models_col = _cfg_columns.get("applications", {}).get("models", "Models")
        _model_inv_cfg = _cfg_columns.get("model_inventory", {})
        _model_inv_pattern = _cfg_columns.get("inventory_files", {}).get("models", "model_inventory_*.xlsx")
        models_src_df = _build_models_source_df(
            legacy_df, _legacy_models_col, _model_inv_cfg, _model_inv_pattern,
        )
        if not models_src_df.empty:
            models_src_df.to_excel(writer, sheet_name="Source - Models", index=False, startrow=1)
        # Key Inventory (hidden) — per-entity "key" app/TP ID sets aggregated
        # from key risks. Non-key items do not drive risk per procedure;
        # HTML report reads this sheet to mark key IDs in drill-down and
        # Inventory views.
        if key_inventory:
            import json as _json
            ki_rows = []
            for eid, sets in key_inventory.items():
                apps_kpa = sets.get("key_apps_kpa", {})
                tps_kpa = sets.get("key_tps_kpa", {})
                # Serialize the per-ID KPA mapping as JSON so the HTML reader
                # can parse it. Sort KPA ids within each list for stable output.
                apps_kpa_json = _json.dumps(
                    {aid: sorted(k) for aid, k in apps_kpa.items()},
                    sort_keys=True,
                )
                tps_kpa_json = _json.dumps(
                    {tid: sorted(k) for tid, k in tps_kpa.items()},
                    sort_keys=True,
                )
                ki_rows.append({
                    "Entity ID": eid,
                    "Key Apps": "; ".join(sorted(sets.get("key_apps", set()))),
                    "Key TPs": "; ".join(sorted(sets.get("key_tps", set()))),
                    "Orphan Apps": "; ".join(sorted(sets.get("orphan_apps", set()))),
                    "Orphan TPs": "; ".join(sorted(sets.get("orphan_tps", set()))),
                    "Key Apps KPA JSON": apps_kpa_json,
                    "Key TPs KPA JSON": tps_kpa_json,
                })
            if ki_rows:
                ki_df = pd.DataFrame(ki_rows)
                ki_df.to_excel(writer, sheet_name="Key_Inventory", index=False)
        if pillar_columns:
            legacy_lookup = _build_legacy_lookup(legacy_df, pillar_columns, entity_id_col)
            legacy_lookup.to_excel(writer, sheet_name="Legacy Ratings Lookup", index=False)

        # --- Upstream Tagging Gaps tab ---
        # One row per dropped/orphan item across findings, OREs (legacy/IRM),
        # PRSA, GRA RAPs, and BMA. Always written (even empty), so reviewers
        # can confirm "no orphans this run" rather than guessing.
        if upstream_orphans_df is not None:
            uog = upstream_orphans_df.copy()
            if uog.empty:
                # Write a single placeholder row so the banner + headers render
                # and reviewers can see "no orphans this run" explicitly.
                uog = pd.DataFrame([{
                    "Source": "(none)",
                    "Item ID": "",
                    "Title": "No orphans captured this run.",
                    "Status": "",
                    "Drop Reason": "",
                    "Source File": "",
                }])
            uog.to_excel(writer, sheet_name="Upstream Tagging Gaps", index=False, startrow=1)

        # --- Risk Owner Review tab ---
        ro_review_df = build_risk_owner_review_df(
            transformed_df, legacy_df, entity_id_col,
            findings_index=findings_index,
            rco_overrides=rco_overrides,
        )
        # Build summary before dropping internal columns (summary uses _priority)
        ro_summary_df = build_ro_summary_df(ro_review_df, findings_index=findings_index)
        # Drop internal columns before writing to Excel
        ro_review_clean = ro_review_df.drop(columns=[c for c in ro_review_df.columns if c.startswith("_")])
        ro_review_clean.to_excel(writer, sheet_name="Risk_Owner_Review", index=False)

        # --- Risk Owner Summary tab ---
        ro_summary_df.to_excel(writer, sheet_name="Risk_Owner_Summary", index=False)

    # Apply formatting
    wb = load_workbook(output_path)

    # Status color fills
    green_fill = PatternFill("solid", fgColor="C6EFCE")
    gray_fill = PatternFill("solid", fgColor="D9D9D9")
    yellow_fill = PatternFill("solid", fgColor="FFFF00")
    blue_fill = PatternFill("solid", fgColor="BDD7EE")
    orange_fill = PatternFill("solid", fgColor="FCE4D6")
    status_fills = {
        Status.APPLICABLE: green_fill,
        Status.NOT_APPLICABLE: gray_fill,
        Status.NO_EVIDENCE: orange_fill,
        Status.UNDETERMINED: yellow_fill,
        Status.NOT_ASSESSED: blue_fill,
    }

    review_type_fills = {
        "Determine Applicability": yellow_fill,
        "Assumed N/A": orange_fill,
    }

    source_banners_cfg = _load_source_banners()

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Source tabs in _SOURCE_TAB_BANNER_KEYS were written with startrow=1
        # so row 1 is reserved for a merged disclosure banner. Header is row 2.
        # Style fill/border come from the YAML `style` field (warn vs info).
        banner_key = _SOURCE_TAB_BANNER_KEYS.get(sheet_name)
        if banner_key and ws.max_column > 0:
            from openpyxl.styles import Alignment, Border, Side, Font as _Font
            from openpyxl.utils import get_column_letter as _gcl
            cfg = source_banners_cfg.get(banner_key, {})
            disclosure = _strip_html(cfg.get("body", ""))
            style = cfg.get("style", "info")
            if style == "warn":
                bg_color, border_color = "FFF3CD", "FFAD1F"
            else:
                bg_color, border_color = "D1ECF1", "0C5460"
            ws.cell(row=1, column=1, value=disclosure)
            ws.merge_cells(start_row=1, start_column=1,
                           end_row=1, end_column=ws.max_column)
            banner_fill = PatternFill("solid", fgColor=bg_color)
            banner_border = Border(
                left=Side(style="thin", color=border_color),
                right=Side(style="thin", color=border_color),
                top=Side(style="thin", color=border_color),
                bottom=Side(style="thin", color=border_color),
            )
            ws.cell(row=1, column=1).fill = banner_fill
            ws.cell(row=1, column=1).border = banner_border
            ws.cell(row=1, column=1).alignment = Alignment(
                wrap_text=True, vertical="center", horizontal="left"
            )
            ws.row_dimensions[1].height = 36
            # Style the header row at row 2 (pandas wrote it via startrow=1).
            hdr_font = _Font(bold=True, color="FFFFFF", name="Arial")
            hdr_fill = PatternFill("solid", fgColor="2F5496")
            for col_idx in range(1, ws.max_column + 1):
                c = ws.cell(row=2, column=col_idx)
                c.font = hdr_font
                c.fill = hdr_fill
                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.freeze_panes = "A3"
            # Column widths: estimate from row 2 (header) + data rows (3+),
            # ignoring the merged disclosure on row 1.
            for col_idx in range(1, ws.max_column + 1):
                col_letter = _gcl(col_idx)
                max_len = len(str(ws.cell(row=2, column=col_idx).value or ""))
                for r in range(3, ws.max_row + 1):
                    v = ws.cell(row=r, column=col_idx).value
                    if v is not None:
                        try:
                            max_len = max(max_len, len(str(v)))
                        except (TypeError, ValueError):
                            pass
                ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 40)
            continue

        style_header(ws, ws.max_column)

        # Auto-width columns
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                except (TypeError, ValueError):
                    pass
            # Cap wider for text-heavy columns
            cap = 60 if sheet_name in ("Review_Queue", "Audit_Review") else 40
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), cap)

        # Audit_Review -- full reviewer worksheet formatting
        if sheet_name == "Audit_Review":
            _format_audit_review_sheet(ws, status_fills)

        # Color-code Review_Queue by Review Type
        if sheet_name == "Review_Queue":
            col = _find_header_column(ws, "Review Type")
            if col:
                _color_rows_by_column(ws, col, review_type_fills, match_contains=True)

        # Highlight needs_review rows in yellow on Side_by_Side
        if sheet_name == "Side_by_Side":
            col = _find_header_column(ws, "needs_review")
            if col:
                _color_rows_by_column(ws, col, {True: yellow_fill})

    # Format Methodology tabs (all three: Methodology, Methodology Detail,
    # RCO Methodology share the same styling rules).
    bold_font = Font(bold=True, size=11, name="Arial")
    title_font = Font(bold=True, size=14, name="Arial", color="2F5496")
    sub_header_font = Font(bold=True, size=10, name="Arial", color="2F5496")
    section_headers = {
        "PURPOSE", "STATUS VALUES", "CONFIDENCE LEVELS",
        "EVIDENCE SOURCES (in priority order)", "ADDITIONAL SIGNALS COLUMN",
        "RATING POLICY", "RATING SOURCE COLUMN",
        "CONTROL EFFECTIVENESS ASSESSMENT",
        "WHAT THE TOOL DOES NOT DO",
        "TABS IN THIS WORKBOOK",
        "FINDING FILTERS APPLIED",
        "CROSS-BOUNDARY KEYWORD THRESHOLD",
        "TOOL-COMPUTED COLUMNS (PRSA SOURCE TAB)",
        "L2 SOURCE COLUMN (PRSA SOURCE TAB)",
        "PG GAP SOURCE", "ORE IRM SOURCE",
        "DEDUPLICATION", "COMMON QUESTIONS",
        "RISK OWNER REVIEW \u2014 COLUMN GUIDE",
        "RISK OWNER REVIEW \u2014 HOW TO USE",
        "RISK OWNER REVIEW \u2014 PRIORITY SCORING",
    }
    sub_headers = {"Status", "Level", "Source", "Signal", "Value", "Tab", "Filter",
                   "Column", "Step", "Score", "Question", "Label"}

    for meth_tab in ("LUminate Methodology", "Methodology", "Methodology Detail", "RCO Methodology"):
        if meth_tab not in wb.sheetnames:
            continue
        ws = wb[meth_tab]
        ws.column_dimensions["A"].width = 45
        ws.column_dimensions["B"].width = 120

        # Prose-style sections (body: in YAML) emit rows where col A is blank
        # and col B carries paragraph text. Wrap text + grow row height so the
        # full paragraph is visible without manually resizing.
        from openpyxl.styles import Alignment as _Alignment
        wrap_align = _Alignment(vertical="top", wrap_text=True)

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            cell_val = str(row[0].value or "")
            body_val = str(row[1].value or "")
            if cell_val.startswith("LUminate"):
                row[0].font = title_font
            elif cell_val in section_headers:
                row[0].font = bold_font
            elif cell_val in sub_headers:
                row[0].font = sub_header_font
                row[1].font = sub_header_font
            elif not cell_val and body_val:
                # Prose paragraph row — wrap in col B, give it room.
                row[1].alignment = wrap_align
                # Rough height estimate at ~110 chars per row at width 120.
                est_lines = max(1, (len(body_val) // 110) + body_val.count("\n") + 1)
                ws.row_dimensions[row[1].row].height = max(18, est_lines * 16)

        # On LUminate Methodology specifically, treat each section's title
        # row as a section header (bold) so reviewers can scan source by source.
        if meth_tab == "LUminate Methodology":
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
                cell_val = str(row[0].value or "")
                if not cell_val:
                    continue
                if cell_val == "LUminate Methodology":
                    row[0].font = title_font
                else:
                    row[0].font = bold_font

    # --- Build Dashboard tab ---
    ar_ws = wb["Audit_Review"]
    _build_dashboard_sheet(wb, ar_ws)

    # --- Format Legacy Ratings Lookup tab ---
    if "Legacy Ratings Lookup" in wb.sheetnames:
        ll_ws = wb["Legacy Ratings Lookup"]
        style_header(ll_ws, ll_ws.max_column)
        ll_ws.column_dimensions["A"].width = 15   # Entity ID
        ll_ws.column_dimensions["B"].width = 25   # Risk Pillar
        ll_ws.column_dimensions["C"].width = 18   # Inherent Risk Rating
        ll_ws.column_dimensions["D"].width = 60   # Inherent Risk Rationale
        ll_ws.column_dimensions["E"].width = 20   # Control Assessment
        ll_ws.column_dimensions["F"].width = 60   # Control Assessment Rationale
        ll_ws.auto_filter.ref = ll_ws.dimensions

    # --- Format Risk_Owner_Review tab ---
    if "Risk_Owner_Review" in wb.sheetnames:
        _format_risk_owner_review_sheet(wb["Risk_Owner_Review"], status_fills)

    # --- Format Risk_Owner_Summary tab ---
    if "Risk_Owner_Summary" in wb.sheetnames:
        _format_risk_owner_summary_sheet(wb["Risk_Owner_Summary"])

    # --- Set tab visibility ---
    # Visible tabs (audit-leader-facing workspace + reference + source data
    # the tester actually used in the HTML view):
    #   Dashboard, Audit_Review, Methodology, Legacy Ratings Lookup,
    #   Source - * (all), Source - L2 Taxonomy
    # Hidden tabs (different audience or unreadable today):
    #   Review_Queue       — filtered subset of Audit_Review (the workspace)
    #   Side_by_Side       — debug-only traceability columns
    #   Risk_Owner_Summary — RCO audience, not audit leader
    #   Risk_Owner_Review  — RCO audience, not audit leader
    #   Key_Inventory      — JSON-serialized cells; programmatic feed for HTML.
    #                        The same key app/TP info is reviewer-readable in
    #                        Source - Key Risks (KEY PRIMARY & SECONDARY columns).
    hidden_tabs = [
        "Review_Queue", "Side_by_Side",
        "Risk_Owner_Summary", "Risk_Owner_Review",
        "Key_Inventory",
        # Detailed methodology + RCO methodology hidden by default;
        # audit leaders open the cut-down "Methodology" tab. Power users
        # can unhide either when they want the deeper docs.
        "Methodology Detail", "RCO Methodology",
    ]
    for tab_name in hidden_tabs:
        if tab_name in wb.sheetnames:
            wb[tab_name].sheet_state = "hidden"

    # --- Reorder tabs ---
    desired_order = [
        "Dashboard", "Audit_Review", "Legacy Ratings Lookup", "Methodology",
        "Risk_Owner_Summary", "Risk_Owner_Review",
        # Hidden tabs
        "Methodology Detail", "RCO Methodology",
        "Review_Queue", "Side_by_Side",
        # LUminate Methodology lands between Side_by_Side and the Source - *
        # block per Lu's spec: it's the first thing reviewers see when they
        # start interpreting source data.
        "LUminate Methodology",
        # Upstream Tagging Gaps surfaces orphans across sources (blank-AE
        # and IRM ORE bridge gaps). Sits at the head of the Source - * block
        # so reviewers see "what dropped out" before the per-source tabs.
        "Upstream Tagging Gaps",
        "Source - Legacy Data", "Source - Findings", "Source - Key Risks",
        "Source - OREs", "Source - ORE IRM",
        "Source - PRSA Issues", "Source - PG Gaps",
        "Source - BM Activities",
        "Source - GRA RAPs", "Source - Models", "Source - L2 Taxonomy",
    ]
    for i, name in enumerate(desired_order):
        if name in wb.sheetnames:
            current_idx = wb.sheetnames.index(name)
            wb.move_sheet(name, offset=i - current_idx)

    # --- Tool-added column header highlights ---
    # In source tabs, columns the tool ADDED to the source data get a blue
    # header tint to distinguish them from the original source columns —
    # mirrors the HTML report's `tool: true` styling.
    _tool_fill = PatternFill("solid", fgColor="E3F2FD")
    _tool_cols_by_tab = {
        "Source - Findings": {"Mapping Status", "Mapped To L2(s)"},
        "Source - Key Risks": {"L2 Keyword Matches"},
        "Source - PRSA Issues": {"Other AEs With This PRSA", "Mapped L2s", "Mapping Status", "PG Gap"},
        "Source - PG Gaps": {"PG Gap"},
        "Source - GRA RAPs": {"Mapped L2s", "Mapping Status"},
        "Source - ORE IRM": {"L2 Source", "Mapped L2s", "Mapping Status"},
        "Audit_Review": {
            "Status", "Inherent Risk Rating", "Decision Basis",
            "Additional Signals", "Impact of Issues", "Control Signals",
        },
    }
    for tab_name, tool_cols in _tool_cols_by_tab.items():
        if tab_name not in wb.sheetnames:
            continue
        ws = wb[tab_name]
        # Header row is row 2 for tabs with a merged disclosure banner at row 1,
        # row 1 elsewhere.
        header_row_idx = 2 if tab_name in _SOURCE_TAB_BANNER_KEYS else 1
        for cell in ws[header_row_idx]:
            if cell.value in tool_cols:
                cell.fill = _tool_fill

    wb.save(output_path)
    logger.info(f"  Output saved: {output_path}")
    logger.info(f"  Sheets: {wb.sheetnames}")
