"""
Excel formatting functions for the Risk Taxonomy Transformer.

All openpyxl styling: header styles, row coloring, column grouping,
and per-sheet formatting for the multi-tab output workbook.
"""

from __future__ import annotations

from datetime import datetime

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from risk_taxonomy_transformer.constants import Status


# ---------------------------------------------------------------------------
# Shared helpers
# ---------------------------------------------------------------------------

def _find_header_column(ws, header_name: str) -> int | None:
    """Find the 1-based column index of a header by name, or None."""
    for cell in ws[1]:
        if cell.value == header_name:
            return cell.column
    return None


def _color_rows_by_column(ws, col_index: int, value_to_fill: dict,
                          match_contains: bool = False):
    """Color entire rows based on the value in a specific column.

    If match_contains is True, checks if any key is a substring of the cell value.
    """
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        raw_val = row[col_index - 1].value
        cell_val = str(raw_val or "")
        fill = None
        if match_contains:
            for key, f in value_to_fill.items():
                if key in cell_val:
                    fill = f
                    break
        else:
            # Try raw value first (for bool/int keys), then stringified
            fill = value_to_fill.get(raw_val) or value_to_fill.get(cell_val)
        if fill:
            for cell in row:
                cell.fill = fill


def style_header(ws, max_col: int):
    """Apply header styling to first row."""
    header_font = Font(bold=True, color="FFFFFF", size=10, name="Arial")
    header_fill = PatternFill("solid", fgColor="2F5496")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    for col_idx in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border


# ---------------------------------------------------------------------------
# Per-sheet formatters (Phase 5 extractions)
# ---------------------------------------------------------------------------

def _format_audit_review_sheet(ws, status_fills: dict):
    """Apply full formatting to the Audit_Review worksheet."""
    header_row = 1
    data_start = 2

    # --- Freeze panes: freeze first 2 columns + header row ---
    ws.freeze_panes = f"C{data_start}"

    # --- Auto-filter on data range ---
    last_col_letter = get_column_letter(ws.max_column)
    ws.auto_filter.ref = f"A{header_row}:{last_col_letter}{ws.max_row}"

    # --- Column widths ---
    col_widths = {
        "Entity ID": 12, "Entity Name": 25, "Entity Overview": 40,
        "Audit Leader": 15, "PGA": 12, "Core Audit Team": 18,
        "New L1": 20, "New L2": 30,
        "Proposed Status": 22, "Proposed Rating": 16,
        "Confidence": 12, "Legacy Source": 18,
        "Decision Basis": 60, "Additional Signals": 50,
        "Source Rationale": 60, "Control Signals": 50,
        "Control Effectiveness Baseline": 22, "Impact of Issues": 20,
        "Source Control Rationale": 40,
        "Reviewer Status": 22, "Reviewer Rating Override": 18, "Reviewer Notes": 40,
    }
    for cell in ws[header_row]:
        if cell.value in col_widths:
            ws.column_dimensions[cell.column_letter].width = col_widths[cell.value]

    # Text wrap for long-text columns
    wrap_align = Alignment(wrap_text=True, vertical="top")
    for col_name in ("Decision Basis", "Additional Signals", "Source Rationale",
                     "Control Signals", "Source Control Rationale",
                     "Impact of Issues", "Reviewer Notes", "Entity Overview"):
        col_idx = None
        for cell in ws[header_row]:
            if cell.value == col_name:
                col_idx = cell.column
                break
        if col_idx:
            for row_idx in range(data_start, ws.max_row + 1):
                ws.cell(row=row_idx, column=col_idx).alignment = wrap_align

    # --- Color-code by Proposed Status ---
    status_col = None
    for cell in ws[header_row]:
        if cell.value == "Proposed Status":
            status_col = cell.column
            break
    if status_col:
        for row_idx in range(data_start, ws.max_row + 1):
            status_val = ws.cell(row=row_idx, column=status_col).value
            fill = status_fills.get(status_val)
            if fill:
                ws.cell(row=row_idx, column=status_col).fill = fill

    # --- Status tier formatting: left border colors ---
    if status_col:
        status_borders = {
            Status.UNDETERMINED: Side(style="thick", color="E8923C"),
            Status.NO_EVIDENCE: Side(style="thick", color="FFC107"),
        }
        for row_idx in range(data_start, ws.max_row + 1):
            status_val = ws.cell(row=row_idx, column=status_col).value
            border_side = status_borders.get(status_val)
            if border_side:
                ws.cell(row=row_idx, column=1).border = Border(left=border_side)

    # --- Reviewer column formatting ---
    reviewer_fill = PatternFill("solid", fgColor="E2EFDA")
    reviewer_cols = []
    for cell in ws[header_row]:
        if cell.value in ("Reviewer Status", "Reviewer Rating Override", "Reviewer Notes"):
            cell.fill = reviewer_fill
            reviewer_cols.append(cell.column)

    # --- Column grouping: hide detail columns ---
    # Group 1: Source Control Rationale (control detail)
    # Group 2: Rating Source through Impact - Regulatory (rating detail)
    # Group 3: L2 Definition (reference)
    hide_col_names = [
        "Source Control Rationale",
        "Rating Source", "Source Rating", "Likelihood", "Overall Impact",
        "Impact - Financial", "Impact - Reputational",
        "Impact - Consumer Harm", "Impact - Regulatory",
        "L2 Definition",
    ]
    for col_name in hide_col_names:
        col_idx = _find_header_column(ws, col_name)
        if col_idx:
            cl = get_column_letter(col_idx)
            ws.column_dimensions[cl].outlineLevel = 1
            ws.column_dimensions[cl].hidden = True

    # --- Row height for readable text ---
    for row_idx in range(data_start, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 45

    # --- Entity group separators: top border on first row of new entity ---
    entity_col = 1  # Entity ID is column A
    prev_entity = None
    for row_idx in range(data_start, ws.max_row + 1):
        current_entity = ws.cell(row=row_idx, column=entity_col).value
        if prev_entity is not None and current_entity != prev_entity:
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                existing = cell.border
                cell.border = Border(
                    top=Side(style="medium", color="2F5496"),
                    left=existing.left,
                    right=existing.right,
                    bottom=existing.bottom,
                )
        prev_entity = current_entity


def _format_risk_owner_review_sheet(ws, status_fills: dict):
    """Apply full formatting to the Risk_Owner_Review worksheet."""
    header_row = 1
    data_start = 2

    # Freeze panes: Entity ID + Entity Name + header row
    ws.freeze_panes = f"C{data_start}"

    # Auto-filter
    last_col = get_column_letter(ws.max_column)
    ws.auto_filter.ref = f"A{header_row}:{last_col}{ws.max_row}"

    # Column widths
    ro_col_widths = {
        "Entity Overview": 40, "Decision Basis": 50,
        "Source Rationale Excerpt": 50, "Applicable Siblings": 45,
        "Sibling Alert": 30, "Business Line Comparison": 40,
        "RCO Comment": 40,
    }
    for cell in ws[header_row]:
        if cell.value in ro_col_widths:
            ws.column_dimensions[cell.column_letter].width = ro_col_widths[cell.value]
        else:
            # Auto-fit capped at 25
            ws.column_dimensions[cell.column_letter].width = min(
                max(len(str(cell.value or "")) + 4, 12), 25
            )

    # Text wrap on long-text columns
    wrap_align = Alignment(wrap_text=True, vertical="top")
    wrap_cols = ("Entity Overview", "Decision Basis", "Source Rationale Excerpt",
                 "Applicable Siblings", "Sibling Alert", "Business Line Comparison",
                 "Impact of Issues", "RCO Comment")
    for col_name in wrap_cols:
        col_idx = _find_header_column(ws, col_name)
        if col_idx:
            for row_idx in range(data_start, ws.max_row + 1):
                ws.cell(row=row_idx, column=col_idx).alignment = wrap_align

    # Row height
    for row_idx in range(data_start, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 45

    # Status cell coloring (same fills as Audit_Review)
    status_col_ro = _find_header_column(ws, "Proposed Status")
    if status_col_ro:
        for row_idx in range(data_start, ws.max_row + 1):
            status_val = ws.cell(row=row_idx, column=status_col_ro).value
            fill = status_fills.get(status_val)
            if fill:
                ws.cell(row=row_idx, column=status_col_ro).fill = fill

    # Sibling Alert cell coloring -- orange when populated
    orange_fill = PatternFill("solid", fgColor="FCE4D6")
    alert_col = _find_header_column(ws, "Sibling Alert")
    if alert_col:
        for row_idx in range(data_start, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=alert_col).value
            if val and str(val).strip():
                ws.cell(row=row_idx, column=alert_col).fill = orange_fill

    # Contradicted N/A row coloring -- light red for priority 100 rows
    priority_col = _find_header_column(ws, "Review Priority")
    red_fill = PatternFill("solid", fgColor="FFC7CE")
    if priority_col:
        for row_idx in range(data_start, ws.max_row + 1):
            if ws.cell(row=row_idx, column=priority_col).value == 100:
                for col_idx in range(1, ws.max_column + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = red_fill

    # RCO action column headers -- green fill
    rco_header_fill = PatternFill("solid", fgColor="E2EFDA")
    for cell in ws[header_row]:
        if cell.value in ("RCO Agrees", "RCO Recommended Status",
                          "RCO Recommended Rating", "RCO Comment"):
            cell.fill = rco_header_fill

    # Column grouping: rating detail columns (Likelihood through Impact of Issues)
    lh_col = _find_header_column(ws, "Likelihood")
    ioi_col = _find_header_column(ws, "Impact of Issues")
    if lh_col and ioi_col:
        for col_idx in range(lh_col, ioi_col + 1):
            cl = get_column_letter(col_idx)
            ws.column_dimensions[cl].outlineLevel = 1
            ws.column_dimensions[cl].hidden = True

    # Decision Basis and Source Rationale Excerpt are visible by default
    # (grouping removed — these columns are essential for RCO review)


def _format_risk_owner_summary_sheet(ws):
    """Apply formatting to the Risk_Owner_Summary worksheet."""
    header_row = 1
    data_start = 2

    ws.freeze_panes = f"C{data_start}"

    # Column widths
    for cell in ws[header_row]:
        if cell.value == "L2":
            ws.column_dimensions[cell.column_letter].width = 35
        elif cell.value == "Applicable %":
            ws.column_dimensions[cell.column_letter].width = 12
        else:
            ws.column_dimensions[cell.column_letter].width = 15

    # Format Applicable % as percentage
    pct_col = _find_header_column(ws, "Applicable %")
    if pct_col:
        for row_idx in range(data_start, ws.max_row + 1):
            ws.cell(row=row_idx, column=pct_col).number_format = '0.0%'

    # Bold Contradicted N/A and Sibling Alerts where > 0
    bold_font_summary = Font(bold=True)
    for col_name in ("Contradicted N/A", "Sibling Alerts"):
        col_idx = _find_header_column(ws, col_name)
        if col_idx:
            for row_idx in range(data_start, ws.max_row + 1):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val and val > 0:
                    ws.cell(row=row_idx, column=col_idx).font = bold_font_summary


def _build_dashboard_sheet(wb, ar_ws):
    """Create and populate the Dashboard sheet with formulas referencing Audit_Review."""
    dash_ws = wb.create_sheet("Dashboard", 0)

    # Find column letters in Audit_Review for formulas
    rs_col = ps_col = cs_col = as_col = db_col = ""
    for cell in ar_ws[1]:
        if cell.value == "Reviewer Status": rs_col = cell.column_letter
        if cell.value == "Proposed Status": ps_col = cell.column_letter
        if cell.value == "Control Signals": cs_col = cell.column_letter
        if cell.value == "Additional Signals": as_col = cell.column_letter
        if cell.value == "Decision Basis": db_col = cell.column_letter
    ar_max = ar_ws.max_row
    section_font = Font(bold=True, size=11, color="2F5496")
    label_font = Font(size=10)
    bold_font = Font(bold=True, size=10)

    r = 1
    dash_ws.cell(row=r, column=1, value="Risk Taxonomy Review \u2014 Dashboard").font = Font(bold=True, size=14, color="2F5496")
    r = 2
    dash_ws.cell(row=r, column=1, value=f"Generated: {datetime.now().strftime('%B %d, %Y %I:%M %p').replace(' 0', ' ')}")

    # --- Section A: Tool Proposals ---
    r = 4
    dash_ws.cell(row=r, column=1, value="TOOL PROPOSALS").font = section_font
    dash_ws.cell(row=r, column=2, value="Count").font = bold_font
    dash_ws.cell(row=r, column=3, value="%").font = bold_font

    total_row = r + 2  # row for total count (used in % formulas)
    proposals = [
        (r+1, "Total Audit Entities", f'=SUMPRODUCT(1/COUNTIF(Audit_Review!A2:A{ar_max},Audit_Review!A2:A{ar_max}))', ""),
        (r+2, "Total Entity-L2 Rows", f'=COUNTA(Audit_Review!A2:A{ar_max})', ""),
        (r+3, "Applicable (evidence found)",
         f'=COUNTIF(Audit_Review!{ps_col}2:{ps_col}{ar_max},"{Status.APPLICABLE}")'
         f'-COUNTIFS(Audit_Review!{ps_col}2:{ps_col}{ar_max},"{Status.APPLICABLE}",Audit_Review!{db_col}2:{db_col}{ar_max},"AI review*")'
         if db_col else f'=COUNTIF(Audit_Review!{ps_col}2:{ps_col}{ar_max},"{Status.APPLICABLE}")', True),
        (r+4, "AI-Resolved",
         f'=COUNTIFS(Audit_Review!{db_col}2:{db_col}{ar_max},"AI review*")'
         if db_col else '', True),
        (r+5, "Applicability Undetermined", f'=COUNTIF(Audit_Review!{ps_col}2:{ps_col}{ar_max},"{Status.UNDETERMINED}")', True),
        (r+6, "Assumed N/A \u2014 Verify", f'=COUNTIF(Audit_Review!{ps_col}2:{ps_col}{ar_max},"Assumed N/A*")', True),
        (r+7, "Not Applicable (legacy N/A)",
         f'=COUNTIF(Audit_Review!{ps_col}2:{ps_col}{ar_max},"{Status.NOT_APPLICABLE}")'
         f'-COUNTIFS(Audit_Review!{ps_col}2:{ps_col}{ar_max},"{Status.NOT_APPLICABLE}",Audit_Review!{db_col}2:{db_col}{ar_max},"AI review*")'
         if db_col else f'=COUNTIF(Audit_Review!{ps_col}2:{ps_col}{ar_max},"{Status.NOT_APPLICABLE}")', True),
        (r+8, "No Legacy Source (structural gap)", f'=COUNTIF(Audit_Review!{ps_col}2:{ps_col}{ar_max},"{Status.NOT_ASSESSED}")', True),
        (r+9, "", "", ""),
        (r+10, "Rows Requiring Your Judgment",
         f'=COUNTIF(Audit_Review!{ps_col}2:{ps_col}{ar_max},"{Status.UNDETERMINED}")'
         f'+COUNTIF(Audit_Review!{ps_col}2:{ps_col}{ar_max},"Assumed N/A*")', True),
    ]
    if cs_col:
        proposals.append((r+11, "Rows With Control Signals",
                          f'=COUNTIF(Audit_Review!{cs_col}2:{cs_col}{ar_max},"<>")', True))
    if as_col:
        proposals.append((r+12, "Rows With Additional Signals",
                          f'=COUNTIF(Audit_Review!{as_col}2:{as_col}{ar_max},"<>")', True))

    for row_num, label, formula, show_pct in proposals:
        if label:
            dash_ws.cell(row=row_num, column=1, value=label).font = label_font
        if formula:
            dash_ws.cell(row=row_num, column=2, value=formula)
        if show_pct is True:
            pct_formula = f'=IF(B${total_row}=0,0,B{row_num}/B${total_row})'
            dash_ws.cell(row=row_num, column=3, value=pct_formula)
            dash_ws.cell(row=row_num, column=3).number_format = '0.0%'

    # Dashboard column widths
    dash_ws.column_dimensions["A"].width = 40
    dash_ws.column_dimensions["B"].width = 15
    dash_ws.column_dimensions["C"].width = 10

    return dash_ws
