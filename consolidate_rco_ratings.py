"""
Consolidate RCO Rating Responses
=================================
Reads ChatGPT response.json files from rco_rating_prompts/<l2>/ batch folders,
validates each response, merges with LUminate suggested status, and writes a
single Excel summary sorted by rating severity.

Output: data/output/rco_ratings_<l2_slug>_<timestamp>.xlsx

The "Ratings" sheet is a per-entity adjudication view — one row per AE with
everything the RCO needs in-row:
  Entity ID | Entity Name | Entity Overview (Optro, Archer fallback) |
  Proposed Rating | Rating Rationale | LUminate Status | Core/Aux |
  <L2> Findings (count + "id | title | severity | status" lines)

For Model, the row additionally carries the 1-1 legacy view and the model
makeup from the inventory join:
  ... | Legacy Rating | Legacy Rationale | ... | Models Tagged | Critical |
  High | Medium | Low | Not In Inventory | Guidance Impact (counts) |
  Model Classes | ... | Flags

Model also gets three analysis sheets built from the model inventory join:
  AE Model Profile — per-AE model counts by impact category, model classes,
    and the counts-based Impact implied by the RCO guidance (>=1 Critical -> C,
    >=1 High -> H, Medium >= 30% or >= 2 -> M, else L). Flags proposals below
    that guidance impact (likelihood may legitimately moderate — the flag asks
    the RCO to confirm), High+ proposals with no models on file, and proposals
    2+ levels below legacy with no tagged models (oversight/governance AEs
    whose exposure the tagged-model lens may not capture)
  Shared Models    — models tagged to 2+ AEs, flagged when the tagged AEs'
    proposed ratings diverge by 2+ levels
  Peer Divergence  — AE pairs sharing 50%+ of their model portfolio whose
    proposed ratings are 2+ levels apart
Thresholds are module constants (GUIDANCE_*, PEER_*) — tune as the models team
reviews real output.

Rating column is color-coded (Critical=red, High=orange, Medium=yellow, Low=green).
A summary block at the top shows counts by rating.

Usage:
    python consolidate_rco_ratings.py --l2 Conduct
    python consolidate_rco_ratings.py --l2 "Internal Fraud"
    python consolidate_rco_ratings.py --l2 Conduct --dry-run   # validate only
"""

from __future__ import annotations

import argparse
import json
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from risk_taxonomy_transformer.utils import latest_input
from export_rco_rating_prompts import (
    _load_config,
    _load_model_risk_legacy,
    _load_model_inventory,
    _load_ae_model_tagging,
    _load_optro_overviews,
    _parse_model_ids,
    load_relationships,
    load_findings_by_ae,
    resolve_l2_name,
    l2_output_slug,
    VALID_RATINGS,
    RATING_LEVEL,
    RATINGS_BY_SEVERITY,
)
from consolidate_llm_responses import _try_parse_json_array
from export_html_report import _norm_id_series

_PROJECT_ROOT = Path(__file__).parent
_PROMPTS_ROOT = _PROJECT_ROOT / "data" / "output" / "rco_rating_prompts"
_OUT_DIR = _PROJECT_ROOT / "data" / "output"

# Tunable thresholds for the Model analysis flags.
# GUIDANCE_* mirror the RCO Model Risk guidance's counts-based impact rules
# ("#Medium >= 30% or at least 2"); >=1 Critical -> C and >=1 High -> H are
# fixed by the guidance and implemented in _guidance_impact.
GUIDANCE_MEDIUM_PCT = 0.30     # Medium-impact models >= this share of the AE's models...
GUIDANCE_MEDIUM_COUNT = 2      # ...or at least this many -> guidance impact Medium
PEER_OVERLAP_MIN = 0.5         # Jaccard overlap on model ID sets to count as peers
PEER_RATING_GAP_MIN = 2        # rating levels apart (Low<Medium<High<Critical) to flag

# openpyxl fill colors per rating
_FILLS = {
    "Critical": PatternFill("solid", fgColor="FF4C4C"),  # red
    "High":     PatternFill("solid", fgColor="FF944C"),  # orange
    "Medium":   PatternFill("solid", fgColor="FFD966"),  # yellow
    "Low":      PatternFill("solid", fgColor="92D050"),  # green
}
_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")    # dark blue
_SUMMARY_FILL = PatternFill("solid", fgColor="D6E4F0")   # light blue
_THIN = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _load_workbook_context(l2_name: str) -> dict:
    """Per-entity context from the latest transformer workbook: LUminate
    suggested status for the L2, Archer entity overviews, Core/Aux
    relationships (Side_by_Side), and open findings tagged to the L2.

    Returns {"status": {}, "archer": {}, "relationships": {}, "findings": {}},
    all keyed by normalized entity ID; empty dicts when no workbook exists.
    """
    ctx: dict = {"status": {}, "archer": {}, "relationships": {}, "findings": {}}
    latest = latest_input(
        _PROJECT_ROOT / "data" / "output",
        ["transformed_risk_taxonomy_*.xlsx"],
        log_label="transformer output",
    )
    if latest is None:
        return ctx
    try:
        xls = pd.ExcelFile(latest)
        audit_df = pd.read_excel(xls, sheet_name="Audit_Review")
        audit_df["Entity ID"] = _norm_id_series(audit_df["Entity ID"])
        subset = audit_df[audit_df["New L2"] == l2_name]
        ctx["status"] = dict(zip(subset["Entity ID"],
                                 subset["Suggested Status"].astype(str)))
        if "Entity Overview" in audit_df.columns:
            for _, r in audit_df.drop_duplicates(subset=["Entity ID"],
                                                 keep="first").iterrows():
                ov = str(r.get("Entity Overview", "")).strip()
                if ov and ov.lower() != "nan":
                    ctx["archer"][r["Entity ID"]] = ov
        ctx["relationships"] = load_relationships(xls, l2_name)
        ctx["findings"] = load_findings_by_ae(xls, l2_name)
    except Exception as e:
        print(f"  [warn] Could not load workbook context: {e}")
    return ctx


# ---------------------------------------------------------------------------
# Model analysis (Model / Model Risk only)
# ---------------------------------------------------------------------------

def _guidance_impact(counts: dict[str, int], total: int) -> str:
    """Counts-based Impact per the RCO Model Risk guidance.

    Purpose-based rules (Regulatory / Capital / P&L / etc. -> Critical) can only
    raise this, so it is a floor on guidance Impact — not on inherent risk,
    which likelihood (Table 3.1) may moderate downward.
    """
    if counts["Critical"] >= 1:
        return "Critical"
    if counts["High"] >= 1:
        return "High"
    if counts["Medium"] >= GUIDANCE_MEDIUM_COUNT or (
            total and counts["Medium"] / total >= GUIDANCE_MEDIUM_PCT):
        return "Medium"
    return "Low"


def _build_model_analysis(rows: list[dict], legacy_data: dict, inventory: dict,
                          tagging: dict | None = None) -> dict:
    """Build per-AE model profiles, shared-model view, and peer-divergence pairs.

    Entity->model mapping comes from the ae_model_tagging file when present
    (authoritative), else the legacy Models field — same precedence as the
    prompt export.

    Returns {"profiles": [...], "shared_models": [...], "peer_pairs": [...]}.
    """
    rating_by_ae = {r["entity_id"]: r["proposed_rating"] for r in rows}

    profiles: list[dict] = []
    ae_models: dict[str, set[str]] = {}
    for row in rows:
        eid = row["entity_id"]
        if tagging is not None:
            models_text = tagging.get(eid, "")
        else:
            models_text = legacy_data.get(eid, {}).get("models_text", "")
        mids = _parse_model_ids(models_text)
        # Only inventory-matched IDs count as models — same as the dashboard,
        # which discards stray tokens (years, versions) that match no row.
        matched = [m for m in mids if m in inventory]
        ae_models[eid] = set(matched)

        counts = {r: 0 for r in VALID_RATINGS}
        for mid in matched:
            impact = inventory[mid].get("impact", "")
            if impact in counts:
                counts[impact] += 1
        not_in_inventory = len(mids) - len(matched)

        guidance = _guidance_impact(counts, len(matched)) if matched else "—"
        classes = sorted({inventory[mid].get("class_", "") for mid in matched} - {""})

        flags: list[str] = []
        proposed = row["proposed_rating"]
        if matched and RATING_LEVEL.get(proposed, 0) < RATING_LEVEL.get(guidance, 0):
            flags.append(
                f"Proposed {proposed} is below counts-based guidance impact {guidance} "
                f"({counts['Critical']} Critical / {counts['High']} High / "
                f"{counts['Medium']} Medium model(s)) — confirm likelihood "
                f"supports the lower rating"
            )
        if not matched and RATING_LEVEL.get(proposed, 0) >= RATING_LEVEL["High"]:
            flags.append("Proposed rating is High or above but no models on file")
        legacy_rating = str(row.get("legacy_rating", "")).strip()
        if not matched and legacy_rating in RATING_LEVEL:
            gap = RATING_LEVEL[legacy_rating] - RATING_LEVEL.get(proposed, 0)
            if gap >= 2:
                flags.append(
                    f"Proposed {proposed} is {gap} levels below legacy {legacy_rating} "
                    f"with no tagged models — the tagged-model lens may not capture "
                    f"this entity's exposure (oversight/governance AEs, untagged "
                    f"model classes)"
                )

        profiles.append({
            "entity_id": eid,
            "entity_name": row["entity_name"],
            "proposed": proposed,
            "legacy": row.get("legacy_rating", "—"),
            "total": len(matched),
            "counts": counts,
            "not_in_inventory": not_in_inventory,
            "guidance_impact": guidance,
            "classes": ", ".join(classes),
            "flags": "; ".join(flags),
        })

    # Shared models: one row per model tagged to 2+ AEs.
    model_to_aes: dict[str, list[str]] = {}
    for eid, mids in ae_models.items():
        for mid in mids:
            model_to_aes.setdefault(mid, []).append(eid)

    shared_models: list[dict] = []
    for mid in sorted(model_to_aes):
        aes = sorted(model_to_aes[mid])
        if len(aes) < 2:
            continue
        inv = inventory.get(mid, {})
        levels = [RATING_LEVEL[rating_by_ae[e]] for e in aes
                  if rating_by_ae.get(e) in RATING_LEVEL]
        divergent = len(levels) >= 2 and (max(levels) - min(levels) >= PEER_RATING_GAP_MIN)
        shared_models.append({
            "model_id": mid,
            "name": inv.get("name", "(not in inventory)"),
            "class_": inv.get("class_", ""),
            "impact": inv.get("impact", ""),
            "ae_count": len(aes),
            "aes": ", ".join(f"{e} ({rating_by_ae.get(e, '?')})" for e in aes),
            "flag": ("Tagged AEs diverge by 2+ rating levels" if divergent else ""),
        })

    # Peer divergence: AE pairs with heavy model overlap but distant ratings.
    peer_pairs: list[dict] = []
    eids = sorted(e for e, m in ae_models.items() if m)
    for i, a in enumerate(eids):
        for b in eids[i + 1:]:
            inter = ae_models[a] & ae_models[b]
            if not inter:
                continue
            overlap = len(inter) / len(ae_models[a] | ae_models[b])
            if overlap < PEER_OVERLAP_MIN:
                continue
            ra, rb = rating_by_ae.get(a), rating_by_ae.get(b)
            if ra not in RATING_LEVEL or rb not in RATING_LEVEL:
                continue
            gap = abs(RATING_LEVEL[ra] - RATING_LEVEL[rb])
            if gap >= PEER_RATING_GAP_MIN:
                peer_pairs.append({
                    "ae_a": a, "rating_a": ra,
                    "ae_b": b, "rating_b": rb,
                    "shared_count": len(inter),
                    "overlap_pct": round(overlap * 100),
                    "why": (
                        f"{len(inter)} shared model(s), {round(overlap * 100)}% portfolio "
                        f"overlap, but ratings are {gap} levels apart ({ra} vs {rb})"
                    ),
                })

    return {"profiles": profiles, "shared_models": shared_models, "peer_pairs": peer_pairs}


# ---------------------------------------------------------------------------
# Core consolidation
# ---------------------------------------------------------------------------

def consolidate(l2_name: str, dry_run: bool = False) -> int:
    """Consolidate responses for one L2. Returns exit code (0 = ok, 1 = errors)."""
    l2_name = resolve_l2_name(l2_name)
    slug = l2_output_slug(l2_name)
    prompt_dir = _PROMPTS_ROOT / slug

    if not prompt_dir.exists():
        print(f'  No prompt folder found at {prompt_dir}')
        print(f'  Run: python export_rco_rating_prompts.py --l2 "{l2_name}" first.')
        return 1

    batch_dirs = sorted(d for d in prompt_dir.iterdir()
                        if d.is_dir() and d.name.startswith("batch_"))
    if not batch_dirs:
        print(f"  No batch_NNN/ folders in {prompt_dir}")
        return 1

    print(f'  Found {len(batch_dirs)} batch folder(s)')

    all_rows: list[dict] = []
    errors: list[str] = []
    warnings: list[str] = []

    for batch_dir in batch_dirs:
        rfile = batch_dir / "response.json"
        mfile = batch_dir / "manifest.json"

        # Load manifest for entity list
        expected_entities: list[str] = []
        if mfile.exists():
            try:
                manifest = json.loads(mfile.read_text(encoding="utf-8"))
                expected_entities = manifest.get("entities", [])
            except Exception as e:
                warnings.append(f"{batch_dir.name}: manifest load error — {e}")

        # Load response
        if not rfile.exists():
            errors.append(f"{batch_dir.name}: response.json missing")
            continue

        text = rfile.read_text(encoding="utf-8")
        data, err = _try_parse_json_array(text)
        if err:
            errors.append(f"{batch_dir.name}: {err}")
            continue
        if not data:
            warnings.append(f"{batch_dir.name}: response.json is empty — no output pasted yet")
            continue

        # Validate rows
        batch_entity_ids: list[str] = []
        for idx, item in enumerate(data):
            if not isinstance(item, dict):
                errors.append(f"{batch_dir.name} item {idx}: expected object")
                continue
            missing = [k for k in ("entity_id", "entity_name", "proposed_rating", "rating_rationale")
                       if k not in item]
            if missing:
                errors.append(f"{batch_dir.name} item {idx}: missing fields {missing}")
                continue
            rating = str(item["proposed_rating"]).strip()
            if rating not in VALID_RATINGS:
                errors.append(
                    f"{batch_dir.name} item {idx} ({item.get('entity_id', '?')}): "
                    f"invalid rating '{rating}' — must be one of {VALID_RATINGS}"
                )
                continue
            row = {
                "entity_id":       str(item["entity_id"]).strip(),
                "entity_name":     str(item["entity_name"]).strip(),
                "proposed_rating": rating,
                "rating_rationale": str(item["rating_rationale"]).strip(),
            }
            all_rows.append(row)
            batch_entity_ids.append(row["entity_id"])

        # Coverage check
        if expected_entities:
            responded = set(batch_entity_ids)
            missing_ae = set(expected_entities) - responded
            if missing_ae:
                warnings.append(
                    f"{batch_dir.name}: response missing {len(missing_ae)} "
                    f"expected entity ID(s): {sorted(missing_ae)}"
                )

        print(f"    {batch_dir.name}: {len(batch_entity_ids)} valid row(s)")

    # Summary
    print()
    if errors:
        for e in errors:
            print(f"  [error] {e}")
    if warnings:
        for w in warnings:
            print(f"  [warn]  {w}")

    if not all_rows:
        print("  No valid rows to consolidate.")
        return 1 if errors else 0

    # Deduplicate — last response wins if an entity appears in multiple batches
    seen: dict[str, dict] = {}
    for row in all_rows:
        seen[row["entity_id"]] = row
    rows = list(seen.values())

    # Sort by rating severity, most severe first
    rows.sort(key=lambda r: -RATING_LEVEL.get(r["proposed_rating"], -99))

    # Attach per-entity context: LUminate status, overview (Optro primary,
    # Archer fallback — same precedence as the prompt), Core/Aux, findings.
    ctx = _load_workbook_context(l2_name)
    optro_overviews = _load_optro_overviews()
    for row in rows:
        eid = row["entity_id"]
        row["luminate_status"] = ctx["status"].get(eid, "—")
        row["overview"] = optro_overviews.get(eid) or ctx["archer"].get(eid, "—")
        row["core_aux"] = ctx["relationships"].get(eid, "Not specified")
        row["findings"] = ctx["findings"].get(eid, [])

    # Model Risk: attach legacy rating + rationale (1-1 mapping) and build
    # the model composition analysis for the extra sheets.
    is_model_risk = l2_name == "Model"
    model_analysis = None
    if is_model_risk:
        cfg = _load_config()
        legacy_data = _load_model_risk_legacy(cfg)
        for row in rows:
            legacy = legacy_data.get(row["entity_id"], {})
            rating = legacy.get("rating", "")
            rationale = legacy.get("rationale", "")
            row["legacy_rating"] = rating if rating.lower() not in ("nan", "none", "") else "—"
            row["legacy_rationale"] = rationale if rationale.lower() not in ("nan", "none", "") else "—"
        inventory = _load_model_inventory(cfg)
        tagging = _load_ae_model_tagging(cfg)
        model_analysis = _build_model_analysis(rows, legacy_data, inventory, tagging)

    print(f"  Total valid entities: {len(rows)}")
    counts = {r: sum(1 for row in rows if row["proposed_rating"] == r) for r in VALID_RATINGS}
    for rating in RATINGS_BY_SEVERITY:
        if counts[rating]:
            print(f"    {rating}: {counts[rating]}")

    if model_analysis:
        flagged = [p for p in model_analysis["profiles"] if p["flags"]]
        divergent = [m for m in model_analysis["shared_models"] if m["flag"]]
        print(f"  Model analysis: {len(flagged)} flagged AE profile(s), "
              f"{len(divergent)} shared model(s) with divergent AE ratings, "
              f"{len(model_analysis['peer_pairs'])} divergent peer pair(s)")

    if dry_run:
        print()
        print("  --dry-run: no file written.")
        return 1 if errors else 0

    # Write Excel
    ts = datetime.now().strftime("%m%d%Y%I%M%p")
    out_path = _OUT_DIR / f"rco_ratings_{slug}_{ts}.xlsx"
    _write_excel(out_path, l2_name, rows, counts,
                 include_legacy=is_model_risk, model_analysis=model_analysis)
    print()
    print(f"  Written: {out_path.name}")
    return 1 if errors else 0


def _write_excel(path: Path, l2_name: str, rows: list[dict], counts: dict[str, int],
                 include_legacy: bool = False, model_analysis: dict | None = None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Ratings"

    prof_by_eid = ({p["entity_id"]: p for p in model_analysis["profiles"]}
                   if model_analysis else {})

    headers = ["Entity ID", "Entity Name", "Entity Overview",
               "Proposed Rating", "Rating Rationale"]
    col_widths = [12, 26, 60, 14, 60]
    if include_legacy:
        headers += ["Legacy Rating", "Legacy Rationale"]
        col_widths += [12, 60]
    headers += ["LUminate Status", "Core/Aux"]
    col_widths += [16, 12]
    if model_analysis is not None:
        headers += ["Models Tagged", "Critical", "High", "Medium", "Low",
                    "Not In Inventory", "Guidance Impact (counts)", "Model Classes"]
        col_widths += [10, 8, 8, 8, 8, 12, 14, 22]
    headers += [f"{l2_name} Findings"]
    col_widths += [50]
    if model_analysis is not None:
        headers += ["Flags"]
        col_widths += [50]

    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    last_col = get_column_letter(len(col_widths))

    row_num = 1

    # Title row
    ws.merge_cells(f"A{row_num}:{last_col}{row_num}")
    title_cell = ws.cell(row_num, 1,
                         value=f"{l2_name} — RCO Proposed Ratings  |  "
                               f"Generated {datetime.now().strftime('%Y-%m-%d')}")
    title_cell.font = Font(bold=True, color="FFFFFF", size=12)
    title_cell.fill = _HEADER_FILL
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row_num].height = 22
    row_num += 1

    # Summary row
    ws.merge_cells(f"A{row_num}:{last_col}{row_num}")
    summary_parts = [f"{r}: {counts[r]}" for r in RATINGS_BY_SEVERITY if counts[r]]
    summary_cell = ws.cell(row_num, 1,
                           value="  |  ".join(summary_parts) +
                                 f"  |  Total: {sum(counts.values())}")
    summary_cell.font = Font(bold=True, size=10)
    summary_cell.fill = _SUMMARY_FILL
    summary_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row_num].height = 18
    row_num += 1

    # Blank separator
    row_num += 1

    # Header row
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row_num, col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _THIN
    ws.row_dimensions[row_num].height = 20
    row_num += 1

    # Data rows — cell values resolved by header name so column blocks can
    # come and go per mode without index bookkeeping.
    _rating_cols = ("Proposed Rating", "Legacy Rating", "Guidance Impact (counts)")
    for row in rows:
        prof = prof_by_eid.get(row["entity_id"], {})
        prof_counts = prof.get("counts", {})
        findings = row.get("findings", [])
        cell_map = {
            "Entity ID": row["entity_id"],
            "Entity Name": row["entity_name"],
            "Entity Overview": row.get("overview", "—"),
            "Proposed Rating": row["proposed_rating"],
            "Rating Rationale": row["rating_rationale"],
            "Legacy Rating": row.get("legacy_rating", "—"),
            "Legacy Rationale": row.get("legacy_rationale", "—"),
            "LUminate Status": row["luminate_status"],
            "Core/Aux": row.get("core_aux", "Not specified"),
            "Models Tagged": prof.get("total", 0),
            "Critical": prof_counts.get("Critical", 0),
            "High": prof_counts.get("High", 0),
            "Medium": prof_counts.get("Medium", 0),
            "Low": prof_counts.get("Low", 0),
            "Not In Inventory": prof.get("not_in_inventory", 0),
            "Guidance Impact (counts)": prof.get("guidance_impact", "—"),
            "Model Classes": prof.get("classes", ""),
            f"{l2_name} Findings": ("None on file" if not findings else
                                    f"{len(findings)} open:\n" + "\n".join(findings)),
            "Flags": prof.get("flags", ""),
        }
        for col, header in enumerate(headers, start=1):
            val = cell_map[header]
            cell = ws.cell(row_num, col, value=val)
            cell.border = _THIN
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if header in _rating_cols and val in _FILLS:
                cell.fill = _FILLS[val]
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif header == "Flags" and val:
                cell.font = Font(bold=True, color="C00000")
        ws.row_dimensions[row_num].height = 40
        row_num += 1

    # Freeze panes below header
    ws.freeze_panes = f"A{5}"  # row 4 is header, freeze below it

    if model_analysis:
        _write_analysis_sheets(wb, model_analysis)

    wb.save(path)


def _sheet_header(ws, headers: list[str], widths: list[int]) -> int:
    """Write a formatted header row on row 1; return the next row number."""
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(1, col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _THIN
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"
    return 2


def _write_analysis_sheets(wb: Workbook, analysis: dict):
    # --- AE Model Profile ---
    ws = wb.create_sheet("AE Model Profile")
    row_num = _sheet_header(
        ws,
        ["Entity ID", "Entity Name", "Proposed Rating", "Legacy Rating",
         "Total Models", "Critical", "High", "Medium", "Low",
         "Not In Inventory", "Guidance Impact (counts)", "Model Classes", "Flags"],
        [12, 30, 16, 14, 12, 9, 9, 9, 9, 14, 16, 24, 70],
    )
    for p in analysis["profiles"]:
        values = [
            p["entity_id"], p["entity_name"], p["proposed"], p["legacy"],
            p["total"],
            p["counts"]["Critical"], p["counts"]["High"],
            p["counts"]["Medium"], p["counts"]["Low"],
            p["not_in_inventory"], p["guidance_impact"], p["classes"], p["flags"],
        ]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row_num, col, value=val)
            cell.border = _THIN
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if col in (3, 11) and val in _FILLS:
                cell.fill = _FILLS[val]
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col == 13 and val:
                cell.font = Font(bold=True, color="C00000")
        row_num += 1

    # --- Shared Models ---
    ws = wb.create_sheet("Shared Models")
    row_num = _sheet_header(
        ws,
        ["Model ID", "Model Name", "Model Class", "Impact",
         "AE Count", "Tagged AEs (proposed rating)", "Flag"],
        [10, 34, 18, 12, 10, 60, 40],
    )
    if not analysis["shared_models"]:
        ws.cell(row_num, 1, value="No models are tagged to more than one entity.")
    for m in analysis["shared_models"]:
        values = [m["model_id"], m["name"], m["class_"], m["impact"],
                  m["ae_count"], m["aes"], m["flag"]]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row_num, col, value=val)
            cell.border = _THIN
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if col == 4 and val in _FILLS:
                cell.fill = _FILLS[val]
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col == 7 and val:
                cell.font = Font(bold=True, color="C00000")
        row_num += 1

    # --- Peer Divergence ---
    ws = wb.create_sheet("Peer Divergence")
    row_num = _sheet_header(
        ws,
        ["Entity A", "Rating A", "Entity B", "Rating B",
         "Shared Models", "Overlap %", "Why Flagged"],
        [12, 12, 12, 12, 14, 10, 80],
    )
    if not analysis["peer_pairs"]:
        ws.cell(row_num, 1,
                value=f"No peer pairs above {int(PEER_OVERLAP_MIN * 100)}% model overlap "
                      f"with ratings {PEER_RATING_GAP_MIN}+ levels apart.")
    for pair in analysis["peer_pairs"]:
        values = [pair["ae_a"], pair["rating_a"], pair["ae_b"], pair["rating_b"],
                  pair["shared_count"], pair["overlap_pct"], pair["why"]]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row_num, col, value=val)
            cell.border = _THIN
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if col in (2, 4) and val in _FILLS:
                cell.fill = _FILLS[val]
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
        row_num += 1


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    import logging
    logging.basicConfig(level=logging.INFO, format="%(message)s")
    parser = argparse.ArgumentParser(
        description="Consolidate RCO rating responses into a single Excel summary"
    )
    parser.add_argument(
        "--l2",
        required=True,
        help='L2 name, e.g. "Conduct" or "Internal Fraud"',
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        dest="dry_run",
        help="Validate responses and print summary without writing the Excel file",
    )
    ns = parser.parse_args()

    print(f'Consolidating RCO ratings for "{ns.l2}"...')
    rc = consolidate(ns.l2, dry_run=ns.dry_run)
    sys.exit(rc)
