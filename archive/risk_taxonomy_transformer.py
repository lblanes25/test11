"""
Risk Taxonomy Transformer
=========================
Transforms legacy 14-pillar risk taxonomy into new 6 L1 / 23 L2 taxonomy.
Handles: LLM override file, sub-risk description lookup, deterministic mapping,
1:many keyword resolution, overlay/amplifier risks, and rating decomposition.

Resolution order for split mappings:
  Override — LLM-classified overrides from Review Queue (highest priority)
  Tier 0  — Sub-risk descriptions (deterministic)
  Tier 1  — Keyword matching on pillar rationale text
  Tier 2  — Default to first L2, flagged for review

Workflow:
  1. Run script without overrides -> produces Review Queue
  2. Batch Review Queue through LLM prompt -> produces override file
  3. Re-run script with override file -> overrides replace low-confidence mappings

SETUP:
1. Fill in CROSSWALK_CONFIG with your actual mappings (Section 1)
2. Update KEYWORD_MAP with rationale keywords per L2 (Section 1)
3. Set file paths in main() (Section 6), including SUB_RISK_PATH
4. Run: python risk_taxonomy_transformer.py
5. (Optional) Batch Review Queue through LLM, save as overrides, re-run
"""

import pandas as pd
import re
import logging
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler("transform_log.txt", mode="w"),
        logging.StreamHandler()
    ]
)
log = logging.getLogger(__name__)

# =============================================================================
# SECTION 1: CONFIGURATION — FILL THESE IN
# =============================================================================

# -- Legacy rating maps --
RISK_RATING_MAP = {
    "low": 1, "medium": 2, "high": 3, "critical": 4,
    "l": 1, "m": 2, "h": 3, "c": 4,
    "1": 1, "2": 2, "3": 3, "4": 4,
}

CONTROL_RATING_MAP = {
    "well controlled": 1,
    "moderately controlled": 2,
    "new/not tested yet": 3,
    "not tested": 3,
    "new": 3,
    "insufficiently controlled": 4,
    "not applicable": None,
    "n/a": None,
}

# -- New taxonomy definition --
# 6 L1s / 23 L2s based on actual AmEx risk taxonomy.
NEW_TAXONOMY = {
    "Strategic": [
        "Earnings",
        "Capital",
    ],
    "Liquidity": [
        "Liquidity",
    ],
    "Reputational": [
        "Reputation",
    ],
    "Market": [
        "Interest Rate",
        "FX and Price",
    ],
    "Credit": [
        "Consumer and Small Business",
        "Commercial",
    ],
    "Operational and Compliance": [
        "Data",
        "Fraud (External and Internal)",
        "Information and Cyber Security",
        "Technology",
        "Processing, Execution and Change",
        "Business Disruption",
        "Human Capital",
        "Financial Reporting",
        "Third Party",
        "Model",
        "Conduct",
        "Prudential & bank administration compliance",
        "Customer / client protection and product compliance",
        "Financial crimes",
        "Privacy",
    ],
}

# Flatten for validation
ALL_L2_RISKS = []
L2_TO_L1 = {}
for l1, l2_list in NEW_TAXONOMY.items():
    for l2 in l2_list:
        ALL_L2_RISKS.append(l2)
        L2_TO_L1[l2] = l1

# -- Crosswalk config --
# Each legacy pillar maps to one or more new L2s.
#
# mapping_type:
#   "direct"    — 1:1 to a single L2
#   "multi"     — 1:many, populates ALL targets (primary, secondary, conditional)
#   "overlay"   — amplifier risk, flagged on target L2s
#
# For "multi" targets:
#   "primary"     — always populated with legacy rating, high confidence
#   "secondary"   — always populated, flagged for team review
#   "conditional" — only populated if rationale/sub-risk keywords match
#
# When multiple old L1s map to the same L2, the higher (more conservative)
# rating is kept and both sources are logged for traceability.

CROSSWALK_CONFIG = {
    "Credit": {
        "mapping_type": "multi",
        "targets": [
            {"l2": "Consumer and Small Business", "relationship": "primary"},
            {"l2": "Commercial", "relationship": "primary"},
        ],
        "notes": "Both populate; teams mark the non-applicable one N/A",
    },
    "Market": {
        "mapping_type": "multi",
        "targets": [
            {"l2": "Interest Rate", "relationship": "primary"},
            {"l2": "FX and Price", "relationship": "primary"},
        ],
        "notes": "Both populate; similarly assessed per VP",
    },
    "Strategic & Business": {
        "mapping_type": "multi",
        "targets": [
            {"l2": "Earnings", "relationship": "primary"},
            {"l2": "Capital", "relationship": "secondary"},
        ],
        "notes": "Capital has low applicability for most AEs",
    },
    "Funding and Liquidity": {
        "mapping_type": "direct",
        "target_l2": "Liquidity",
        "notes": "Direct 1:1 mapping",
    },
    "Reputational": {
        "mapping_type": "direct",
        "target_l2": "Reputation",
        "notes": "Direct 1:1 mapping",
    },
    "Model": {
        "mapping_type": "direct",
        "target_l2": "Model",
        "notes": "Direct 1:1 mapping",
    },
    "Third Party": {
        "mapping_type": "direct",
        "target_l2": "Third Party",
        "notes": "Direct 1:1 mapping",
    },
    "Financial Reporting": {
        "mapping_type": "direct",
        "target_l2": "Financial Reporting",
        "notes": "Direct 1:1 mapping",
    },
    "External Fraud": {
        "mapping_type": "direct",
        "target_l2": "Fraud (External and Internal)",
        "notes": "Direct 1:1 mapping",
    },
    "Information Technology": {
        "mapping_type": "multi",
        "targets": [
            {"l2": "Technology", "relationship": "primary"},
            {"l2": "Data", "relationship": "conditional",
             "conditions": ["data governance", "data quality", "data management",
                            "data lineage", "data lifecycle", "data usability"]},
        ],
        "notes": "Data only if rationale mentions data governance topics",
    },
    "Information Security": {
        "mapping_type": "multi",
        "targets": [
            {"l2": "Information and Cyber Security", "relationship": "primary"},
            {"l2": "Data", "relationship": "conditional",
             "conditions": ["data protection", "data loss", "data breach",
                            "data compliance", "PII", "personal data",
                            "data privacy"]},
        ],
        "notes": "Data only if rationale mentions data protection/privacy",
    },
    "Operational": {
        "mapping_type": "multi",
        "targets": [
            {"l2": "Processing, Execution and Change", "relationship": "primary"},
            {"l2": "Business Disruption", "relationship": "primary"},
            {"l2": "Human Capital", "relationship": "primary"},
            {"l2": "Conduct", "relationship": "secondary"},
            {"l2": "Privacy", "relationship": "secondary"},
            {"l2": "Data", "relationship": "conditional",
             "conditions": ["data risk", "data volume", "data depend",
                            "data governance", "data quality", "data process"]},
        ],
        "notes": "Old operational covered operations/process, legal, privacy, people/conduct, data. Largest split.",
    },
    "Compliance": {
        "mapping_type": "multi",
        "targets": [
            {"l2": "Prudential & bank administration compliance", "relationship": "primary"},
            {"l2": "Customer / client protection and product compliance", "relationship": "primary"},
            {"l2": "Financial crimes", "relationship": "primary"},
            {"l2": "Conduct", "relationship": "secondary"},
        ],
        "notes": "Old compliance covered enterprise, consumer, fair lending, new products, global financial crimes",
    },
    "Country": {
        "mapping_type": "overlay",
        "target_l2s": [
            "Prudential & bank administration compliance",
            "Financial crimes",
            "Consumer and Small Business",
            "Commercial",
        ],
        "notes": "Amplifier — flags relevant L2s, does not create own row",
    },
}

# -- Keyword map for conditional and split resolution --
# Used for:
#   1. "conditional" targets in multi mappings (checks rationale + sub-risks)
#   2. Legacy "split" mappings (if any remain)
#   3. Sub-risk description scoring (Tier 0)
#
# Keys must match L2 names exactly. Values are lists of keywords/phrases
# found in legacy rationale text. Case-insensitive matching.
KEYWORD_MAP = {
    # Conditional target keywords — used to decide if a conditional L2 applies
    "Data": [
        "data governance", "data quality", "data management", "data lineage",
        "data lifecycle", "data usability", "data risk", "data volume",
        "data depend", "data process", "data protection", "data loss",
        "data breach", "data compliance", "PII", "personal data",
        "data privacy", "data classification", "record retention",
    ],
    # General L2 keywords — used for sub-risk description scoring
    "Technology": [
        "technology", "system", "IT ", "infrastructure", "processing",
        "stability", "capacity", "performance", "information system",
        "platform", "application", "software", "hardware", "network",
    ],
    "Information and Cyber Security": [
        "cyber", "breach", "phishing", "intrusion", "vulnerab",
        "ransomware", "infosec", "information security", "penetration",
        "firewall", "endpoint", "malware", "DDoS", "unauthorized access",
        "confidentiality", "integrity", "availability",
    ],
    "Processing, Execution and Change": [
        "process", "execution", "change management", "control failure",
        "error", "manual", "reconciliation", "settlement", "transaction",
        "operational loss", "procedure", "safeguard",
    ],
    "Business Disruption": [
        "business continuity", "BCP", "disaster recovery", "DR ",
        "resilience", "outage", "disruption", "pandemic",
        "crisis management", "incident", "hazard",
    ],
    "Human Capital": [
        "talent", "retention", "hiring", "workforce", "culture",
        "training", "succession", "employee", "human resource",
        "HR ", "headcount", "attrition", "compensation", "benefits",
        "workplace environment", "organizational structure",
    ],
    "Conduct": [
        "conduct", "misconduct", "ethics", "code of conduct",
        "blue box values", "integrity", "trust", "colleague",
        "intentional", "unintentional", "responsibilities",
    ],
    "Privacy": [
        "privacy", "personal data", "GDPR", "CCPA", "PII",
        "data subject", "consent", "opt-out", "data processing",
        "privacy law", "privacy regulation",
    ],
    "Prudential & bank administration compliance": [
        "prudential", "bank admin", "governance", "oversight",
        "regulatory commitment", "enterprise compliance",
        "compliance program", "regulatory", "examination",
    ],
    "Customer / client protection and product compliance": [
        "consumer", "fair lending", "UDAAP", "CRA", "disclosure",
        "complaint", "customer harm", "fee", "servicing",
        "product compliance", "new product", "client protection",
    ],
    "Financial crimes": [
        "financial crime", "money laundering", "AML", "BSA",
        "sanctions", "OFAC", "suspicious activity", "terrorism",
        "corruption", "bribery", "KYC", "know your customer",
    ],
    "Fraud (External and Internal)": [
        "fraud", "identity theft", "account takeover", "counterfeit",
        "scheme", "defraud", "false pretenses", "embezzlement",
        "internal fraud", "external fraud",
    ],
    "Financial Reporting": [
        "financial report", "regulatory report", "10-K", "10-Q",
        "SEC filing", "GAAP", "accounting", "material misstatement",
        "restatement", "financial statement",
    ],
    "Third Party": [
        "vendor", "third.party", "outsourc", "supplier",
        "third party", "partner", "service provider",
        "subcontract", "offshore",
    ],
    "Model": [
        "model", "validation", "back-test", "backtest", "challenger",
        "model risk", "algorithm", "methodology", "MRM",
        "model governance", "model performance",
    ],
    "Earnings": [
        "earnings", "revenue", "profitability", "margin",
        "income", "NII", "fee income", "expense", "pricing",
        "customer base", "product diversification",
    ],
    "Capital": [
        "capital", "capital adequacy", "capital raising",
        "capital allocation", "capital distribution",
        "regulatory capital", "CCAR", "stress test",
    ],
    "Interest Rate": [
        "interest rate", "repricing", "yield curve", "basis risk",
        "options risk", "rate sensitivity", "NII",
    ],
    "FX and Price": [
        "foreign exchange", "FX ", "currency", "price risk",
        "market-making", "dealing", "position taking",
        "equity market",
    ],
    "Consumer and Small Business": [
        "consumer", "small business", "cardholder", "cardmember",
        "retail", "personal", "individual",
    ],
    "Commercial": [
        "commercial", "corporate", "counterpart", "institutional",
        "wholesale", "large corporate",
    ],
    "Reputation": [
        "reputation", "brand", "media coverage", "stakeholder",
        "public perception", "trust",
    ],
    "Liquidity": [
        "liquidity", "funding", "cash flow", "obligation",
        "deposit", "borrowing",
    ],
}

# Pre-lowercase keyword lists so inner loops don't repeat .lower() per keyword
for _l2_key in KEYWORD_MAP:
    KEYWORD_MAP[_l2_key] = [kw.lower() for kw in KEYWORD_MAP[_l2_key]]


# =============================================================================
# SECTION 2: INGESTION
# =============================================================================

def ingest_legacy_data(filepath: str) -> pd.DataFrame:
    """Read the legacy entity-level risk data from Excel.

    Expects a wide-format file: one row per audit entity with columns for
    each legacy pillar's rating, rationale, control assessment, and control
    rationale. Adjust column name patterns below to match your file.
    """
    log.info(f"Reading legacy data from {filepath}")
    df = pd.read_excel(filepath)
    log.info(f"  Loaded {len(df)} audit entities, {len(df.columns)} columns")

    # Normalize column names: strip whitespace, lowercase
    df.columns = [c.strip() for c in df.columns]
    return df


def ingest_crosswalk(filepath: str = None) -> dict:
    """Return the crosswalk config.

    If a filepath is provided, reads from Excel and builds the config dict.
    Otherwise returns the hardcoded CROSSWALK_CONFIG above.
    """
    if filepath:
        log.info(f"Reading crosswalk from {filepath}")
        xw = pd.read_excel(filepath)
        log.warning("  File-based crosswalk parsing not yet implemented; using hardcoded config")
    return CROSSWALK_CONFIG


def ingest_sub_risks(filepath: str, entity_id_col: str, legacy_l1_col: str,
                     risk_desc_col: str, risk_id_col: str = None,
                     rating_col: str = None) -> pd.DataFrame:
    """Read sub-risk descriptions file.

    Expected columns (configure names in main()):
      - entity_id_col:  Audit Entity ID
      - risk_id_col:    Risk ID (optional, for traceability)
      - risk_desc_col:  Risk description text
      - legacy_l1_col:  Legacy L1 pillar(s), tab-separated if multiple
      - rating_col:     Inherent risk rating (optional, not used for scoring)

    Returns DataFrame with one row per sub-risk, with legacy L1s exploded
    so each row maps to a single L1.
    """
    log.info(f"Reading sub-risk descriptions from {filepath}")
    df = pd.read_excel(filepath)
    df.columns = [c.strip() for c in df.columns]
    log.info(f"  Loaded {len(df)} sub-risk rows")

    # Rename to standard internal names
    col_map = {entity_id_col: "entity_id", risk_desc_col: "risk_description",
               legacy_l1_col: "legacy_l1_raw"}
    if risk_id_col:
        col_map[risk_id_col] = "risk_id"
    if rating_col:
        col_map[rating_col] = "sub_risk_rating"
    df = df.rename(columns=col_map)

    # Ensure entity_id is string
    df["entity_id"] = df["entity_id"].astype(str).str.strip()

    # Explode multi-value L1 pillars so each row has one L1
    # Excel cells may use newlines, tabs, semicolons, or pipes as separators
    df["legacy_l1_list"] = df["legacy_l1_raw"].astype(str).str.split(r"\n|\t|;|\|")
    df = df.explode("legacy_l1_list")
    df["legacy_l1"] = df["legacy_l1_list"].str.strip()
    df = df[df["legacy_l1"] != ""]  # drop empty strings from splitting
    df = df.drop(columns=["legacy_l1_list"])

    # Clean up description text
    df["risk_description"] = df["risk_description"].astype(str).str.strip()

    log.info(f"  After L1 explosion: {len(df)} sub-risk-to-L1 rows")
    log.info(f"  Unique entities with sub-risks: {df['entity_id'].nunique()}")

    return df


def build_sub_risk_index(sub_risks_df: pd.DataFrame) -> dict:
    """Build lookup: {entity_id: {legacy_pillar: [list of risk descriptions]}}.

    This enables fast lookup during entity transformation.
    """
    index = {}
    all_l1s_in_subrisk = set()
    for _, row in sub_risks_df.iterrows():
        eid = row["entity_id"]
        l1 = row["legacy_l1"]
        desc = row["risk_description"]
        all_l1s_in_subrisk.add(l1)
        if eid not in index:
            index[eid] = {}
        if l1 not in index[eid]:
            index[eid][l1] = []
        index[eid][l1].append(desc)

    # Diagnostic: show which sub-risk L1 values match crosswalk keys
    crosswalk_keys = set(CROSSWALK_CONFIG.keys())
    matched = all_l1s_in_subrisk & crosswalk_keys
    unmatched = all_l1s_in_subrisk - crosswalk_keys
    unused = crosswalk_keys - all_l1s_in_subrisk
    log.info(f"  Sub-risk L1 values found: {sorted(all_l1s_in_subrisk)}")
    log.info(f"  Matched to crosswalk keys: {sorted(matched)}")
    if unmatched:
        log.warning(f"  Sub-risk L1s NOT in crosswalk (will be ignored): {sorted(unmatched)}")
    if unused:
        log.info(f"  Crosswalk keys with NO sub-risks: {sorted(unused)}")

    return index


def load_overrides(filepath: str) -> dict:
    """Load LLM-classified overrides from Excel/CSV.

    Expected columns:
      - entity_id:        Audit entity identifier
      - source_legacy_pillar: Legacy pillar name (must match CROSSWALK_CONFIG keys)
      - classified_l2:    The LLM-assigned L2 risk name (must match NEW_TAXONOMY)
      - llm_confidence:   Optional — High/Medium/Low from the LLM

    Returns dict: {(entity_id, legacy_pillar): {"l2": str, "confidence": str}}
    """
    log.info(f"Loading LLM overrides from {filepath}")

    if filepath.endswith(".csv"):
        df = pd.read_csv(filepath)
    else:
        df = pd.read_excel(filepath)

    df.columns = [c.strip() for c in df.columns]
    df["entity_id"] = df["entity_id"].astype(str).str.strip()
    df["source_legacy_pillar"] = df["source_legacy_pillar"].astype(str).str.strip()
    df["classified_l2"] = df["classified_l2"].astype(str).str.strip()

    overrides = {}
    validated = 0
    skipped = 0

    for _, row in df.iterrows():
        l2 = row["classified_l2"]
        if l2 not in L2_TO_L1:
            log.warning(f"  Override skipped: '{l2}' not in taxonomy "
                        f"(entity={row['entity_id']}, pillar={row['source_legacy_pillar']})")
            skipped += 1
            continue

        key = (row["entity_id"], row["source_legacy_pillar"])
        conf = str(row.get("llm_confidence", "high")).strip().lower()
        if conf not in ("high", "medium", "low"):
            conf = "high"

        overrides[key] = {"l2": l2, "confidence": conf}
        validated += 1

    log.info(f"  Loaded {validated} valid overrides, skipped {skipped} invalid")
    return overrides


def ingest_findings(filepath: str, col_map: dict) -> pd.DataFrame:
    """Read findings/issues data.

    Expected columns (configure names via col_map):
      entity_id, issue_id, l2_risk, severity, status, issue_title, remediation_date
    """
    log.info(f"Reading findings from {filepath}")
    df = pd.read_excel(filepath)
    df.columns = [c.strip() for c in df.columns]

    rename = {}
    for internal, actual in col_map.items():
        if actual and actual in df.columns:
            rename[actual] = internal
    df = df.rename(columns=rename)

    df["entity_id"] = df["entity_id"].astype(str).str.strip()
    df["l2_risk"] = df["l2_risk"].astype(str).str.strip()

    # Validate L2 names match taxonomy
    valid = df["l2_risk"].isin(ALL_L2_RISKS)
    invalid_l2s = df[~valid]["l2_risk"].unique()
    if len(invalid_l2s) > 0:
        log.warning(f"  Findings L2s NOT in taxonomy (will be ignored): {list(invalid_l2s)}")
    df = df[valid]

    log.info(f"  Loaded {len(df)} valid findings across {df['entity_id'].nunique()} entities")
    log.info(f"  L2s covered by findings: {sorted(df['l2_risk'].unique())}")
    return df


def build_findings_index(findings_df: pd.DataFrame) -> dict:
    """Build lookup: {entity_id: {l2_risk: [list of finding dicts]}}.

    Each finding dict: {issue_id, severity, status, issue_title, remediation_date}
    """
    index = {}
    for _, row in findings_df.iterrows():
        eid = row["entity_id"]
        l2 = row["l2_risk"]
        finding = {
            "issue_id": str(row.get("issue_id", "")),
            "severity": str(row.get("severity", "")),
            "status": str(row.get("status", "")),
            "issue_title": str(row.get("issue_title", "")),
            "remediation_date": str(row.get("remediation_date", "")),
        }
        if eid not in index:
            index[eid] = {}
        if l2 not in index[eid]:
            index[eid][l2] = []
        index[eid][l2].append(finding)

    total_findings = sum(len(fs) for eid in index for fs in index[eid].values())
    log.info(f"  Findings index built: {len(index)} entities, {total_findings} total findings")
    return index


# =============================================================================
# SECTION 3: RATING CONVERSION & RATIONALE PARSING
# =============================================================================

def _make_row(
    entity_id: str, l1: str, l2: str, *,
    likelihood=None, impact_financial=None, impact_reputational=None,
    impact_consumer_harm=None, impact_regulatory=None,
    iag_control_effectiveness=None, aligned_assurance_rating=None,
    management_awareness_rating=None,
    source_legacy_pillar=None, source_risk_rating_raw=None,
    source_rationale="", source_control_raw=None, source_control_rationale="",
    mapping_type="", confidence="", method="",
    dims_parsed_from_rationale=False, sub_risk_evidence="", needs_review=False,
) -> dict:
    """Build a single transformed row dict with consistent keys."""
    return {
        "entity_id": entity_id,
        "new_l1": l1,
        "new_l2": l2,
        "composite_key": f"{l2} {entity_id}",
        "likelihood": likelihood,
        "impact_financial": impact_financial,
        "impact_reputational": impact_reputational,
        "impact_consumer_harm": impact_consumer_harm,
        "impact_regulatory": impact_regulatory,
        "iag_control_effectiveness": iag_control_effectiveness,
        "aligned_assurance_rating": aligned_assurance_rating,
        "management_awareness_rating": management_awareness_rating,
        "source_legacy_pillar": source_legacy_pillar,
        "source_risk_rating_raw": source_risk_rating_raw,
        "source_rationale": source_rationale,
        "source_control_raw": source_control_raw,
        "source_control_rationale": source_control_rationale,
        "mapping_type": mapping_type,
        "confidence": confidence,
        "method": method,
        "dims_parsed_from_rationale": dims_parsed_from_rationale,
        "sub_risk_evidence": sub_risk_evidence,
        "needs_review": needs_review,
    }


def convert_risk_rating(value) -> int | None:
    """Convert legacy risk rating to 1-4 numeric scale."""
    if pd.isna(value):
        return None
    return RISK_RATING_MAP.get(str(value).strip().lower())


def convert_control_rating(value) -> int | None:
    """Convert legacy control assessment to 1-4 numeric scale."""
    if pd.isna(value):
        return None
    return CONTROL_RATING_MAP.get(str(value).strip().lower())


def parse_rationale_for_dimensions(rationale: str) -> dict:
    """Extract explicit likelihood/impact mentions from rationale text.

    Handles many free-text formats:
      "likelihood is high"           "impact: medium"
      "likelihood(high)"             "impact (medium)"
      "the likelihood is medium"     "impact is high because..."
      "L: Low, I: High"             "high likelihood"
      "likelihood - low"             "likelihood = critical"
      "likelihood rating: high"      "impact rating is medium"

    Returns dict with any found dimensions; empty dict if none found.
    """
    if not rationale or pd.isna(rationale):
        return {}

    text = str(rationale).lower()
    found = {}
    rating_words = "low|medium|high|critical"

    for dimension in ["likelihood", "impact"]:
        # Pattern 1: "dimension <separator> rating"
        # Handles: is, :, -, =, (, and optional words like "is rated", "rating:"
        match = re.search(
            rf"{dimension}\s*(?:rating\s*)?(?:is\s*(?:rated\s*)?|:|–|-|=|\()\s*({rating_words})",
            text
        )
        if match:
            found[dimension] = RISK_RATING_MAP.get(match.group(1))
            continue

        # Pattern 2: "the dimension ... is/of rating" (words in between, up to 5)
        match = re.search(
            rf"(?:the\s+)?{dimension}\s+(?:\w+\s+){{0,5}}(?:is|of)\s+({rating_words})",
            text
        )
        if match:
            found[dimension] = RISK_RATING_MAP.get(match.group(1))
            continue

        # Pattern 3: "rating dimension" (e.g., "high likelihood")
        match = re.search(
            rf"({rating_words})\s+{dimension}",
            text
        )
        if match:
            found[dimension] = RISK_RATING_MAP.get(match.group(1))

    # Abbreviation patterns: "L: Low" / "I: High" / "L-Low, I-Medium"
    abbrev_match = re.search(
        rf"\bL\s*[:\-=]\s*({rating_words})", text
    )
    if abbrev_match and "likelihood" not in found:
        found["likelihood"] = RISK_RATING_MAP.get(abbrev_match.group(1))

    abbrev_match = re.search(
        rf"\bI\s*[:\-=]\s*({rating_words})", text
    )
    if abbrev_match and "impact" not in found:
        found["impact"] = RISK_RATING_MAP.get(abbrev_match.group(1))

    # Specific impact types: financial, reputational, regulatory, consumer
    for impact_type in ["financial", "reputational", "regulatory", "consumer"]:
        # "financial impact <sep> rating" or "impact <sep> financial <sep> rating"
        match = re.search(
            rf"{impact_type}\s+impact\s*(?:rating\s*)?(?:is\s*(?:rated\s*)?|:|–|-|=|\()?\s*({rating_words})",
            text
        )
        if match:
            found[f"impact_{impact_type}"] = RISK_RATING_MAP.get(match.group(1))
            continue

        # "rating financial impact"
        match = re.search(
            rf"({rating_words})\s+{impact_type}\s+impact",
            text
        )
        if match:
            found[f"impact_{impact_type}"] = RISK_RATING_MAP.get(match.group(1))
            continue

        # "impact - financial: rating" or "impact (financial): rating"
        match = re.search(
            rf"impact\s*[\-\(]\s*{impact_type}\s*[\)\:]?\s*(?:is\s*)?({rating_words})",
            text
        )
        if match:
            found[f"impact_{impact_type}"] = RISK_RATING_MAP.get(match.group(1))

    return found


# =============================================================================
# SECTION 4: MAPPING ENGINE
# =============================================================================

def score_rationale_against_keywords(rationale: str, target_l2s: list[str]) -> dict:
    """Score rationale text against keyword sets for candidate L2s.

    Returns dict of {l2_name: score}. Higher = better match.
    """
    if not rationale or pd.isna(rationale):
        return {l2: 0 for l2 in target_l2s}

    text = str(rationale).lower()
    scores = {}
    for l2 in target_l2s:
        keywords = KEYWORD_MAP.get(l2, [])
        score = 0
        for kw in keywords:
            if kw in text:
                score += 1
        scores[l2] = score
    return scores


def resolve_via_sub_risks(
    entity_id: str,
    legacy_pillar: str,
    target_l2s: list[str],
    sub_risk_index: dict,
) -> tuple[str | None, str, str, list[str]]:
    """Tier 0: Resolve split mapping using sub-risk descriptions.

    Scores each sub-risk description against keyword sets for candidate L2s.
    A single sub-risk can evidence multiple L2s simultaneously.

    Returns (selected_l2, confidence, method, matched_descriptions).
    Returns (None, ...) if sub-risks don't resolve it.
    """
    entity_subs = sub_risk_index.get(entity_id, {})
    descriptions = entity_subs.get(legacy_pillar, [])

    if not descriptions:
        return None, "", "", []

    # Score each description against all candidate L2s
    l2_scores = {l2: 0 for l2 in target_l2s}
    l2_evidence = {l2: [] for l2 in target_l2s}

    for desc in descriptions:
        desc_lower = desc.lower()
        for l2 in target_l2s:
            keywords = KEYWORD_MAP.get(l2, [])
            hits = sum(1 for kw in keywords if kw in desc_lower)
            if hits > 0:
                l2_scores[l2] += hits
                l2_evidence[l2].append(desc[:100])  # truncate for logging

    max_score = max(l2_scores.values())
    if max_score == 0:
        return None, "", "", []

    top_l2s = [l2 for l2, s in l2_scores.items() if s == max_score]
    matched_descs = []
    for l2 in top_l2s:
        matched_descs.extend(l2_evidence[l2])

    if len(top_l2s) == 1:
        confidence = "high" if max_score >= 3 else "medium"
        return top_l2s[0], confidence, "sub_risk_lookup", matched_descs
    else:
        # Multiple L2s scored — pick highest, but this is actually useful info:
        # it means the entity has sub-risks spanning multiple L2s.
        # Return the top one; the others will still get gap-filled or caught
        # by the rationale keyword fallback for other legacy pillars.
        return top_l2s[0], "medium", "sub_risk_tie", matched_descs


def resolve_split_mapping(
    legacy_pillar: str,
    rationale: str,
    target_l2s: list[str],
    entity_id: str = None,
    sub_risk_index: dict = None,
    overrides: dict = None,
) -> tuple[str, str, str, list[str]]:
    """Resolve a 1:many mapping. Checks Override first, then Tier 0 (sub-risks),
    then falls back to Tier 1 (keyword matching on rationale).

    Returns (selected_l2, confidence, method, sub_risk_evidence).
    """
    # Override: LLM-classified result from previous run
    if overrides and entity_id:
        key = (entity_id, legacy_pillar)
        if key in overrides:
            ovr = overrides[key]
            return ovr["l2"], ovr["confidence"], "llm_override", []

    # Tier 0: Sub-risk descriptions
    if sub_risk_index and entity_id:
        result_l2, conf, method, evidence = resolve_via_sub_risks(
            entity_id, legacy_pillar, target_l2s, sub_risk_index
        )
        if result_l2:
            return result_l2, conf, method, evidence

    # Tier 1: Keyword matching on rationale text
    scores = score_rationale_against_keywords(rationale, target_l2s)
    max_score = max(scores.values()) if scores else 0

    if max_score == 0:
        return target_l2s[0], "low", "default_first_l2", []

    top_l2s = [l2 for l2, s in scores.items() if s == max_score]
    if len(top_l2s) == 1:
        confidence = "high" if max_score >= 3 else "medium"
        return top_l2s[0], confidence, "keyword_match", []
    else:
        return top_l2s[0], "low", "keyword_tie", []


def transform_entity(
    entity_id: str,
    entity_row: pd.Series,
    crosswalk: dict,
    pillar_columns: dict,
    sub_risk_index: dict = None,
    overrides: dict = None,
    findings_index: dict = None,
) -> tuple[list[dict], list[dict]]:
    """Transform one audit entity from legacy to new taxonomy.

    Args:
        entity_id: The audit entity identifier (e.g., "AE-7")
        entity_row: The full row from the legacy DataFrame
        crosswalk: The CROSSWALK_CONFIG dict
        pillar_columns: Dict mapping legacy pillar name to its column names
        sub_risk_index: Optional dict from build_sub_risk_index()
        overrides: Optional dict from load_overrides()
        findings_index: Optional dict from build_findings_index()

    Returns:
        (transformed_rows, overlay_flags)
    """
    transformed = []
    overlays = []
    mapped_l2s = set()

    # --- Pre-check: findings-confirmed L2s ---
    # If an entity has findings tagged to a new L2, that L2 is confirmed applicable.
    # Create placeholder rows with no ratings (ratings come from legacy pillar data).
    # These will be overridden by dedup if the crosswalk also produces rated rows.
    findings_confirmed_l2s = set()
    if findings_index:
        entity_findings = findings_index.get(entity_id, {})
        for l2, findings_list in entity_findings.items():
            if l2 in L2_TO_L1:
                findings_confirmed_l2s.add(l2)
                issue_summaries = [
                    f"{f['issue_id']}: {f['issue_title']} ({f['severity']}, {f['status']})"
                    for f in findings_list[:5]
                ]
                l1 = L2_TO_L1[l2]
                mapped_l2s.add(l2)
                transformed.append(_make_row(
                    entity_id, l1, l2,
                    source_legacy_pillar="Findings",
                    mapping_type="findings",
                    confidence="high",
                    method="issue_confirmed",
                    sub_risk_evidence="; ".join(issue_summaries),
                ))
                log.info(f"  Entity {entity_id}: '{l2}' confirmed applicable by {len(findings_list)} finding(s)")

    for legacy_pillar, config in crosswalk.items():
        # Get legacy data for this pillar
        cols = pillar_columns.get(legacy_pillar)
        if not cols:
            log.warning(f"  Entity {entity_id}: No columns found for '{legacy_pillar}'")
            continue

        rating_raw = entity_row.get(cols.get("rating"))
        rationale = entity_row.get(cols.get("rationale"), "")
        control_raw = entity_row.get(cols.get("control"))
        control_rationale = entity_row.get(cols.get("control_rationale"), "")

        rating_numeric = convert_risk_rating(rating_raw)
        control_numeric = convert_control_rating(control_raw)

        # Skip N/A ratings — flag all candidate L2s as not applicable
        raw_str = str(rating_raw).strip().lower() if rating_raw and not pd.isna(rating_raw) else ""
        is_na = (rating_numeric is None and raw_str in ("not applicable", "n/a", "na", ""))

        if is_na and config.get("mapping_type") != "overlay":
            # Determine which L2s this pillar would have mapped to
            mt = config.get("mapping_type", "")
            if mt in ("direct", "broader"):
                na_l2s = [config["target_l2"]]
            elif mt == "multi":
                na_l2s = [t["l2"] for t in config["targets"]]
            elif mt == "split":
                na_l2s = config.get("target_l2s", [])
            else:
                na_l2s = []

            for l2_name in na_l2s:
                l1 = L2_TO_L1.get(l2_name, "UNKNOWN")
                mapped_l2s.add(l2_name)
                transformed.append(_make_row(
                    entity_id, l1, l2_name,
                    source_legacy_pillar=legacy_pillar,
                    source_risk_rating_raw=rating_raw,
                    source_rationale=str(rationale) if rationale else "",
                    source_control_raw=control_raw,
                    mapping_type=config.get("mapping_type", ""),
                    confidence="n/a",
                    method="source_not_applicable",
                ))
            log.info(f"  Entity {entity_id}: '{legacy_pillar}' -> N/A, flagged {len(na_l2s)} L2s as not applicable")
            continue

        # Parse rationale for explicit dimension mentions
        parsed_dims = parse_rationale_for_dimensions(str(rationale))

        # Build the 5 risk dimension values
        likelihood = parsed_dims.get("likelihood", rating_numeric)
        impact_financial = parsed_dims.get("impact_financial", rating_numeric)
        impact_reputational = parsed_dims.get("impact_reputational", rating_numeric)
        impact_consumer = parsed_dims.get("impact_consumer", rating_numeric)
        impact_regulatory = parsed_dims.get("impact_regulatory", rating_numeric)
        # If generic "impact" was parsed, use it as default for all impact cols
        if "impact" in parsed_dims:
            generic_impact = parsed_dims["impact"]
            impact_financial = parsed_dims.get("impact_financial", generic_impact)
            impact_reputational = parsed_dims.get("impact_reputational", generic_impact)
            impact_consumer = parsed_dims.get("impact_consumer", generic_impact)
            impact_regulatory = parsed_dims.get("impact_regulatory", generic_impact)

        mapping_type = config["mapping_type"]

        if mapping_type == "overlay":
            for target_l2 in config["target_l2s"]:
                overlays.append({
                    "entity_id": entity_id,
                    "target_l2": target_l2,
                    "overlay_source": legacy_pillar,
                    "overlay_rating": rating_numeric,
                    "overlay_rationale": str(rationale),
                })
            log.info(f"  Entity {entity_id}: '{legacy_pillar}' -> overlay on {config['target_l2s']}")
            continue

        # Build list of target L2s to create rows for
        if mapping_type in ("direct", "broader"):
            targets_to_create = [{
                "l2": config["target_l2"],
                "confidence": "high",
                "method": mapping_type,
                "sub_risk_evidence": [],
            }]

        elif mapping_type == "multi":
            # Evidence-based multi-target: score each candidate L2 against
            # sub-risk descriptions and rationale keywords. Only populate
            # L2s that have actual evidence. If none match, default to
            # first primary and flag for review.
            targets_to_create = []

            # Build combined text from rationale + sub-risk descriptions
            entity_subs = (sub_risk_index or {}).get(entity_id, {})
            sub_descs = entity_subs.get(legacy_pillar, [])
            combined_text = " ".join(
                [str(rationale).lower()] + [d.lower() for d in sub_descs]
            )

            # Look up LLM override once for this entity+pillar
            override_entry = None
            if overrides and entity_id:
                override_entry = overrides.get((entity_id, legacy_pillar))

            first_primary_l2 = None
            for target in config["targets"]:
                if target["relationship"] == "primary" and not first_primary_l2:
                    first_primary_l2 = target["l2"]

                # Check LLM override for this target L2
                if override_entry and override_entry["l2"] == target["l2"]:
                    targets_to_create.append({
                        "l2": target["l2"],
                        "confidence": override_entry["confidence"],
                        "method": "llm_override",
                        "sub_risk_evidence": [],
                    })
                    continue

                # Score this L2 against combined text using KEYWORD_MAP
                l2_name = target["l2"]
                keywords = KEYWORD_MAP.get(l2_name, [])
                hits = [kw for kw in keywords if kw in combined_text]
                score = len(hits)

                # Also check conditional-specific keywords if present
                conditions = target.get("conditions", [])
                condition_hits = [c for c in conditions if c.lower() in combined_text]
                score += len(condition_hits)
                all_evidence = (hits + condition_hits)[:8]

                rel = target["relationship"]

                if score > 0:
                    # Evidence found — populate this L2
                    if score >= 3:
                        confidence = "high"
                    elif score >= 1:
                        confidence = "medium"
                    method = f"evidence_match ({rel})"
                    targets_to_create.append({
                        "l2": l2_name,
                        "confidence": confidence,
                        "method": method,
                        "sub_risk_evidence": all_evidence,
                    })

            # If no L2s had evidence, default to first primary and flag
            if not targets_to_create:
                if first_primary_l2:
                    targets_to_create.append({
                        "l2": first_primary_l2,
                        "confidence": "low",
                        "method": "default_no_evidence",
                        "sub_risk_evidence": [],
                    })
                    log.info(
                        f"  Entity {entity_id}: '{legacy_pillar}' -> no evidence for any L2, "
                        f"defaulted to '{first_primary_l2}' — FLAGGED FOR REVIEW"
                    )
                else:
                    log.warning(
                        f"  Entity {entity_id}: '{legacy_pillar}' multi mapping "
                        f"produced no targets and has no primary fallback"
                    )
                    continue

            # Track candidate L2s that were evaluated but had no evidence
            matched_l2s_this_pillar = {t["l2"] for t in targets_to_create}
            for target in config["targets"]:
                candidate_l2 = target["l2"]
                if candidate_l2 not in matched_l2s_this_pillar:
                    # This L2 was evaluated and didn't match — create explicit blank row
                    l1 = L2_TO_L1.get(candidate_l2, "UNKNOWN")
                    mapped_l2s.add(candidate_l2)
                    transformed.append(_make_row(
                        entity_id, l1, candidate_l2,
                        source_legacy_pillar=legacy_pillar,
                        source_risk_rating_raw=rating_raw,
                        source_rationale=str(rationale) if rationale else "",
                        source_control_raw=control_raw,
                        mapping_type=mapping_type,
                        confidence="none",
                        method="evaluated_no_evidence",
                    ))

        elif mapping_type == "split":
            selected_l2, confidence, method, sub_risk_evidence = resolve_split_mapping(
                legacy_pillar, str(rationale), config["target_l2s"],
                entity_id=entity_id, sub_risk_index=sub_risk_index,
                overrides=overrides,
            )
            targets_to_create = [{
                "l2": selected_l2,
                "confidence": confidence,
                "method": method,
                "sub_risk_evidence": sub_risk_evidence,
            }]
        else:
            log.error(f"  Unknown mapping_type '{mapping_type}' for '{legacy_pillar}'")
            continue

        dims_were_parsed = bool(parsed_dims)

        for tgt in targets_to_create:
            selected_l2 = tgt["l2"]
            l1 = L2_TO_L1.get(selected_l2, "UNKNOWN")
            mapped_l2s.add(selected_l2)

            row = {
                "entity_id": entity_id,
                "new_l1": l1,
                "new_l2": selected_l2,
                "composite_key": f"{selected_l2} {entity_id}",
                "likelihood": likelihood,
                "impact_financial": impact_financial,
                "impact_reputational": impact_reputational,
                "impact_consumer_harm": impact_consumer,
                "impact_regulatory": impact_regulatory,
                "iag_control_effectiveness": control_numeric,
                "aligned_assurance_rating": control_numeric,
                "management_awareness_rating": control_numeric,
                "source_legacy_pillar": legacy_pillar,
                "source_risk_rating_raw": rating_raw,
                "source_rationale": str(rationale) if rationale else "",
                "source_control_raw": control_raw,
                "source_control_rationale": str(control_rationale) if control_rationale else "",
                "mapping_type": mapping_type,
                "confidence": tgt["confidence"],
                "method": tgt["method"],
                "dims_parsed_from_rationale": dims_were_parsed,
                "sub_risk_evidence": "; ".join(tgt["sub_risk_evidence"]) if tgt["sub_risk_evidence"] else "",
                "needs_review": tgt["confidence"] == "low",
            }
            transformed.append(row)
            log.info(
                f"  Entity {entity_id}: '{legacy_pillar}' -> {l1} / {selected_l2} "
                f"[{tgt['method']}, conf={tgt['confidence']}]"
            )

    # --- Deduplicate: when multiple old L1s map to the same new L2 ---
    # Rules:
    #   - If one row has ratings and the other doesn't, keep the one WITH ratings
    #     but append issue info from findings rows to sub_risk_evidence
    #   - If both have ratings, keep the higher (more conservative) rating
    #   - If issue_confirmed vs evaluated_no_evidence/gap_fill, keep issue_confirmed
    seen = {}
    deduped = []
    for row in transformed:
        l2 = row["new_l2"]
        if l2 not in seen:
            seen[l2] = len(deduped)
            deduped.append(row)
        else:
            existing = deduped[seen[l2]]
            existing_rating = existing.get("likelihood") or 0
            new_rating = row.get("likelihood") or 0
            existing_method = existing.get("method", "")
            new_method = row.get("method", "")

            # Priority methods (override blanks)
            blank_methods = ("evaluated_no_evidence", "gap_fill", "true_gap_fill")

            if new_method == "issue_confirmed" and existing_method in blank_methods:
                # Findings confirmation overrides blank rows
                deduped[seen[l2]] = row
            elif existing_method == "issue_confirmed" and new_method in blank_methods:
                # Keep existing findings confirmation
                pass
            elif existing_method == "issue_confirmed" and new_rating > 0:
                # Existing is findings-only (no ratings), new has ratings — keep new but append issue info
                row["sub_risk_evidence"] = (
                    (row.get("sub_risk_evidence", "") + " | " + existing.get("sub_risk_evidence", "")).strip(" | ")
                )
                row["source_legacy_pillar"] = (
                    f"{row['source_legacy_pillar']} (also: Findings)"
                )
                deduped[seen[l2]] = row
            elif new_method == "issue_confirmed" and existing_rating > 0:
                # New is findings-only, existing has ratings — keep existing but append issue info
                existing["sub_risk_evidence"] = (
                    (existing.get("sub_risk_evidence", "") + " | " + row.get("sub_risk_evidence", "")).strip(" | ")
                )
                existing["source_legacy_pillar"] = (
                    f"{existing['source_legacy_pillar']} (also: Findings)"
                )
            elif new_rating > existing_rating:
                row["source_legacy_pillar"] = (
                    f"{row['source_legacy_pillar']} (also: {existing['source_legacy_pillar']})"
                )
                row["method"] = f"{row['method']} (dedup: kept higher)"
                deduped[seen[l2]] = row
            else:
                existing["source_legacy_pillar"] = (
                    f"{existing['source_legacy_pillar']} (also: {row['source_legacy_pillar']})"
                )
                if "dedup" not in existing_method:
                    existing["method"] = f"{existing_method} (dedup: kept higher)"

            log.info(
                f"  Entity {entity_id}: DEDUP '{l2}' — "
                f"'{row.get('source_legacy_pillar', '')}' [{new_method}] vs "
                f"existing [{existing_method}]"
            )
    transformed = deduped

    # Identify any new L2 risks with NO legacy mapping at all (true gaps)
    # With the current crosswalk this should be zero.
    for l2 in ALL_L2_RISKS:
        if l2 not in mapped_l2s:
            l1 = L2_TO_L1[l2]
            transformed.append(_make_row(
                entity_id, l1, l2,
                mapping_type="no_legacy_source",
                confidence="none",
                method="true_gap_fill",
            ))

    return transformed, overlays


# =============================================================================
# SECTION 5: PIPELINE & EXPORT
# =============================================================================

def run_pipeline(
    legacy_df: pd.DataFrame,
    crosswalk: dict,
    entity_id_col: str,
    pillar_columns: dict,
    sub_risk_index: dict = None,
    overrides: dict = None,
    findings_index: dict = None,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Run the full transformation pipeline across all entities."""
    all_transformed = []
    all_overlays = []

    for idx, row in legacy_df.iterrows():
        entity_id = str(row[entity_id_col]).strip()
        log.info(f"Processing entity {entity_id} ({idx + 1}/{len(legacy_df)})")
        transformed, overlays = transform_entity(
            entity_id, row, crosswalk, pillar_columns,
            sub_risk_index=sub_risk_index,
            overrides=overrides,
            findings_index=findings_index,
        )
        all_transformed.extend(transformed)
        all_overlays.extend(overlays)

    transformed_df = pd.DataFrame(all_transformed)
    overlays_df = pd.DataFrame(all_overlays) if all_overlays else pd.DataFrame()

    total = len(transformed_df)
    needs_review = transformed_df["needs_review"].sum()
    high_conf = (transformed_df["confidence"] == "high").sum()
    med_conf = (transformed_df["confidence"] == "medium").sum()
    low_conf = (transformed_df["confidence"] == "low").sum()

    log.info("=" * 60)
    log.info("TRANSFORMATION COMPLETE")
    log.info(f"  Total rows: {total}")
    log.info(f"  High confidence: {high_conf} ({high_conf/total*100:.1f}%)")
    log.info(f"  Medium confidence: {med_conf} ({med_conf/total*100:.1f}%)")
    log.info(f"  Low confidence / needs review: {low_conf} ({low_conf/total*100:.1f}%)")
    na_rows = (transformed_df["method"] == "source_not_applicable").sum()
    log.info(f"  Source N/A (skipped): {na_rows}")
    eval_no_ev = (transformed_df["method"] == "evaluated_no_evidence").sum()
    log.info(f"  Evaluated no evidence: {eval_no_ev}")
    true_gaps = (transformed_df["method"] == "true_gap_fill").sum()
    log.info(f"  True gap fills (no legacy pillar maps): {true_gaps}")
    evidence_matches = transformed_df["method"].str.contains("evidence_match", na=False).sum()
    evidence_high = (transformed_df["method"].str.contains("evidence_match", na=False) & (transformed_df["confidence"] == "high")).sum()
    evidence_med = (transformed_df["method"].str.contains("evidence_match", na=False) & (transformed_df["confidence"] == "medium")).sum()
    defaults_no_evidence = (transformed_df["method"] == "default_no_evidence").sum()
    issue_confirmed = (transformed_df["method"] == "issue_confirmed").sum()
    llm_overrides = transformed_df["method"].str.contains("llm_override", na=False).sum()
    dedup_count = transformed_df["method"].str.contains("dedup", na=False).sum()
    dims_parsed = transformed_df["dims_parsed_from_rationale"].sum()
    log.info(f"  Evidence-based matches: {evidence_matches} (high: {evidence_high}, medium: {evidence_med})")
    log.info(f"  Issue-confirmed applicable: {issue_confirmed}")
    log.info(f"  Default no evidence (flagged for review): {defaults_no_evidence}")
    log.info(f"  Resolved via LLM overrides: {llm_overrides}")
    log.info(f"  Deduplicated (multiple sources -> same L2): {dedup_count}")
    log.info(f"  Dimensions parsed from rationale: {dims_parsed}")
    log.info(f"  Overlay flags: {len(overlays_df)}")
    log.info(f"  Flagged for review: {needs_review}")
    log.info("=" * 60)

    return transformed_df, overlays_df


def apply_overlay_flags(transformed_df: pd.DataFrame, overlays_df: pd.DataFrame) -> pd.DataFrame:
    """Join overlay flags onto the transformed data."""
    if overlays_df.empty:
        transformed_df["overlay_flag"] = False
        transformed_df["overlay_source"] = ""
        transformed_df["overlay_rating"] = None
        transformed_df["overlay_rationale"] = ""
        return transformed_df

    # Aggregate overlays per entity+L2
    overlay_agg = (
        overlays_df
        .groupby(["entity_id", "target_l2"])
        .agg({
            "overlay_source": lambda x: "; ".join(x),
            "overlay_rating": "max",
            "overlay_rationale": lambda x: " | ".join(x),
        })
        .reset_index()
    )

    merged = transformed_df.merge(
        overlay_agg,
        left_on=["entity_id", "new_l2"],
        right_on=["entity_id", "target_l2"],
        how="left",
        suffixes=("", "_overlay"),
    )
    merged["overlay_flag"] = merged["target_l2"].notna()
    merged.drop(columns=["target_l2"], errors="ignore", inplace=True)
    merged["overlay_source"] = merged["overlay_source"].fillna("")
    merged["overlay_rationale"] = merged["overlay_rationale"].fillna("")

    return merged


def flag_control_contradictions(transformed_df: pd.DataFrame, findings_index: dict) -> pd.DataFrame:
    """Flag rows where control rating contradicts open findings.

    Adds a 'control_flag' column with human-readable warnings.
    """
    if not findings_index:
        transformed_df["control_flag"] = ""
        return transformed_df

    flags = []
    for _, row in transformed_df.iterrows():
        eid = str(row.get("entity_id", ""))
        l2 = row.get("new_l2", "")
        control_eff = row.get("iag_control_effectiveness")

        entity_findings = findings_index.get(eid, {})
        l2_findings = entity_findings.get(l2, [])

        # Only open/in-progress findings
        open_findings = [
            f for f in l2_findings
            if str(f.get("status", "")).strip().lower() not in ("closed", "")
        ]

        if not open_findings or control_eff is None:
            flags.append("")
            continue

        flag_parts = []
        for f in open_findings[:3]:
            sev = f.get("severity", "")
            iid = f.get("issue_id", "")
            title = f.get("issue_title", "")[:80]

            if control_eff == 1:
                flag_parts.append(
                    f"Open {sev} issue ({iid}: {title}) — "
                    f"review whether Well Controlled rating reflects current state"
                )
            elif control_eff == 2 and str(sev).strip().lower() == "high":
                flag_parts.append(
                    f"Open High issue ({iid}: {title}) — "
                    f"consider whether Moderately Controlled rating is appropriate"
                )

        flags.append(" | ".join(flag_parts))

    transformed_df["control_flag"] = flags
    return transformed_df


def _derive_decision_basis(row) -> str:
    """Plain-language explanation of mapping method for a transformed row."""
    method = str(row.get("method", ""))
    pillar = str(row.get("source_legacy_pillar", ""))
    evidence = str(row.get("sub_risk_evidence", ""))

    if method == "direct":
        return f"Direct mapping from {pillar} pillar"
    if method == "issue_confirmed":
        return f"Confirmed applicable — open finding: {evidence}"
    if "evidence_match" in method:
        if evidence:
            return f"Keywords matched in rationale/sub-risks: {evidence}"
        return f"Keywords matched in rationale from {pillar} pillar"
    if method == "llm_override":
        return "Classified by AI review of rationale and sub-risk descriptions"
    if method == "source_not_applicable":
        return f"Legacy {pillar} pillar rated Not Applicable"
    if method == "evaluated_no_evidence":
        return f"Evaluated from {pillar} pillar — no matching keywords found"
    if method == "default_no_evidence":
        return f"Could not determine applicable L2 from {pillar} pillar — team review required"
    if method in ("true_gap_fill", "gap_fill"):
        return "No legacy pillar maps to this L2"
    if "dedup" in method:
        return f"Mapped from {pillar} (multiple legacy sources)"
    return method


def build_audit_review_df(transformed_df: pd.DataFrame) -> pd.DataFrame:
    """Build the auditor-facing Audit Review dataframe with plain-language columns."""
    df = transformed_df.copy()

    # Status column
    def derive_status(method):
        method = str(method)
        if method in ("direct", "broader") or "evidence_match" in method or method == "llm_override" or method == "issue_confirmed":
            return "Applicable"
        if method in ("source_not_applicable", "evaluated_no_evidence"):
            return "Not Applicable"
        if method == "default_no_evidence":
            return "Needs Review"
        if method in ("true_gap_fill", "gap_fill"):
            return "Not Assessed"
        if "dedup" in method:
            return "Applicable"
        return "Needs Review"

    df["Status"] = df["method"].apply(derive_status)
    df["Decision Basis"] = df.apply(_derive_decision_basis, axis=1)

    # Rating Source column
    def derive_rating_source(row):
        has_ratings = row.get("likelihood") is not None and not pd.isna(row.get("likelihood", None))
        if not has_ratings:
            return "No ratings — legacy source was N/A or not assessed"

        parts = []
        if row.get("dims_parsed_from_rationale"):
            parts.append("Inherent Risk: Parsed from rationale — likelihood and impact stated separately")
        else:
            raw = row.get("source_risk_rating_raw", "")
            parts.append(f"Inherent Risk: Default from composite rating {raw}")

        ctrl = row.get("source_control_raw", "")
        if row.get("iag_control_effectiveness") is not None and not pd.isna(row.get("iag_control_effectiveness")):
            parts.append(f"Control: All 3 set from {ctrl}")
        else:
            parts.append("Control: Not assessed")

        return " | ".join(parts)

    df["Rating Source"] = df.apply(derive_rating_source, axis=1)

    # Select and rename columns
    audit_cols = {
        "entity_id": "Entity ID",
        "new_l1": "New L1",
        "new_l2": "New L2",
        "Status": "Status",
        "Decision Basis": "Decision Basis",
        "Rating Source": "Rating Source",
        "likelihood": "Likelihood",
        "impact_financial": "Impact - Financial",
        "impact_reputational": "Impact - Reputational",
        "impact_consumer_harm": "Impact - Consumer Harm",
        "impact_regulatory": "Impact - Regulatory",
        "iag_control_effectiveness": "IAG Control Effectiveness",
        "aligned_assurance_rating": "Aligned Assurance Rating",
        "management_awareness_rating": "Management Awareness Rating",
        "control_flag": "Control Flag",
        "source_legacy_pillar": "Legacy Source",
        "confidence": "Confidence",
    }

    available = {k: v for k, v in audit_cols.items() if k in df.columns}
    result = df[list(available.keys())].copy()
    result.columns = list(available.values())

    # Sort: Needs Review first, then Applicable, then Not Applicable, then Not Assessed
    status_order = {"Needs Review": 0, "Applicable": 1, "Not Applicable": 2, "Not Assessed": 3}
    result["_sort"] = result["Status"].map(status_order).fillna(4)
    result = result.sort_values(["Entity ID", "_sort"]).drop(columns=["_sort"])

    return result


def build_review_queue_df(transformed_df: pd.DataFrame) -> pd.DataFrame:
    """Build redesigned Review Queue including both defaults and evaluated-no-evidence."""
    mask = transformed_df["method"].isin(["default_no_evidence", "evaluated_no_evidence"])
    df = transformed_df[mask].copy()

    if df.empty:
        return df

    # Review Type column
    def derive_review_type(method):
        if method == "default_no_evidence":
            return "Determine Applicability — could not match to specific L2"
        if method == "evaluated_no_evidence":
            return "Confirm Not Applicable — evaluated, no evidence found"
        return ""

    df["Review Type"] = df["method"].apply(derive_review_type)

    df["Decision Basis"] = df.apply(_derive_decision_basis, axis=1)

    review_cols = [
        "entity_id", "Review Type", "new_l1", "new_l2", "Decision Basis",
        "source_legacy_pillar", "source_risk_rating_raw", "source_rationale",
        "sub_risk_evidence",
    ]
    available = [c for c in review_cols if c in df.columns]
    result = df[available].copy()

    col_rename = {
        "entity_id": "Entity ID", "new_l1": "New L1", "new_l2": "New L2",
        "source_legacy_pillar": "Legacy Source", "source_risk_rating_raw": "Source Rating",
        "source_rationale": "Source Rationale", "sub_risk_evidence": "Sub-Risk Evidence",
    }
    result = result.rename(columns=col_rename)

    # Sort: Determine Applicability first, then Confirm Not Applicable
    result = result.sort_values(["Review Type", "Entity ID"])

    return result


def style_header(ws, max_col: int):
    """Apply header styling to first row."""
    header_font = Font(bold=True, color="FFFFFF", size=10, name="Arial")
    header_fill = PatternFill("solid", fgColor="2F5496")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    for col_idx in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border


def export_results(
    transformed_df: pd.DataFrame,
    overlays_df: pd.DataFrame,
    legacy_df: pd.DataFrame,
    entity_id_col: str,
    output_path: str,
):
    """Write multi-sheet Excel output."""
    log.info(f"Writing output to {output_path}")

    # --- Sheet 1: Transformed Upload ---
    upload_cols = [
        "composite_key", "entity_id", "new_l1", "new_l2",
        "likelihood", "impact_financial", "impact_reputational",
        "impact_consumer_harm", "impact_regulatory",
        "iag_control_effectiveness", "aligned_assurance_rating",
        "management_awareness_rating",
    ]
    upload_df = transformed_df[upload_cols].copy()
    upload_df.columns = [
        "Risk-Entity Key", "Entity ID", "L1 Risk Pillar", "L2 Risk",
        "Likelihood", "Impact - Financial", "Impact - Reputational",
        "Impact - Consumer Harm", "Impact - Regulatory",
        "IAG Control Effectiveness", "Aligned Assurance Rating",
        "Management Awareness Rating",
    ]

    # --- Sheet 2: Audit Review (auditor-facing) ---
    audit_df = build_audit_review_df(transformed_df)

    # --- Sheet 3: Review Queue (redesigned) ---
    review_df = build_review_queue_df(transformed_df)

    # --- Sheet 4: Side-by-side (debugging) ---
    trace_cols = [
        "composite_key", "entity_id", "new_l1", "new_l2",
        "likelihood", "impact_financial", "impact_reputational",
        "impact_consumer_harm", "impact_regulatory",
        "iag_control_effectiveness", "aligned_assurance_rating",
        "management_awareness_rating",
        "source_legacy_pillar", "source_risk_rating_raw", "source_rationale",
        "source_control_raw", "source_control_rationale",
        "mapping_type", "confidence", "method",
        "dims_parsed_from_rationale", "sub_risk_evidence", "needs_review",
        "control_flag",
        "overlay_flag", "overlay_source", "overlay_rating", "overlay_rationale",
    ]
    available_trace_cols = [c for c in trace_cols if c in transformed_df.columns]
    trace_df = transformed_df[available_trace_cols].copy()

    # --- Sheet 5: Legacy original ---
    legacy_out = legacy_df.copy()

    # --- Sheet 6: Overlay flags ---
    overlay_out = overlays_df.copy() if not overlays_df.empty else pd.DataFrame()

    # Write sheets
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        upload_df.to_excel(writer, sheet_name="Transformed_Upload", index=False)
        audit_df.to_excel(writer, sheet_name="Audit_Review", index=False)
        review_df.to_excel(writer, sheet_name="Review_Queue", index=False)
        trace_df.to_excel(writer, sheet_name="Side_by_Side", index=False)
        legacy_out.to_excel(writer, sheet_name="Legacy_Original", index=False)
        if not overlay_out.empty:
            overlay_out.to_excel(writer, sheet_name="Overlay_Flags", index=False)

    # Apply formatting
    wb = load_workbook(output_path)

    # Status color fills
    green_fill = PatternFill("solid", fgColor="C6EFCE")
    gray_fill = PatternFill("solid", fgColor="D9D9D9")
    yellow_fill = PatternFill("solid", fgColor="FFFF00")
    blue_fill = PatternFill("solid", fgColor="BDD7EE")
    orange_fill = PatternFill("solid", fgColor="FCE4D6")
    light_red = PatternFill("solid", fgColor="FFC7CE")

    status_fills = {
        "Applicable": green_fill,
        "Not Applicable": gray_fill,
        "Needs Review": yellow_fill,
        "Not Assessed": blue_fill,
    }

    review_type_fills = {
        "Determine Applicability": yellow_fill,
        "Confirm Not Applicable": orange_fill,
    }

    for ws_name in wb.sheetnames:
        ws = wb[ws_name]
        style_header(ws, ws.max_column)

        # Auto-width columns
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                except (TypeError, ValueError):
                    pass
            # Cap wider for text-heavy columns
            cap = 60 if ws_name in ("Review_Queue", "Audit_Review") else 40
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), cap)

        # Color-code Audit_Review by Status
        if ws_name == "Audit_Review":
            status_col = None
            for cell in ws[1]:
                if cell.value == "Status":
                    status_col = cell.column
                    break
            if status_col:
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    status_val = row[status_col - 1].value
                    fill = status_fills.get(status_val)
                    if fill:
                        for cell in row:
                            cell.fill = fill

        # Color-code Review_Queue by Review Type
        if ws_name == "Review_Queue":
            rt_col = None
            for cell in ws[1]:
                if cell.value == "Review Type":
                    rt_col = cell.column
                    break
            if rt_col:
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    rt_val = str(row[rt_col - 1].value or "")
                    for key, fill in review_type_fills.items():
                        if key in rt_val:
                            for cell in row:
                                cell.fill = fill
                            break

        # Highlight needs_review rows in yellow on Side_by_Side
        if ws_name == "Side_by_Side":
            review_col = None
            for cell in ws[1]:
                if cell.value == "needs_review":
                    review_col = cell.column
                    break
            if review_col:
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    if row[review_col - 1].value is True:
                        for cell in row:
                            cell.fill = yellow_fill

    wb.save(output_path)
    log.info(f"  Output saved: {output_path}")
    log.info(f"  Sheets: {wb.sheetnames}")


# =============================================================================
# SECTION 6: MAIN — CONFIGURE AND RUN
# =============================================================================

def main():
    # -------------------------------------------------------------------------
    # CONFIGURE THESE PATHS AND COLUMN NAMES
    # -------------------------------------------------------------------------
    LEGACY_DATA_PATH = "legacy_risk_data.xlsx"       # Your entity data file
    CROSSWALK_PATH = None                             # Set path or None to use hardcoded
    OUTPUT_PATH = "transformed_risk_taxonomy.xlsx"
    ENTITY_ID_COL = "Entity ID"                       # Column name for audit entity ID

    # Sub-risk descriptions file (optional but recommended for split accuracy)
    # Set to None to skip sub-risk lookup and use keyword matching only.
    SUB_RISK_PATH = "sub_risk_descriptions.xlsx"      # Your sub-risk file
    SUB_RISK_COLS = {
        "entity_id": "Audit Entity ID",     # Column: audit entity identifier
        "risk_id": "Risk ID",               # Column: risk ID (optional, for traceability)
        "risk_desc": "Risk Description",    # Column: free-text risk description
        "legacy_l1": "L1 Risk Pillars",     # Column: legacy L1 pillar(s), tab-separated
        "rating": "Inherent Risk Rating",   # Column: rating (optional, not used for scoring)
    }

    # LLM Override file (optional — produced by batching Review Queue through LLM)
    # Set to None on first run. After LLM classification, point to the output file.
    # Expected columns: entity_id, source_legacy_pillar, classified_l2, llm_confidence
    OVERRIDE_PATH = None  # e.g., "llm_overrides.xlsx" or "llm_overrides.csv"

    # Findings/Issues file (optional — confirms L2 applicability and flags control contradictions)
    # Set to None to skip findings integration.
    FINDINGS_PATH = None  # e.g., "findings_data.xlsx"
    FINDINGS_COLS = {
        "entity_id": "Entity ID",           # Column: audit entity identifier
        "issue_id": "Issue ID",             # Column: issue/finding ID
        "l2_risk": "L2 Risk",              # Column: new L2 risk name (must match NEW_TAXONOMY)
        "severity": "Severity",             # Column: High/Medium/Low
        "status": "Status",                 # Column: Open/Closed/In Progress
        "issue_title": "Issue Title",       # Column: description text
        "remediation_date": "Remediation Date",  # Column: target date
    }

    # Map each legacy pillar to its column names in the Excel file.
    # Keys MUST match CROSSWALK_CONFIG keys exactly.
    # Adjust column headers to match your actual file.
    PILLAR_COLUMNS = {
        "Credit": {
            "rating": "Credit - Inherent Risk Rating",
            "rationale": "Credit - Rationale",
            "control": "Credit - Control Assessment",
            "control_rationale": "Credit - Control Rationale",
        },
        "Market": {
            "rating": "Market - Inherent Risk Rating",
            "rationale": "Market - Rationale",
            "control": "Market - Control Assessment",
            "control_rationale": "Market - Control Rationale",
        },
        # ... REPEAT FOR ALL 14: Strategic & Business, Funding and Liquidity,
        #     Reputational, Model, Third Party, Financial Reporting,
        #     External Fraud, Information Technology, Information Security,
        #     Operational, Compliance, Country ...
        #
        # TIP: If your columns follow a consistent pattern like
        # "{Pillar} - Inherent Risk Rating", use the auto-generator below.
    }

    # -------------------------------------------------------------------------
    # HELPER: Auto-generate PILLAR_COLUMNS if naming is consistent
    # Uncomment and adjust the pattern if your columns follow a convention.
    # -------------------------------------------------------------------------
    # PILLAR_COLUMNS = {}
    # for pillar_name in CROSSWALK_CONFIG.keys():
    #     PILLAR_COLUMNS[pillar_name] = {
    #         "rating":            f"{pillar_name} - Inherent Risk Rating",
    #         "rationale":         f"{pillar_name} - Rationale",
    #         "control":           f"{pillar_name} - Control Assessment",
    #         "control_rationale": f"{pillar_name} - Control Rationale",
    #     }

    # -------------------------------------------------------------------------
    # RUN
    # -------------------------------------------------------------------------
    crosswalk = ingest_crosswalk(CROSSWALK_PATH)
    legacy_df = ingest_legacy_data(LEGACY_DATA_PATH)

    # Load sub-risk descriptions if configured
    sub_risk_index = None
    if SUB_RISK_PATH:
        sub_risks_df = ingest_sub_risks(
            SUB_RISK_PATH,
            entity_id_col=SUB_RISK_COLS["entity_id"],
            legacy_l1_col=SUB_RISK_COLS["legacy_l1"],
            risk_desc_col=SUB_RISK_COLS["risk_desc"],
            risk_id_col=SUB_RISK_COLS.get("risk_id"),
            rating_col=SUB_RISK_COLS.get("rating"),
        )
        sub_risk_index = build_sub_risk_index(sub_risks_df)
        log.info(f"  Sub-risk index built: {len(sub_risk_index)} entities with sub-risks")

    # Load LLM overrides if configured
    overrides = None
    if OVERRIDE_PATH:
        overrides = load_overrides(OVERRIDE_PATH)
        log.info(f"  Override index built: {len(overrides)} entity-pillar overrides")

    # Load findings if configured
    findings_index = None
    if FINDINGS_PATH:
        findings_df = ingest_findings(FINDINGS_PATH, FINDINGS_COLS)
        findings_index = build_findings_index(findings_df)

    transformed_df, overlays_df = run_pipeline(
        legacy_df, crosswalk, ENTITY_ID_COL, PILLAR_COLUMNS,
        sub_risk_index=sub_risk_index,
        overrides=overrides,
        findings_index=findings_index,
    )

    transformed_df = apply_overlay_flags(transformed_df, overlays_df)
    transformed_df = flag_control_contradictions(transformed_df, findings_index)

    export_results(
        transformed_df, overlays_df, legacy_df,
        ENTITY_ID_COL, OUTPUT_PATH
    )

    print(f"\nDone! Output: {OUTPUT_PATH}")
    print(f"Review queue: {transformed_df['needs_review'].sum()} items flagged")


if __name__ == "__main__":
    main()
